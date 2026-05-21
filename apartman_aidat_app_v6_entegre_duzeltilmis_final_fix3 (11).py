# -*- coding: utf-8 -*-
"""
Apartman Aidat Sistemi (Mini v4) + Devir (Sakin Geçmişi) - UI Revamp

Masaüstü PySide6 uygulaması.

Özellikler:
- Daire ekle/güncelle/aktif-pasif
- Sakin geçmişi: Devir (taşınma) + geçmiş liste
- Tahakkuk (dönem borç yaz) + manuel geçmiş borç
- Excel'den borç içe aktar (daire yoksa otomatik oluşturur)
- Ödeme al (Banka/Elden) + peşin çok ay paylaştırma + Makbuz PDF
- Gider ekle/sil
- Duyuru ekle/sil + panoya kopyala
- Rapor: gelir/gider/net + borçlu/gecikmiş filtre + Excel export
- WhatsApp: güncel sakinin telefonuna wa.me linki açar

Not: WhatsApp otomatik göndermez, sadece tarayıcıda mesaj taslağı açar.
"""

import sys
import sqlite3
from pathlib import Path
from datetime import date, timedelta
import webbrowser
from urllib.parse import quote


from PySide6.QtCore import Qt, QDate
from PySide6.QtWidgets import (
    QApplication, QWidget, QTabWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QMessageBox, QComboBox, QGroupBox, QFileDialog, QSpinBox, QCheckBox,
    QTextEdit, QDateEdit, QSplitter, QHeaderView, QScrollArea, QDialog  # ⭐ QDialog ekle
)
from PySide6.QtGui import QColor

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


DB_PATH = Path("apartman_aidat.db")


# ----------------- helpers -----------------
def connect():
    con = sqlite3.connect(str(DB_PATH))
    con.execute("PRAGMA foreign_keys = ON;")
    return con


def iso_today() -> str:
    return date.today().isoformat()


def ym_of_today() -> str:
    d = date.today()
    return f"{d.year:04d}-{d.month:02d}"


def safe_float(s: str) -> float:
    t = (s or "").strip()
    if not t:
        return 0.0
    t = t.replace(" ", "")
    # TR: 1.234,56 -> 1234.56
    t = t.replace(".", "")
    t = t.replace(",", ".")
    return float(t)


def validate_period(donem: str) -> bool:
    if not donem or len(donem) != 7 or donem[4] != "-":
        return False
    try:
        y = int(donem[:4])
        m = int(donem[5:])
        return 2000 <= y <= 2100 and 1 <= m <= 12
    except Exception:
        return False


def add_months(ym: str, add: int) -> str:
    y = int(ym[:4])
    m = int(ym[5:])
    m2 = m + add
    y += (m2 - 1) // 12
    m2 = ((m2 - 1) % 12) + 1
    return f"{y:04d}-{m2:02d}"


def months_range(ym_start: str, ym_end: str):
    ys = int(ym_start[:4]); ms = int(ym_start[5:])
    ye = int(ym_end[:4]); me = int(ym_end[5:])
    cur_y, cur_m = ys, ms
    out = []
    while (cur_y, cur_m) <= (ye, me):
        out.append(f"{cur_y:04d}-{cur_m:02d}")
        cur_m += 1
        if cur_m == 13:
            cur_m = 1
            cur_y += 1
    return out


def ensure_db():
    con = connect()
    con.executescript("""
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS daireler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        daire_no TEXT NOT NULL UNIQUE,
        ad_soyad TEXT NOT NULL,
        telefon TEXT DEFAULT '',
        aidat REAL NOT NULL DEFAULT 0,
        aktif INTEGER NOT NULL DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS odeme_detay (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    odeme_id INTEGER NOT NULL,
    donem TEXT NOT NULL,
    tutar REAL NOT NULL,
    FOREIGN KEY(odeme_id) REFERENCES odemeler(id)
    );
    CREATE TABLE IF NOT EXISTS sakinler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        daire_id INTEGER NOT NULL,
        ad_soyad TEXT NOT NULL,
        telefon TEXT DEFAULT '',
        baslangic_tarihi TEXT NOT NULL,
        bitis_tarihi TEXT,
        FOREIGN KEY(daire_id) REFERENCES daireler(id)
    );
    CREATE INDEX IF NOT EXISTS idx_sakinler_daire ON sakinler(daire_id);
    CREATE INDEX IF NOT EXISTS idx_sakinler_aktif ON sakinler(daire_id, bitis_tarihi);

    CREATE TABLE IF NOT EXISTS tahakkuk (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        donem TEXT NOT NULL,
        daire_id INTEGER NOT NULL,
        tutar REAL NOT NULL,
        created_at TEXT NOT NULL,
        UNIQUE(donem, daire_id),
        FOREIGN KEY(daire_id) REFERENCES daireler(id)
    );

    CREATE TABLE IF NOT EXISTS odemeler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        donem TEXT NOT NULL,
        tarih TEXT NOT NULL,
        daire_id INTEGER NOT NULL,
        tutar REAL NOT NULL,
        yontem TEXT NOT NULL,
        makbuz_no TEXT NOT NULL,
        aciklama TEXT DEFAULT '',
        FOREIGN KEY(daire_id) REFERENCES daireler(id)
    );

    CREATE INDEX IF NOT EXISTS idx_odemeler_donem ON odemeler(donem);
    CREATE INDEX IF NOT EXISTS idx_odemeler_daire ON odemeler(daire_id);

    CREATE TABLE IF NOT EXISTS giderler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        donem TEXT NOT NULL,
        tarih TEXT NOT NULL,
        kategori TEXT NOT NULL,
        tutar REAL NOT NULL,
        yontem TEXT NOT NULL,
        aciklama TEXT DEFAULT ''
    );
    CREATE INDEX IF NOT EXISTS idx_giderler_donem ON giderler(donem);

    CREATE TABLE IF NOT EXISTS duyurular (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tarih TEXT NOT NULL,
        baslik TEXT NOT NULL,
        mesaj TEXT NOT NULL
    );
    """)
    con.commit()

    def set_default(k, v):
        if con.execute("SELECT value FROM settings WHERE key=?", (k,)).fetchone() is None:
            con.execute("INSERT INTO settings(key,value) VALUES(?,?)", (k, str(v)))

    set_default("vade_gun", 10)
    set_default("gecikme_gun", 5)
    con.commit()
    con.close()


def get_setting_int(key: str, default: int) -> int:
    con = connect()
    row = con.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    con.close()
    try:
        return int(row[0]) if row else default
    except Exception:
        return default


def set_setting_int(key: str, value: int):
    con = connect()
    con.execute(
        "INSERT INTO settings(key,value) VALUES(?,?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, str(int(value)))
    )
    con.commit()
    con.close()


def receipt_next_for_period(donem: str) -> str:
    yyyymm = donem.replace("-", "")
    con = connect()
    row = con.execute("""
        SELECT makbuz_no FROM odemeler
        WHERE donem=?
        ORDER BY id DESC
        LIMIT 1
    """, (donem,)).fetchone()
    con.close()
    if not row:
        return f"{yyyymm}-00001"
    last = row[0]
    try:
        seq = int(last.split("-")[1])
    except Exception:
        seq = 0
    return f"{yyyymm}-{seq+1:05d}"


def compute_vade_date(donem: str) -> date:
    vade_gun = get_setting_int("vade_gun", 10)
    y, m = donem.split("-")
    y = int(y); m = int(m)
    # last day
    if m == 12:
        next_month = date(y + 1, 1, 1)
    else:
        next_month = date(y, m + 1, 1)
    last_day = next_month - timedelta(days=1)
    day = min(vade_gun, last_day.day)
    return date(y, m, day)


def status_for_balance(donem: str, bakiye: float) -> str:
    if bakiye <= 0.00001:
        return "Ödedi"
    vade = compute_vade_date(donem)
    gecikme = get_setting_int("gecikme_gun", 5)
    limit = vade + timedelta(days=gecikme)
    if date.today() > limit:
        return "Gecikmiş"
    return "Borçlu"


def autosize_worksheet(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            v = str(v)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)
# ============ MODAL DIALOGS ============

class EditPaymentDialog(QDialog):
    """Ödeme düzenleme modalı"""
    def __init__(self, payment_id: int, parent=None):
        super().__init__(parent)
        self.payment_id = payment_id
        self.setWindowTitle("Ödemeyi Düzenle")
        self.setGeometry(100, 100, 500, 300)
        self.setModal(True)
        self.init_ui()
        self.load_payment()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        form = QFormLayout()
        
        self.in_donem = QLineEdit()
        self.in_donem.setReadOnly(True)
        self.dt_tarih = QDateEdit()
        self.dt_tarih.setCalendarPopup(True)
        self.dt_tarih.setDisplayFormat("dd-MM-yyyy")
        self.in_daire = QLineEdit()
        self.in_daire.setReadOnly(True)
        self.in_tutar = QLineEdit()
        self.cmb_yontem = QComboBox()
        self.cmb_yontem.addItems(["Banka", "Elden"])
        self.in_makbuz = QLineEdit()
        self.in_acik = QLineEdit()
        
        form.addRow("Dönem", self.in_donem)
        form.addRow("Tarih", self.dt_tarih)
        form.addRow("Daire", self.in_daire)
        form.addRow("Tutar (TL)", self.in_tutar)
        form.addRow("Yöntem", self.cmb_yontem)
        form.addRow("Makbuz No", self.in_makbuz)
        form.addRow("Açıklama", self.in_acik)
        
        layout.addLayout(form)
        
        btns = QHBoxLayout()
        self.btn_save = QPushButton("Güncelle")
        self.btn_save.clicked.connect(self.save_payment)
        self.btn_delete = QPushButton("Sil")
        self.btn_delete.clicked.connect(self.delete_payment)
        self.btn_close = QPushButton("Kapat")
        self.btn_close.clicked.connect(self.close)
        
        btns.addWidget(self.btn_save)
        btns.addWidget(self.btn_delete)
        btns.addStretch(1)
        btns.addWidget(self.btn_close)
        
        layout.addLayout(btns)
    
    def load_payment(self):
        """Ödeme verilerini yükle"""
        con = connect()
        row = con.execute("""
            SELECT o.donem, o.tarih, d.daire_no, o.tutar, o.yontem, o.makbuz_no, o.aciklama
            FROM odemeler o
            JOIN daireler d ON d.id = o.daire_id
            WHERE o.id = ?
        """, (self.payment_id,)).fetchone()
        con.close()
        
        if not row:
            QMessageBox.warning(self, "Hata", "Ödeme bulunamadı.")
            self.close()
            return
        
        donem, tarih, daire_no, tutar, yontem, makbuz_no, acik = row
        
        self.in_donem.setText(donem)
        self.dt_tarih.setDate(QDate.fromString(tarih, "yyyy-MM-dd"))
        self.in_daire.setText(daire_no)
        self.in_tutar.setText(f"{float(tutar):.2f}")
        self.cmb_yontem.setCurrentText(yontem)
        self.in_makbuz.setText(makbuz_no or "")
        self.in_acik.setText(acik or "")
    
    def save_payment(self):
        """Ödemeyi güncelle"""
        try:
            tutar = safe_float(self.in_tutar.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Tutar sayısal olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        tarih = self.dt_tarih.date().toPython().isoformat()
        yontem = self.cmb_yontem.currentText()
        acik = self.in_acik.text().strip()
        
        con = connect()
        con.execute("""
            UPDATE odemeler
            SET tarih=?, tutar=?, yontem=?, aciklama=?
            WHERE id=?
        """, (tarih, float(tutar), yontem, acik, self.payment_id))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Ödeme güncellendi.")
        self.accept()
    
    def delete_payment(self):
        """Ödemeyi sil"""
        reply = QMessageBox.question(
            self, "Onay", "Seçili ödemeyi silmek istiyor musunuz?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        
        con = connect()
        con.execute("DELETE FROM odemeler WHERE id=?", (self.payment_id,))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Ödeme silindi.")
        self.accept()



    # ... rest of the code ...
class EditExpenseDialog(QDialog):
    """Gider düzenleme modalı"""
    def __init__(self, expense_id: int, parent=None):
        super().__init__(parent)
        self.expense_id = expense_id
        self.setWindowTitle("Gideri Düzenle")
        self.setGeometry(100, 100, 500, 300)
        self.setModal(True)
        self.init_ui()
        self.load_expense()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        form = QFormLayout()
        
        self.in_donem = QLineEdit()
        self.in_donem.setReadOnly(True)
        
        self.dt_tarih = QDateEdit()
        self.dt_tarih.setCalendarPopup(True)
        self.dt_tarih.setDisplayFormat("dd-MM-yyyy")
        
        self.cmb_kat = QComboBox()
        self.cmb_kat.addItems(["Elektrik", "Temizlik", "Bakım", "Diğer"])
        
        self.in_tutar = QLineEdit()
        
        self.cmb_yontem = QComboBox()
        self.cmb_yontem.addItems(["Banka", "Elden"])
        
        self.in_acik = QLineEdit()
        
        form.addRow("Dönem", self.in_donem)
        form.addRow("Tarih", self.dt_tarih)
        form.addRow("Kategori", self.cmb_kat)
        form.addRow("Tutar (TL)", self.in_tutar)
        form.addRow("Yöntem", self.cmb_yontem)
        form.addRow("Açıklama", self.in_acik)
        
        layout.addLayout(form)
        
        btns = QHBoxLayout()
        self.btn_save = QPushButton("Güncelle")
        self.btn_save.clicked.connect(self.save_expense)
        self.btn_delete = QPushButton("Sil")
        self.btn_delete.clicked.connect(self.delete_expense)
        self.btn_close = QPushButton("Kapat")
        self.btn_close.clicked.connect(self.close)
        
        btns.addWidget(self.btn_save)
        btns.addWidget(self.btn_delete)
        btns.addStretch(1)
        btns.addWidget(self.btn_close)
        
        layout.addLayout(btns)
    
    def load_expense(self):
        """Gider verilerini yükle"""
        con = connect()
        row = con.execute("""
            SELECT donem, tarih, kategori, tutar, yontem, aciklama
            FROM giderler
            WHERE id = ?
        """, (self.expense_id,)).fetchone()
        con.close()
        
        if not row:
            QMessageBox.warning(self, "Hata", "Gider bulunamadı.")
            self.close()
            return
        
        donem, tarih, kat, tutar, yontem, acik = row
        
        self.in_donem.setText(donem)
        self.dt_tarih.setDate(QDate.fromString(tarih, "yyyy-MM-dd"))
        self.cmb_kat.setCurrentText(kat)
        self.in_tutar.setText(f"{float(tutar):.2f}")
        self.cmb_yontem.setCurrentText(yontem)
        self.in_acik.setText(acik or "")   
    
    def save_expense(self):
        """Gideri güncelle"""
        try:
            tutar = safe_float(self.in_tutar.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Tutar sayısal olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        tarih = self.dt_tarih.date().toPython().isoformat()
        kat = self.cmb_kat.currentText()
        yontem = self.cmb_yontem.currentText()
        acik = self.in_acik.text().strip()
        
        con = connect()
        con.execute("""
            UPDATE giderler
            SET tarih=?, kategori=?, tutar=?, yontem=?, aciklama=?
            WHERE id=?
        """, (tarih, kat, float(tutar), yontem, acik, self.expense_id))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Gider güncellendi.")
        self.accept()
    
    def delete_expense(self):
        """Gideri sil"""
        reply = QMessageBox.question(
            self, "Onay", "Seçili gideri silmek istiyor musunuz?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        
        con = connect()
        con.execute("DELETE FROM giderler WHERE id=?", (self.expense_id,))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Gider silindi.")
        self.accept()

# ----------------- App -----------------
class ApartmanAidatApp(QWidget):
    def __init__(self):
        super().__init__()
        self.selected_daire_id = None
        ensure_db()
        self.setWindowTitle("Apartman Aidat Sistemi (Mini v4 - Devir)")
        screen = QApplication.primaryScreen().availableGeometry()
        w = int(screen.width() * 0.85)
        h = int(screen.height() * 0.85)
        self.resize(w, h)
        self.setMinimumSize(min(1100, w), min(700, h))
        self._build_ui()
        self.refresh_all()
       
   
    
    # ---------------- UI ----------------
    def _build_ui(self):
    
        root = QVBoxLayout(self)
        
        self.kasa_label = QLabel()
        self.kasa_label.setStyleSheet("""
            font-size: 16px;
            font-weight: bold;
            padding: 8px;
        """)

       
        topbar = QHBoxLayout()
        root.addLayout(topbar)

        topbar.addWidget(QLabel("Dönem:"))
        self.in_donem = QLineEdit(ym_of_today())
        self.in_donem.setMaximumWidth(110)
        topbar.addWidget(self.in_donem)

        self.btn_refresh = QPushButton("Yenile")
        self.btn_refresh.clicked.connect(self.refresh_all)
        topbar.addWidget(self.btn_refresh)

        topbar.addStretch(1)

        topbar.addWidget(QLabel("Vade günü:"))
        self.sp_vade = QSpinBox()
        self.sp_vade.setRange(1, 28)
        self.sp_vade.setValue(get_setting_int("vade_gun", 10))
        topbar.addWidget(self.sp_vade)

        topbar.addWidget(QLabel("Gecikme + gün:"))
        self.sp_gec = QSpinBox()
        self.sp_gec.setRange(0, 60)
        self.sp_gec.setValue(get_setting_int("gecikme_gun", 5))
        topbar.addWidget(self.sp_gec)

        self.btn_save_settings = QPushButton("Ayarları Kaydet")
        self.btn_save_settings.clicked.connect(self.save_settings)
        topbar.addWidget(self.btn_save_settings)

        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)

        self._tab_daireler()
        self._tab_tahakkuk()
        self._tab_odeme()
        self._tab_gider()
        self._tab_duyuru()
        self._tab_rapor()
        self._tab_kasa() 
        self._tab_ekstre()

    def _tab_daireler(self):
        w = QWidget()
        lay = QVBoxLayout(w)

        gb = QGroupBox("Daire Ekle / Güncelle")
        form = QFormLayout(gb)
        self.d_in_no = QLineEdit()
        self.d_in_ad = QLineEdit()
        self.d_in_tel = QLineEdit()
        self.d_in_aidat = QLineEdit()
        self.d_chk_aktif = QCheckBox("Aktif")
        self.d_chk_aktif.setChecked(True)

        form.addRow("Daire No", self.d_in_no)
        form.addRow("İsim", self.d_in_ad)
        form.addRow("Telefon (GSM)", self.d_in_tel)
        form.addRow("Aidat (TL)", self.d_in_aidat)
        form.addRow("", self.d_chk_aktif)

        btns = QHBoxLayout()
        self.btn_d_save = QPushButton("Kaydet (Ekle/Güncelle)")
        self.btn_d_save.clicked.connect(self.daire_save)
        self.btn_d_clear = QPushButton("Temizle")
        self.btn_d_clear.clicked.connect(self.daire_clear)
        self.btn_d_deactivate = QPushButton("Seçileni Pasife Al")
        self.btn_d_deactivate.clicked.connect(self.daire_deactivate_selected)
        self.btn_d_activate = QPushButton("Seçileni Aktif Yap")
        self.btn_d_activate.clicked.connect(self.daire_activate_selected)

        btns.addWidget(self.btn_d_save)
        btns.addWidget(self.btn_d_clear)
        btns.addStretch(1)
        btns.addWidget(self.btn_d_deactivate)
        btns.addWidget(self.btn_d_activate)

        lay.addWidget(gb)
        lay.addLayout(btns)

        rowwrap = QSplitter(Qt.Horizontal)
        lay.addWidget(rowwrap, 1)

        # --- Sol: Daire listesi ---
        self.tbl_daire = QTableWidget(0, 6)
        self.tbl_daire.setHorizontalHeaderLabels(["ID", "Daire", "İsim", "Telefon", "Aidat", "Aktif"])
        self.tbl_daire.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_daire.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_daire.cellClicked.connect(self.daire_row_clicked)
        self.tbl_daire.horizontalHeader().setStretchLastSection(True)
        self.tbl_daire.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        rowwrap.addWidget(self.tbl_daire)

        # --- Sağ: Özet + Devir + Sakin + Dönem listesi ---
        right_w = QWidget()
        right = QVBoxLayout(right_w)
        right.setContentsMargins(0, 0, 0, 0)
        right.setSpacing(8)

        gb_sum = QGroupBox("Seçili Daire Özeti")
        fsum = QFormLayout(gb_sum)

        self.lbl_sel = QLabel("-")
        self.lbl_total_borc = QLabel("0.00")
        self.lbl_total_odeme = QLabel("0.00")
        self.lbl_total_bakiye = QLabel("0.00")

        self.lbl_period_borc = QLabel("0.00")
        self.lbl_period_odeme = QLabel("0.00")
        self.lbl_period_bakiye = QLabel("0.00")
        self.lbl_period_durum = QLabel("-")

        fsum.addRow("Daire", self.lbl_sel)
        fsum.addRow("Genel Borç", self.lbl_total_borc)
        fsum.addRow("Genel Ödeme", self.lbl_total_odeme)
        fsum.addRow("Genel Bakiye", self.lbl_total_bakiye)
        fsum.addRow("Seçili Dönem Borç", self.lbl_period_borc)
        fsum.addRow("Seçili Dönem Ödeme", self.lbl_period_odeme)
        fsum.addRow("Seçili Dönem Bakiye", self.lbl_period_bakiye)
        fsum.addRow("Seçili Dönem Durum", self.lbl_period_durum)
        right.addWidget(gb_sum, 1)

        # Devir paneli
        gb_dev = QGroupBox("Daire Devir (Taşındı)")
        fdev = QFormLayout(gb_dev)
        self.dev_cikis = QDateEdit()
        self.dev_cikis.setCalendarPopup(True)
        self.dev_cikis.setDisplayFormat("dd-MM-yyyy")
        self.dev_cikis.setDate(QDate.currentDate())

        self.dev_basla = QDateEdit()
        self.dev_basla.setCalendarPopup(True)
        self.dev_basla.setDisplayFormat("dd-MM-yyyy")
        self.dev_basla.setDate(QDate.currentDate())

        self.dev_ad = QLineEdit()
        self.dev_tel = QLineEdit()

        self.btn_dev = QPushButton("DEVRİ KAYDET")
        self.btn_dev.clicked.connect(self.devir_yap_selected)

        fdev.addRow("Çıkış Tarihi", self.dev_cikis)
        fdev.addRow("Yeni Başlangıç", self.dev_basla)
        fdev.addRow("Yeni İsim", self.dev_ad)
        fdev.addRow("Yeni Telefon", self.dev_tel)
        fdev.addRow("", self.btn_dev)
        right.addWidget(gb_dev, 1)

        gb_sakin = QGroupBox("Sakin Geçmişi (Seçili Daire)")
        vsh = QVBoxLayout(gb_sakin)
        self.tbl_sakin_hist = QTableWidget(0, 4)
        self.tbl_sakin_hist.setHorizontalHeaderLabels(["Ad Soyad", "Telefon", "Başlangıç", "Bitiş"])
        self.tbl_sakin_hist.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_sakin_hist.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_sakin_hist.horizontalHeader().setStretchLastSection(True)
        self.tbl_sakin_hist.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        vsh.addWidget(self.tbl_sakin_hist, 1)
        right.addWidget(gb_sakin, 2)

        # Sağ paneli scroll içine al (küçük ekranda taşma olmasın)
        scr = QScrollArea()
        scr.setWidgetResizable(True)
        scr.setFrameShape(QScrollArea.NoFrame)
        scr.setWidget(right_w)
        rowwrap.addWidget(scr)

        # Oran: sol geniş, sağ dar (responsive)
        rowwrap.setStretchFactor(0, 3)
        rowwrap.setStretchFactor(1, 2)

        self.tabs.addTab(w, "Daireler")
    def _tab_tahakkuk(self):
        w = QWidget()
        lay = QVBoxLayout(w)

        top = QGroupBox("Tahakkuk / Borç")
        form = QFormLayout(top)

        self.t_cmb_daire = QComboBox()
        self.t_in_from = QLineEdit(ym_of_today())
        self.t_in_to = QLineEdit(ym_of_today())
        self.t_in_tutar = QLineEdit()
        self.t_in_tutar.setPlaceholderText("Örn: 400")

        btns = QHBoxLayout()
        self.btn_t_add = QPushButton("Tahakkuk Oluştur")
        self.btn_t_import = QPushButton("Excel'den Borç İçe Aktar")
        self.btn_t_delete = QPushButton("Seçili Borcu Sil")
        btns.addWidget(self.btn_t_add)
        btns.addWidget(self.btn_t_import)
        btns.addWidget(self.btn_t_delete)

        form.addRow("Daire", self.t_cmb_daire)
        form.addRow("Başlangıç", self.t_in_from)
        form.addRow("Bitiş", self.t_in_to)
        form.addRow("Tutar", self.t_in_tutar)
        form.addRow("", btns)
        lay.addWidget(top)

        self.tbl_tah = QTableWidget(0, 5)
        self.tbl_tah.setHorizontalHeaderLabels(["ID", "Daire", "Dönem", "Tutar", "Tarih"])
        self.tbl_tah.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_tah.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_tah.horizontalHeader().setStretchLastSection(True)
        self.tbl_tah.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_tah, 1)

        self.btn_t_add.clicked.connect(self.add_tahakkuk_range)
        self.btn_t_import.clicked.connect(self.import_debts_from_excel)
        self.btn_t_delete.clicked.connect(self.delete_selected_tahakkuk)

        self.tabs.addTab(w, "Tahakkuk")
        
    def delete_selected_tahakkuk(self):
        row = self.tbl_tah.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Silmek için bir tahakkuk seçin.")
            return

        try:
            tah_id = int(self.tbl_tah.item(row, 0).text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Geçerli bir tahakkuk satırı seçin.")
            return

        if QMessageBox.question(self, "Onay", "Seçili borç kaydı silinsin mi?") != QMessageBox.Yes:
            return

        con = connect()
        con.execute("DELETE FROM tahakkuk WHERE id=?", (tah_id,))
        con.commit()
        con.close()

        self.refresh_all()

    
    def _tab_odeme(self):
        w = QWidget()
        lay = QVBoxLayout(w)

        gb = QGroupBox("Aidat Ödeme")
        form = QFormLayout(gb)

        self.o_cmb_mode = QComboBox()  # ⭐ FORM BAŞINDA OLMALI
        self.o_cmb_mode.addItems([
            "Tek Dönem Borç Kapatma",
            "Peşin Çok Dönem"
        ])

        self.o_in_donem = QLineEdit(ym_of_today())
        self.o_cmb_daire = QComboBox()

        self.o_dt_tarih = QDateEdit()
        self.o_dt_tarih.setCalendarPopup(True)
        self.o_dt_tarih.setDisplayFormat("dd-MM-yyyy")
        self.o_dt_tarih.setDate(QDate.currentDate())

        self.o_sp_ay_sayisi = QSpinBox()
        self.o_sp_ay_sayisi.setRange(1, 24)
        self.o_sp_ay_sayisi.setValue(1)
        self.o_cmb_daire.currentIndexChanged.connect(self.load_borc_listesi)

        self.o_in_tutar = QLineEdit()
        self.o_in_tutar.setPlaceholderText(
            "Tek dönemde boşsa kalan bakiye alınır. Peşin modda boşsa aidat×ay."
        )

        self.o_lbl_bakiye = QLabel("Borç: 0.00 TL | Ödeme: 0.00 TL | Kalan: 0.00 TL")

        self.o_btn_fill_balance = QPushButton("Kalanı Doldur")
        self.o_btn_fill_balance.clicked.connect(self.fill_selected_period_balance)

        self.o_cmb_yontem = QComboBox()
        self.o_cmb_yontem.addItems(["Banka", "Elden"])

        self.o_in_acik = QLineEdit()

        self.btn_o_save = QPushButton("Kaydet")
        self.btn_o_save.clicked.connect(self.odeme_save_multi)

        self.btn_o_whatsapp = QPushButton("Seçili Daireye WhatsApp Mesaj Aç")
        self.btn_o_whatsapp.clicked.connect(self.open_whatsapp_for_payment_tab)

        form.addRow("Ödeme Modu", self.o_cmb_mode)
        form.addRow("Borç / Başlangıç Dönem", self.o_in_donem)
        form.addRow("Daire", self.o_cmb_daire)
        form.addRow("Ödeme Tarihi", self.o_dt_tarih)
        form.addRow("Seçili Dönem Bakiye", self.o_lbl_bakiye)
        form.addRow("", self.o_btn_fill_balance)
        form.addRow("Kaç Ay Peşin?", self.o_sp_ay_sayisi)
        form.addRow("Toplam Tutar (TL)", self.o_in_tutar)
        form.addRow("Yöntem", self.o_cmb_yontem)
        form.addRow("Açıklama", self.o_in_acik)
        form.addRow("", self.btn_o_save)
        form.addRow("", self.btn_o_whatsapp)
        self.tbl_borc_list = QTableWidget(0, 5)
        self.tbl_borc_list.setHorizontalHeaderLabels([
            "Seç",
            "Dönem",
            "Borç",
            "Ödenen",
            "Kalan"
        ])

        self.tbl_borc_list.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl_borc_list.setSelectionBehavior(QTableWidget.SelectRows)

        lay.addWidget(QLabel("Borçlu Dönemler"))
        lay.addWidget(self.tbl_borc_list)
        lay.addWidget(gb)

        # ============ SPLITTER: Üst (yeni ödeme) + Alt (geçmiş) ============
        splitter = QSplitter(Qt.Vertical)
    
        # Üst: Ödeme kaydı tablosu
        gb_top = QGroupBox("Seçili Dönem Ödemeleri")
        vlay_top = QVBoxLayout(gb_top)
    
        self.tbl_odeme = QTableWidget(0, 7)
        self.tbl_odeme.setHorizontalHeaderLabels(["ID", "Dönem", "Tarih", "Daire", "Tutar", "Yöntem", "Makbuz"])
        self.tbl_odeme.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_odeme.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_odeme.horizontalHeader().setStretchLastSection(True)
        self.tbl_odeme.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
        btns = QHBoxLayout()
        self.btn_odeme_del = QPushButton("Seçili Ödemeyi Sil")
        self.btn_odeme_del.clicked.connect(self.delete_selected_payment)
        btns.addWidget(self.btn_odeme_del)
        btns.addStretch(1)
    
        vlay_top.addLayout(btns)
        vlay_top.addWidget(self.tbl_odeme)
    
        splitter.addWidget(gb_top)
    
        # Alt: Seçili dairenin geçmiş ödemeleri
        gb_bottom = QGroupBox("Seçili Daire - Geçmiş Ödemeler (Düzenlemek için çift tıkla)")
        vlay_bottom = QVBoxLayout(gb_bottom)
    
        self.tbl_odeme_gecmis = QTableWidget(0, 6)
        self.tbl_odeme_gecmis.setHorizontalHeaderLabels(["ID", "Dönem", "Tarih", "Tutar", "Yöntem", "Makbuz"])
        self.tbl_odeme_gecmis.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_odeme_gecmis.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_odeme_gecmis.doubleClicked.connect(self.open_payment_for_edit)
        self.tbl_odeme_gecmis.horizontalHeader().setStretchLastSection(True)
        self.tbl_odeme_gecmis.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
        vlay_bottom.addWidget(self.tbl_odeme_gecmis)
        splitter.addWidget(gb_bottom)
    
        # Ratio: üst 40%, alt 60%
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 3)
    
        lay.addWidget(splitter, 1)

        # ============ CONNECTIONS (Form tanımından SONRA) ============
        self.o_cmb_mode.currentIndexChanged.connect(self._set_payment_mode_ui)
        self.o_cmb_daire.currentIndexChanged.connect(self.update_payment_period_balance)
        self.o_cmb_daire.currentIndexChanged.connect(self.load_payment_history)
        self.o_in_donem.textChanged.connect(self.update_payment_period_balance)

        self.tabs.addTab(w, "Ödeme")
    def _tab_gider(self):
        w = QWidget()
        lay = QVBoxLayout(w)

        gb = QGroupBox("Gider Ekle")
        form = QFormLayout(gb)

        self.g_in_donem = QLineEdit(ym_of_today())
        self.g_dt_tarih = QDateEdit()
        self.g_dt_tarih.setCalendarPopup(True)
        self.g_dt_tarih.setDisplayFormat("dd-MM-yyyy")
        self.g_dt_tarih.setDate(QDate.currentDate())

        self.g_cmb_kat = QComboBox()
        self.g_cmb_kat.addItems(["Elektrik", "Temizlik", "Bakım", "Diğer"])

        self.g_in_tutar = QLineEdit()

        self.g_cmb_yontem = QComboBox()
        self.g_cmb_yontem.addItems(["Banka", "Elden"])

        self.g_in_acik = QLineEdit()

        self.btn_g_save = QPushButton("Gider Kaydet")
        self.btn_g_save.clicked.connect(self.gider_save)

        form.addRow("Dönem", self.g_in_donem)
        form.addRow("Tarih", self.g_dt_tarih)
        form.addRow("Kategori", self.g_cmb_kat)
        form.addRow("Tutar (TL)", self.g_in_tutar)
        form.addRow("Yöntem", self.g_cmb_yontem)
        form.addRow("Açıklama", self.g_in_acik)
        form.addRow("", self.btn_g_save)
        lay.addWidget(gb)

        # ============ SPLITTER: Üst (yeni gider) + Alt (geçmiş) ============
        splitter = QSplitter(Qt.Vertical)
    
        # Üst: Gider kaydı tablosu
        gb_top = QGroupBox("Seçili Dönem Giderleri")
        vlay_top = QVBoxLayout(gb_top)
    
        self.tbl_gider = QTableWidget(0, 7)
        self.tbl_gider.setHorizontalHeaderLabels(["ID", "Dönem", "Tarih", "Kategori", "Tutar", "Yöntem", "Açıklama"])
        self.tbl_gider.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_gider.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_gider.horizontalHeader().setStretchLastSection(True)
        self.tbl_gider.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        btns = QHBoxLayout()
        self.btn_gider_del = QPushButton("Seçili Gideri Sil")
        self.btn_gider_del.clicked.connect(self.delete_selected_expense)
        btns.addWidget(self.btn_gider_del)
        btns.addStretch(1)

        vlay_top.addLayout(btns)
        vlay_top.addWidget(self.tbl_gider)
    
        splitter.addWidget(gb_top)
    
        # Alt: Geçmiş giderler (tüm zamanlar)
        gb_bottom = QGroupBox("Tüm Giderler (Düzenlemek için çift tıkla)")
        vlay_bottom = QVBoxLayout(gb_bottom)
    
        self.tbl_gider_gecmis = QTableWidget(0, 7)
        self.tbl_gider_gecmis.setHorizontalHeaderLabels(["ID", "Dönem", "Tarih", "Kategori", "Tutar", "Yöntem", "Açıklama"])
        self.tbl_gider_gecmis.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_gider_gecmis.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_gider_gecmis.doubleClicked.connect(self.open_expense_for_edit)  # ⭐ DOUBLE CLICK
        self.tbl_gider_gecmis.horizontalHeader().setStretchLastSection(True)
        self.tbl_gider_gecmis.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
        vlay_bottom.addWidget(self.tbl_gider_gecmis)
        splitter.addWidget(gb_bottom)
    
        # Ratio: üst 30%, alt 70%
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
    
        lay.addWidget(splitter, 1)

        self.tabs.addTab(w, "Gider")
    def load_expense_history(self):
        """Geçmiş giderleri yükle"""
        self.tbl_gider_gecmis.setRowCount(0)
    
        con = connect()
        rows = con.execute("""
            SELECT id, donem, tarih, kategori, tutar, yontem, aciklama
            FROM giderler
            ORDER BY tarih DESC, donem DESC
        """).fetchall()
        con.close()
    
        for gid, donem, tarih, kat, tutar, yontem, acik in rows:
            row = self.tbl_gider_gecmis.rowCount()
            self.tbl_gider_gecmis.insertRow(row)
        
            vals = [str(gid), donem, tarih, kat, f"{float(tutar):.2f}", yontem, acik or ""]
        
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                if col == 4:  # Tutar
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item.setBackground(QColor(255, 245, 220))  # Açık turuncu
                self.tbl_gider_gecmis.setItem(row, col, item)

    def open_expense_for_edit(self, index):
        """Çift tıklanan gideri düzenleme penceresinde aç"""
        row = index.row()
    
        try:
            expense_id = int(self.tbl_gider_gecmis.item(row, 0).text())
        except Exception:
            QMessageBox.warning(self, "Hata", "Gider ID'si alınamadı.")
            return
    
        # Modal dialog aç
        dialog = EditExpenseDialog(expense_id, self)
        dialog.exec()
    
        # Dialog kapandıktan sonra tabloları yenile
        self.load_expense_history()
        self.refresh_expenses_table()
        self.refresh_all()    

    def _tab_duyuru(self):
        w = QWidget()
        lay = QVBoxLayout(w)

        gb = QGroupBox("Apartman Duyurusu")
        form = QFormLayout(gb)

        self.duy_dt = QDateEdit()
        self.duy_dt.setCalendarPopup(True)
        self.duy_dt.setDisplayFormat("dd-MM-yyyy")
        self.duy_dt.setDate(QDate.currentDate())

        self.duy_baslik = QLineEdit()
        self.duy_mesaj = QTextEdit()

        self.btn_duy_save = QPushButton("Duyuru Kaydet")
        self.btn_duy_save.clicked.connect(self.duyuru_save)

        self.btn_duy_copy = QPushButton("Metni Panoya Kopyala")
        self.btn_duy_copy.clicked.connect(self.duyuru_copy)

        form.addRow("Tarih", self.duy_dt)
        form.addRow("Başlık", self.duy_baslik)
        form.addRow("Mesaj", self.duy_mesaj)
        form.addRow("", self.btn_duy_save)
        form.addRow("", self.btn_duy_copy)
        lay.addWidget(gb)

        self.tbl_duyuru = QTableWidget(0, 4)
        self.tbl_duyuru.setHorizontalHeaderLabels(["ID", "Tarih", "Başlık", "Mesaj (kısa)"])
        self.tbl_duyuru.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_duyuru.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_duyuru.horizontalHeader().setStretchLastSection(True)
        self.tbl_duyuru.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_duyuru, 1)

        btns = QHBoxLayout()
        self.btn_duy_del = QPushButton("Seçileni Sil")
        self.btn_duy_del.clicked.connect(self.duyuru_delete_selected)
        btns.addWidget(self.btn_duy_del)
        btns.addStretch(1)
        lay.addLayout(btns)

        self.tabs.addTab(w, "Duyuru")
    def load_borc_listesi(self):
        self.tbl_borc_list.setRowCount(0)

        daire_id = self.o_cmb_daire.currentData()
        if not daire_id:
            return

        con = connect()

        rows = con.execute("""
            SELECT DISTINCT donem
            FROM tahakkuk
            WHERE daire_id=?
            ORDER BY donem
        """, (daire_id,)).fetchall()

        for (donem,) in rows:
            borc = con.execute(
                "SELECT COALESCE(SUM(tutar),0) FROM tahakkuk WHERE daire_id=? AND donem=?",
                (daire_id, donem)
            ).fetchone()[0] or 0

            odeme = con.execute(
                "SELECT COALESCE(SUM(od.tutar),0) "
                "FROM odeme_detay od "
                "JOIN odemeler o ON o.id = od.odeme_id "
                "WHERE o.daire_id=? AND od.donem=?",
                (daire_id, donem)
            ).fetchone()[0] or 0

            kalan = float(borc) - float(odeme)

            if kalan <= 0:
                continue

            row = self.tbl_borc_list.rowCount()
            self.tbl_borc_list.insertRow(row)

            chk = QCheckBox()
            self.tbl_borc_list.setCellWidget(row, 0, chk)

            vals = [
                donem,
                f"{float(borc):.2f}",
                f"{float(odeme):.2f}",
                f"{float(kalan):.2f}"
            ]

            for col, val in enumerate(vals, start=1):
                item = QTableWidgetItem(val)
                self.tbl_borc_list.setItem(row, col, item)

        con.close()
    def _tab_rapor(self):
        w = QWidget()
        lay = QVBoxLayout(w)

        head = QHBoxLayout()
        lay.addLayout(head)

        head.addWidget(QLabel("Rapor Dönem:"))
        self.r_in_donem = QLineEdit(ym_of_today())
        self.r_in_donem.setMaximumWidth(110)
        head.addWidget(self.r_in_donem)

        self.r_chk_only_unpaid = QCheckBox("Sadece borçlu")
        self.r_chk_only_overdue = QCheckBox("Sadece geciken")
        head.addWidget(self.r_chk_only_unpaid)
        head.addWidget(self.r_chk_only_overdue)

        self.btn_r_calc = QPushButton("Hesapla/Yenile")
        self.btn_r_calc.clicked.connect(self.refresh_report)
        head.addWidget(self.btn_r_calc)

        self.btn_export_excel = QPushButton("Excel'e Aktar")
        self.btn_export_excel.clicked.connect(self.export_excel)
        head.addWidget(self.btn_export_excel)

        head.addStretch(1)

        self.lbl_summary = QLabel("Özet: -")
        self.lbl_summary.setTextInteractionFlags(Qt.TextSelectableByMouse)
        lay.addWidget(self.lbl_summary)

        self.tbl_report = QTableWidget(0, 8)
        self.tbl_report.setHorizontalHeaderLabels([
            "Daire", "İsim", "Telefon", "Aidat", "Borç (Dönem)", "Ödeme (Dönem)", "Bakiye", "Durum"
        ])
        self.tbl_report.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_report.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_report.horizontalHeader().setStretchLastSection(True)
        self.tbl_report.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_report, 1)

        foot = QHBoxLayout()
        lay.addLayout(foot)

        self.btn_msg_selected = QPushButton("Seçili Daire WhatsApp")
        self.btn_msg_selected.clicked.connect(self.report_whatsapp_selected)
        foot.addWidget(self.btn_msg_selected)

        self.btn_msg_all_debt = QPushButton("Borçlulara Toplu Metin Kopyala")
        self.btn_msg_all_debt.clicked.connect(self.copy_bulk_message_text)
        foot.addWidget(self.btn_msg_all_debt)

        foot.addStretch(1)

        self.tabs.addTab(w, "Rapor")
    def _tab_kasa(self):
        """Kasa (Cari) Raporu Sekmesi"""
        w = QWidget()
        lay = QVBoxLayout(w)

        # Filtreleme paneli
        head = QHBoxLayout()
        lay.addLayout(head)

        head.addWidget(QLabel("Başlangıç:"))
        self.k_dt_from = QDateEdit()
        self.k_dt_from.setCalendarPopup(True)
        self.k_dt_from.setDisplayFormat("dd-MM-yyyy")
        self.k_dt_from.setDate(QDate.currentDate().addMonths(-24))
        head.addWidget(self.k_dt_from)

        head.addWidget(QLabel("Bitiş:"))
        self.k_dt_to = QDateEdit()
        self.k_dt_to.setCalendarPopup(True)
        self.k_dt_to.setDisplayFormat("dd-MM-yyyy")
        self.k_dt_to.setDate(QDate.currentDate())
        head.addWidget(self.k_dt_to)

        self.btn_k_refresh = QPushButton("Yenile")
        self.btn_k_refresh.clicked.connect(self.refresh_kasa_report)
        head.addWidget(self.btn_k_refresh)

        self.btn_k_excel = QPushButton("Excel'e Aktar")
        self.btn_k_excel.clicked.connect(self.export_kasa_excel)
        head.addWidget(self.btn_k_excel)

        head.addStretch(1)

        # Özet paneli
        gb_ozet = QGroupBox("Kasa Özeti")
        form_ozet = QFormLayout(gb_ozet)

        self.k_lbl_toplam_gelir = QLabel("0.00 TL")
        self.k_lbl_toplam_gider = QLabel("0.00 TL")
        self.k_lbl_net = QLabel("0.00 TL")
        self.k_lbl_durum = QLabel("-")

        self.k_lbl_gelir_enko = QLabel("0 kayıt")
        self.k_lbl_gider_enko = QLabel("0 kayıt")

        form_ozet.addRow("📥 Toplam Gelir (Tahsilat):", self.k_lbl_toplam_gelir)
        form_ozet.addRow("   Kayıt Sayısı:", self.k_lbl_gelir_enko)
        form_ozet.addRow("📤 Toplam Gider:", self.k_lbl_toplam_gider)
        form_ozet.addRow("   Kayıt Sayısı:", self.k_lbl_gider_enko)
        form_ozet.addRow("📊 NET KALAN:", self.k_lbl_net)
        form_ozet.addRow("Durum:", self.k_lbl_durum)

        lay.addWidget(gb_ozet)

        # Gelir tablosu
        gb_gelir = QGroupBox("Gelir Kaynakları (Ödemeler)")
        vg = QVBoxLayout(gb_gelir)

        self.tbl_k_gelir = QTableWidget(0, 5)
        self.tbl_k_gelir.setHorizontalHeaderLabels(["Dönem", "Tarih", "Daire", "Yöntem", "Tutar"])
        self.tbl_k_gelir.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_k_gelir.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_k_gelir.horizontalHeader().setStretchLastSection(True)
        self.tbl_k_gelir.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        vg.addWidget(self.tbl_k_gelir)

        lay.addWidget(gb_gelir, 2)

        # Gider tablosu
        gb_gider = QGroupBox("Gider Kaynakları")
        vx = QVBoxLayout(gb_gider)

        self.tbl_k_gider = QTableWidget(0, 5)
        self.tbl_k_gider.setHorizontalHeaderLabels(["Dönem", "Tarih", "Kategori", "Yöntem", "Tutar"])
        self.tbl_k_gider.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_k_gider.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_k_gider.horizontalHeader().setStretchLastSection(True)
        self.tbl_k_gider.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        vx.addWidget(self.tbl_k_gider)

        lay.addWidget(gb_gider, 2)

        self.tabs.addTab(w, "Kasa")
        

    def _tab_ekstre(self):
        w = QWidget()
        lay = QVBoxLayout(w)

        head = QHBoxLayout()
        lay.addLayout(head)

        head.addWidget(QLabel("Daire:"))
        self.e_cmb_daire = QComboBox()
        head.addWidget(self.e_cmb_daire, 2)

        head.addWidget(QLabel("Başlangıç:"))
        self.e_from = QLineEdit("2025-01")
        self.e_from.setMaximumWidth(90)
        head.addWidget(self.e_from)

        head.addWidget(QLabel("Bitiş:"))
        self.e_to = QLineEdit(ym_of_today())
        self.e_to.setMaximumWidth(90)
        head.addWidget(self.e_to)

        self.btn_e_list = QPushButton("Listele")
        self.btn_e_list.clicked.connect(self.load_ekstre)
        head.addWidget(self.btn_e_list)

        self.btn_e_excel = QPushButton("Excel'e Aktar")
        self.btn_e_excel.clicked.connect(self.export_ekstre_excel)
        head.addWidget(self.btn_e_excel)

        self.btn_e_print = QPushButton("Yazdır")
        self.btn_e_print.clicked.connect(self.print_ekstre)
        head.addWidget(self.btn_e_print)

        head.addStretch(1)

        self.tbl_ekstre = QTableWidget(0, 6)
        self.tbl_ekstre.setHorizontalHeaderLabels(["Daire", "Dönem", "Borç", "Ödeme", "Bakiye", "Durum"])
        self.tbl_ekstre.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_ekstre.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_ekstre.horizontalHeader().setStretchLastSection(True)
        self.tbl_ekstre.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        lay.addWidget(self.tbl_ekstre, 1)

        self.tabs.addTab(w, "Ekstre")


    # ---------------- actions ----------------
    def save_settings(self):
        set_setting_int("vade_gun", int(self.sp_vade.value()))
        set_setting_int("gecikme_gun", int(self.sp_gec.value()))
        QMessageBox.information(self, "OK", "Ayarlar kaydedildi.")
        self.refresh_all()

    # --- daireler ---
    def daire_clear(self):
        self.selected_daire_id = None
        self.d_in_no.clear()
        self.d_in_ad.clear()
        self.d_in_tel.clear()
        self.d_in_aidat.clear()
        self.d_chk_aktif.setChecked(True)
        self.tbl_daire.clearSelection()
        self.refresh_selected_summary(clear=True)
        self.tbl_sakin_hist.setRowCount(0)
        self.dev_ad.clear()
        self.dev_tel.clear()

    def daire_row_clicked(self, row, col):
        try:
            did = int(self.tbl_daire.item(row, 0).text())
            dno = self.tbl_daire.item(row, 1).text()
            ad = self.tbl_daire.item(row, 2).text()
            tel = self.tbl_daire.item(row, 3).text()
            aidat = self.tbl_daire.item(row, 4).text()
            aktif = self.tbl_daire.item(row, 5).text() == "1"
        except Exception:
            return

        self.selected_daire_id = did

        self.d_in_no.setText(dno)
        self.d_in_ad.setText(ad)
        self.d_in_tel.setText(tel)
        self.d_in_aidat.setText(aidat)
        self.d_chk_aktif.setChecked(aktif)

        self._ensure_current_sakin(did)
        self.refresh_sakin_history_table(did)
        self.refresh_selected_summary(did)

    def _selected_daire_id(self):
        rows = self.tbl_daire.selectionModel().selectedRows()
        if not rows:
            return None
        r = rows[0].row()
        try:
            return int(self.tbl_daire.item(r, 0).text())
        except Exception:
            return None

    def daire_save(self):
        dno = self.d_in_no.text().strip()
        ad = self.d_in_ad.text().strip()
        tel = self.d_in_tel.text().strip()

        try:
            aidat = safe_float(self.d_in_aidat.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Aidat sayısal olmalı.")
            return

        aktif = 1 if self.d_chk_aktif.isChecked() else 0

        if not dno or not ad:
            QMessageBox.warning(self, "Uyarı", "Daire No ve İsim zorunlu.")
            return

        con = connect()
        try:
            if self.selected_daire_id:
                # başka bir daire aynı numarayı kullanıyor mu?
                conflict = con.execute(
                    "SELECT id FROM daireler WHERE daire_no=? AND id<>?",
                    (dno, int(self.selected_daire_id))
                ).fetchone()

                if conflict:
                    QMessageBox.warning(self, "Uyarı", "Bu daire no başka bir kayıtta kullanılıyor.")
                    con.close()
                    return

                con.execute("""
                    UPDATE daireler
                    SET daire_no=?, ad_soyad=?, telefon=?, aidat=?, aktif=?
                    WHERE id=?
                """, (dno, ad, tel, float(aidat), int(aktif), int(self.selected_daire_id)))

                did = int(self.selected_daire_id)

            else:
                con.execute("""
                    INSERT INTO daireler(daire_no, ad_soyad, telefon, aidat, aktif)
                    VALUES(?,?,?,?,?)
                """, (dno, ad, tel, float(aidat), int(aktif)))

                did = int(con.execute(
                    "SELECT id FROM daireler WHERE daire_no=?",
                    (dno,)
                ).fetchone()[0])

            # aktif sakin sync
            has = con.execute(
                "SELECT id FROM sakinler WHERE daire_id=? AND bitis_tarihi IS NULL ORDER BY id DESC LIMIT 1",
                (did,)
            ).fetchone()

            if not has:
                con.execute("""
                    INSERT INTO sakinler(daire_id, ad_soyad, telefon, baslangic_tarihi, bitis_tarihi)
                    VALUES(?,?,?,?,NULL)
                """, (did, ad, tel, iso_today()))
            else:
                con.execute(
                    "UPDATE sakinler SET ad_soyad=?, telefon=? WHERE id=?",
                    (ad, tel, int(has[0]))
                )

            con.commit()

        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Uyarı", "Bu daire no zaten var.")
        finally:
            con.close()

        self.refresh_all()
        self.daire_clear()

    def daire_deactivate_selected(self):
        did = self._selected_daire_id()
        if not did:
            QMessageBox.warning(self, "Uyarı", "Tablodan bir daire seçin.")
            return
        con = connect()
        con.execute("UPDATE daireler SET aktif=0 WHERE id=?", (did,))
        con.commit()
        con.close()
        self.refresh_all()

    def daire_activate_selected(self):
        did = self._selected_daire_id()
        if not did:
            QMessageBox.warning(self, "Uyarı", "Tablodan bir daire seçin.")
            return
        con = connect()
        con.execute("UPDATE daireler SET aktif=1 WHERE id=?", (did,))
        con.commit()
        con.close()
        self.refresh_all()

    # --- sakin (oturan) yönetimi ---
    def _ensure_current_sakin(self, daire_id: int):
        con = connect()
        has = con.execute(
            "SELECT 1 FROM sakinler WHERE daire_id=? AND bitis_tarihi IS NULL LIMIT 1",
            (int(daire_id),)
        ).fetchone()
        if has:
            con.close()
            return
        d = con.execute("SELECT ad_soyad, telefon FROM daireler WHERE id=?", (int(daire_id),)).fetchone()
        if not d:
            con.close()
            return
        ad, tel = d
        con.execute(
            "INSERT INTO sakinler(daire_id, ad_soyad, telefon, baslangic_tarihi, bitis_tarihi) "
            "VALUES(?,?,?,?,NULL)",
            (int(daire_id), str(ad or '').strip() or f"Daire {daire_id}", str(tel or '').strip(), iso_today())
        )
        con.commit()
        con.close()

    def refresh_sakin_history_table(self, daire_id):
        self.tbl_sakin_hist.setRowCount(0)
        if not daire_id:
            return
        con = connect()
        rows = con.execute(
            "SELECT ad_soyad, telefon, baslangic_tarihi, COALESCE(bitis_tarihi,'') "
            "FROM sakinler WHERE daire_id=? ORDER BY id DESC",
            (int(daire_id),)
        ).fetchall()
        con.close()
        for ad, tel, bas, bit in rows:
            r = self.tbl_sakin_hist.rowCount()
            self.tbl_sakin_hist.insertRow(r)
            for c, v in enumerate([ad, tel, bas, bit]):
                it = QTableWidgetItem(str(v))
                if c in (2, 3):
                    it.setTextAlignment(Qt.AlignCenter)
                self.tbl_sakin_hist.setItem(r, c, it)

    def devir_yap_selected(self):
        daire_id = self._selected_daire_id()
        if not daire_id:
            QMessageBox.warning(self, "Uyarı", "Devir için tablodan bir daire seçin.")
            return

        cikis = self.dev_cikis.date().toPython().isoformat()
        basla = self.dev_basla.date().toPython().isoformat()
        yeni_ad = self.dev_ad.text().strip()
        yeni_tel = self.dev_tel.text().strip()

        if not yeni_ad:
            QMessageBox.warning(self, "Uyarı", "Yeni isim zorunlu.")
            return

        self._ensure_current_sakin(int(daire_id))

        con = connect()
        cur = con.execute(
            "SELECT id FROM sakinler WHERE daire_id=? AND bitis_tarihi IS NULL ORDER BY id DESC LIMIT 1",
            (int(daire_id),)
        ).fetchone()
        if cur:
            con.execute("UPDATE sakinler SET bitis_tarihi=? WHERE id=?", (cikis, int(cur[0])))

        con.execute(
            "INSERT INTO sakinler(daire_id, ad_soyad, telefon, baslangic_tarihi, bitis_tarihi) "
            "VALUES(?,?,?,?,NULL)",
            (int(daire_id), yeni_ad, yeni_tel, basla)
        )

        # daireler tablosunda güncel kişi/telefonu da göster
        con.execute("UPDATE daireler SET ad_soyad=?, telefon=? WHERE id=?", (yeni_ad, yeni_tel, int(daire_id)))

        con.commit()
        con.close()

        QMessageBox.information(self, "OK", "Devir işlemi tamamlandı.")
        self.refresh_all()
        self.refresh_sakin_history_table(int(daire_id))

    # --- ekstre (dönem dönem liste ayrı sekme) ---
    def load_ekstre(self):
        daire_id = getattr(self, "e_cmb_daire", None).currentData() if hasattr(self, "e_cmb_daire") else None
        if not daire_id:
            QMessageBox.warning(self, "Uyarı", "Daire seçin.")
            return

        ym_from = self.e_from.text().strip()
        ym_to = self.e_to.text().strip()
        if not validate_period(ym_from) or not validate_period(ym_to):
            QMessageBox.warning(self, "Uyarı", "Dönem formatı YYYY-MM olmalı. Örn: 2025-01")
            return

        periods = months_range(ym_from, ym_to)

        con = connect()
        d = con.execute("SELECT daire_no, ad_soyad FROM daireler WHERE id=?", (int(daire_id),)).fetchone()
        daire_tag = f"Daire {d[0]} - {d[1]}" if d else f"ID {daire_id}"

        self.tbl_ekstre.setRowCount(0)

        for p in periods:
            b = con.execute(
                "SELECT COALESCE(SUM(tutar),0) FROM tahakkuk WHERE daire_id=? AND donem=?",
                (int(daire_id), p)
            ).fetchone()[0] or 0.0
            o = con.execute(
                "SELECT COALESCE(SUM(tutar),0) FROM odemeler WHERE daire_id=? AND donem=?",
                (int(daire_id), p)
            ).fetchone()[0] or 0.0
            bakiye = float(b) - float(o)
            dur = status_for_balance(p, bakiye)

            row_index = self.tbl_ekstre.rowCount()
            self.tbl_ekstre.insertRow(row_index)
            vals = [daire_tag, p, f"{float(b):.2f}", f"{float(o):.2f}", f"{float(bakiye):.2f}", dur]
            for c, v in enumerate(vals):
                it = QTableWidgetItem(str(v))
                if c in (2, 3, 4):
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if c == 1:
                    it.setTextAlignment(Qt.AlignCenter)
                self.tbl_ekstre.setItem(row_index, c, it)

            self._apply_status_row_color(self.tbl_ekstre, row_index, dur)

        con.close()


    def _apply_status_row_color(self, table, row, durum):
        if durum == "Ödedi":
            color = QColor(210, 245, 218)
        else:
            color = QColor(255, 220, 220)
        for col in range(table.columnCount()):
            item = table.item(row, col)
            if item is not None:
                item.setBackground(color)

    def export_ekstre_excel(self):
        if self.tbl_ekstre.rowCount() == 0:
            QMessageBox.information(self, "Bilgi", "Önce ekstreyi listeleyin.")
            return

        daire_text = self.e_cmb_daire.currentText().strip().replace("/", "-")
        ym_from = self.e_from.text().strip()
        ym_to = self.e_to.text().strip()
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Ekstre Excel Kaydet",
            f"ekstre_{daire_text}_{ym_from}_{ym_to}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Ekstre"

        headers = [self.tbl_ekstre.horizontalHeaderItem(c).text() for c in range(self.tbl_ekstre.columnCount())]
        ws.append(headers)
        for r in range(self.tbl_ekstre.rowCount()):
            row_vals = []
            for c in range(self.tbl_ekstre.columnCount()):
                item = self.tbl_ekstre.item(r, c)
                row_vals.append(item.text() if item else "")
            ws.append(row_vals)

        autosize_worksheet(ws)
        wb.save(path)
        QMessageBox.information(self, "OK", f"Ekstre Excel oluşturuldu:\n{path}")

    def print_ekstre(self):
        if self.tbl_ekstre.rowCount() == 0:
            QMessageBox.information(self, "Bilgi", "Önce ekstreyi listeleyin.")
            return

        daire_text = self.e_cmb_daire.currentText().strip() or "Daire"
        ym_from = self.e_from.text().strip()
        ym_to = self.e_to.text().strip()
        rows_html = []
        for r in range(self.tbl_ekstre.rowCount()):
            vals = []
            for c in range(self.tbl_ekstre.columnCount()):
                item = self.tbl_ekstre.item(r, c)
                vals.append(item.text() if item else "")
            durum = vals[5] if len(vals) > 5 else ""
            bg = "#d4f5da" if durum == "Ödedi" else "#ffdcdc"
            cells = "".join(f"<td>{v}</td>" for v in vals)
            rows_html.append(f"<tr style='background:{bg};'>{cells}</tr>")

        headers_html = "".join(
            f"<th>{self.tbl_ekstre.horizontalHeaderItem(c).text()}</th>"
            for c in range(self.tbl_ekstre.columnCount())
        )
        html = f"""<!DOCTYPE html>
<html lang='tr'>
<head>
<meta charset='utf-8'>
<title>Ekstre Yazdır</title>
<style>
body {{ font-family: Arial, sans-serif; margin: 24px; }}
h2 {{ margin-bottom: 4px; }}
p {{ margin-top: 0; color: #444; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ border: 1px solid #999; padding: 8px; text-align: left; }}
th {{ background: #efefef; }}
@media print {{ button {{ display:none; }} body {{ margin:0; }} }}
</style>
</head>
<body>
<button onclick='window.print()'>Yazdır</button>
<h2>{daire_text} Ekstresi</h2>
<p>Dönem aralığı: {ym_from} - {ym_to}</p>
<table>
<thead><tr>{headers_html}</tr></thead>
<tbody>{''.join(rows_html)}</tbody>
</table>
</body>
</html>"""

        out_path = Path.cwd() / "ekstre_yazdir.html"
        out_path.write_text(html, encoding="utf-8")
        webbrowser.open(out_path.resolve().as_uri())

    # --- selected summary panel ---
    def refresh_selected_summary_from_ui(self):
        did = self._selected_daire_id()
        if did:
            self.refresh_selected_summary(did)
        else:
            QMessageBox.information(self, "Bilgi", "Özet için tablodan daire seçin.")

    def refresh_selected_summary(self, daire_id=None, clear=False):
        if clear or not daire_id:
            self.lbl_sel.setText("-")
            for lbl in (
                self.lbl_total_borc, self.lbl_total_odeme, self.lbl_total_bakiye,
                self.lbl_period_borc, self.lbl_period_odeme, self.lbl_period_bakiye
            ):
                lbl.setText("0.00")
            self.lbl_period_durum.setText("-")
            if hasattr(self, 'tbl_sel_hist'):
                self.tbl_sel_hist.setRowCount(0)
            return

        donem = self.in_donem.text().strip()
        if not validate_period(donem):
            donem = ym_of_today()

        con = connect()
        row = con.execute("SELECT daire_no, ad_soyad, telefon FROM daireler WHERE id=?", (int(daire_id),)).fetchone()
        if not row:
            con.close()
            self.refresh_selected_summary(clear=True)
            return
        dno, ad, _tel = row
        self.lbl_sel.setText(f"Daire {dno} - {ad}")

        total_borc, total_odeme, total_bakiye = self._balance_for_daire_total(int(daire_id))

        self.lbl_total_borc.setText(f"{float(total_borc):.2f}")
        self.lbl_total_odeme.setText(f"{float(total_odeme):.2f}")
        self.lbl_total_bakiye.setText(f"{float(total_bakiye):.2f}")
        p_borc = con.execute("SELECT COALESCE(SUM(tutar),0) FROM tahakkuk WHERE daire_id=? AND donem=?",
                             (int(daire_id), donem)).fetchone()[0] or 0.0
        p_odeme = con.execute("SELECT COALESCE(SUM(tutar),0) FROM odemeler WHERE daire_id=? AND donem=?",
                              (int(daire_id), donem)).fetchone()[0] or 0.0
        p_bakiye = float(p_borc) - float(p_odeme)

        self.lbl_period_borc.setText(f"{float(p_borc):.2f}")
        self.lbl_period_odeme.setText(f"{float(p_odeme):.2f}")
        self.lbl_period_bakiye.setText(f"{float(p_bakiye):.2f}")
        self.lbl_period_durum.setText(status_for_balance(donem, p_bakiye))


        con.close()

    # --- tahakkuk ---
   
    def add_tahakkuk_range(self):
        daire_id = self.t_cmb_daire.currentData()
        if not daire_id:
            QMessageBox.warning(self, "Uyarı", "Daire seçin.")
            return

        start = self.t_in_from.text().strip()
        end = self.t_in_to.text().strip()

        if not validate_period(start) or not validate_period(end):
            QMessageBox.warning(self, "Uyarı", "Dönem formatı YYYY-MM olmalı.")
            return

        if start > end:
            QMessageBox.warning(self, "Uyarı", "Başlangıç dönemi bitişten büyük olamaz.")
            return

        try:
            tutar = safe_float(self.t_in_tutar.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Tutar sayısal olmalı.")
            return

        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return

        con = connect()
        adet = 0
        for donem in months_range(start, end):
            con.execute("""
                INSERT INTO tahakkuk(donem, daire_id, tutar, created_at)
                VALUES(?,?,?,?)
                ON CONFLICT(donem, daire_id)
                DO UPDATE SET
                    tutar=excluded.tutar,
                    created_at=excluded.created_at
            """, (donem, int(daire_id), float(tutar), iso_today()))
            adet += 1

        con.commit()
        con.close()

        QMessageBox.information(self, "OK", f"{adet} dönem için tahakkuk oluşturuldu / güncellendi.")
        self.refresh_all()
    def _set_payment_mode_ui(self):
        tek_donem = (self.o_cmb_mode.currentText() == "Tek Dönem Borç Kapatma")
        self.o_sp_ay_sayisi.setEnabled(not tek_donem)
        self.o_btn_fill_balance.setEnabled(tek_donem)

        if tek_donem:
            self.o_sp_ay_sayisi.setValue(1)

        self.update_payment_period_balance()

    def update_payment_period_balance(self):
        if not hasattr(self, "o_lbl_bakiye"):
            return

        donem = self.o_in_donem.text().strip()
        daire_id = self.o_cmb_daire.currentData()

        if not daire_id or not validate_period(donem):
            self.o_lbl_bakiye.setText("Borç: 0.00 TL | Ödeme: 0.00 TL | Kalan: 0.00 TL")
            return

        borc, odeme, bakiye = self._balance_for_daire_period(int(daire_id), donem)
        self.o_lbl_bakiye.setText(
            f"Borç: {borc:.2f} TL | Ödeme: {odeme:.2f} TL | Kalan: {bakiye:.2f} TL"
        )

    def fill_selected_period_balance(self):
        donem = self.o_in_donem.text().strip()
        daire_id = self.o_cmb_daire.currentData()

        if not daire_id:
            QMessageBox.warning(self, "Uyarı", "Daire seçin.")
            return
        if not validate_period(donem):
            QMessageBox.warning(self, "Uyarı", "Dönem formatı YYYY-MM olmalı.")
            return

        borc, odeme, bakiye = self._balance_for_daire_period(int(daire_id), donem)
        if bakiye <= 0.00001:
            QMessageBox.information(self, "Bilgi", "Bu dönemde kalan borç yok.")
            self.o_in_tutar.clear()
            return

        self.o_in_tutar.setText(f"{bakiye:.2f}")

   

    def import_debts_from_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Borç Excel'i Seç", "", "Excel (*.xlsx)")
        if not path:
            return

        try:
            wb = load_workbook(path, data_only=True)
            ws = wb.active
        except Exception as e:
            QMessageBox.warning(self, "Hata", f"Excel açılamadı:\n{e}")
            return

        header = []
        for cell in ws[1]:
            header.append((str(cell.value).strip().lower() if cell.value is not None else ""))

        def find_col(*names):
            for i, h in enumerate(header):
                for n in names:
                    if h == n:
                        return i
            for i, h in enumerate(header):
                h2 = h.replace(" ", "").replace("_", "")
                for n in names:
                    n2 = n.replace(" ", "").replace("_", "")
                    if h2 == n2:
                        return i
            return None

        col_daire = find_col("daire_no", "daire", "daire no", "no")
        col_donem = find_col("donem", "dönem", "period")
        col_tutar = find_col("tutar", "borc", "borç", "amount")

        col_ad = find_col("ad_soyad", "ad", "isim", "name")
        col_tel = find_col("telefon", "gsm", "tel", "phone")
        col_aidat = find_col("aidat", "aidatı", "monthly", "fee")
        col_aktif = find_col("aktif", "status", "active")

        if col_daire is None or col_donem is None or col_tutar is None:
            QMessageBox.warning(self, "Uyarı",
                                "Başlıklar bulunamadı.\nGerekli sütunlar: daire_no / donem / tutar")
            return

        con = connect()
        dmap = {str(r[0]).strip(): int(r[1]) for r in con.execute("SELECT daire_no, id FROM daireler").fetchall()}

        ok = 0
        skipped = 0
        created_flats = 0
        updated_flats = 0
        errors = 0
        msgs = []

        for r in range(2, ws.max_row + 1):
            rowvals = [ws.cell(row=r, column=c + 1).value for c in range(ws.max_column)]

            daire_no = rowvals[col_daire] if col_daire < len(rowvals) else None
            donem = rowvals[col_donem] if col_donem < len(rowvals) else None
            tutar = rowvals[col_tutar] if col_tutar < len(rowvals) else None

            if daire_no is None and donem is None and tutar is None:
                continue

            daire_no = str(daire_no).strip() if daire_no is not None else ""
            donem = str(donem).strip() if donem is not None else ""
            donem = donem.replace(".", "-").replace("/", "-")

            if not daire_no or not validate_period(donem):
                skipped += 1
                msgs.append(f"Satır {r}: daire_no/dönem geçersiz -> {daire_no} / {donem}")
                continue

            try:
                tutar_f = safe_float(str(tutar))
            except Exception:
                errors += 1
                msgs.append(f"Satır {r}: tutar sayı değil -> {tutar}")
                continue

            if tutar_f <= 0:
                skipped += 1
                msgs.append(f"Satır {r}: tutar <= 0 -> {tutar_f}")
                continue

            # daire yoksa oluştur
            if daire_no not in dmap:
                ad = ""
                tel = ""
                aidat_val = 0.0
                aktif_val = 1

                if col_ad is not None and col_ad < len(rowvals) and rowvals[col_ad] is not None:
                    ad = str(rowvals[col_ad]).strip()
                if not ad:
                    ad = f"Daire {daire_no}"

                if col_tel is not None and col_tel < len(rowvals) and rowvals[col_tel] is not None:
                    tel = str(rowvals[col_tel]).strip()

                if col_aidat is not None and col_aidat < len(rowvals) and rowvals[col_aidat] is not None:
                    try:
                        aidat_val = float(safe_float(str(rowvals[col_aidat])))
                    except Exception:
                        aidat_val = 0.0

                if col_aktif is not None and col_aktif < len(rowvals) and rowvals[col_aktif] is not None:
                    a = str(rowvals[col_aktif]).strip().lower()
                    aktif_val = 1 if a in ("1", "true", "evet", "yes", "aktif") else 0

                try:
                    con.execute(
                        "INSERT INTO daireler(daire_no, ad_soyad, telefon, aidat, aktif) VALUES(?,?,?,?,?)",
                        (daire_no, ad, tel, float(aidat_val), int(aktif_val))
                    )
                    daire_id = con.execute("SELECT id FROM daireler WHERE daire_no=?", (daire_no,)).fetchone()[0]
                    dmap[daire_no] = int(daire_id)
                    created_flats += 1

                    # ilk sakin
                    con.execute(
                        "INSERT INTO sakinler(daire_id, ad_soyad, telefon, baslangic_tarihi, bitis_tarihi) "
                        "VALUES(?,?,?,?,NULL)",
                        (int(daire_id), ad, tel, iso_today())
                    )
                except Exception as e:
                    errors += 1
                    msgs.append(f"Satır {r}: daire oluşturulamadı ({daire_no}) -> {e}")
                    continue
            else:
                # opsiyonel bilgileri doluysa güncelle (boşu ezmeyelim)
                daire_id = dmap[daire_no]
                new_ad = None
                new_tel = None
                new_aidat = None
                new_aktif = None

                if col_ad is not None and col_ad < len(rowvals) and rowvals[col_ad] is not None:
                    t = str(rowvals[col_ad]).strip()
                    if t:
                        new_ad = t
                if col_tel is not None and col_tel < len(rowvals) and rowvals[col_tel] is not None:
                    t = str(rowvals[col_tel]).strip()
                    if t:
                        new_tel = t
                if col_aidat is not None and col_aidat < len(rowvals) and rowvals[col_aidat] is not None:
                    try:
                        new_aidat = float(safe_float(str(rowvals[col_aidat])))
                    except Exception:
                        new_aidat = None
                if col_aktif is not None and col_aktif < len(rowvals) and rowvals[col_aktif] is not None:
                    a = str(rowvals[col_aktif]).strip().lower()
                    new_aktif = 1 if a in ("1", "true", "evet", "yes", "aktif") else 0

                if any(v is not None for v in (new_ad, new_tel, new_aidat, new_aktif)):
                    old_ad, old_tel, old_aidat, old_aktif = con.execute(
                        "SELECT ad_soyad, telefon, aidat, aktif FROM daireler WHERE id=?",
                        (daire_id,)
                    ).fetchone()
                    ad_final = new_ad if new_ad is not None else old_ad
                    tel_final = new_tel if new_tel is not None else old_tel
                    aidat_final = float(new_aidat) if new_aidat is not None else float(old_aidat)
                    aktif_final = int(new_aktif) if new_aktif is not None else int(old_aktif)
                    con.execute(
                        "UPDATE daireler SET ad_soyad=?, telefon=?, aidat=?, aktif=? WHERE id=?",
                        (ad_final, tel_final, aidat_final, aktif_final, daire_id)
                    )
                    # güncel sakin sync (tarihler sabit)
                    self._ensure_current_sakin(int(daire_id))
                    cur = con.execute(
                        "SELECT id FROM sakinler WHERE daire_id=? AND bitis_tarihi IS NULL ORDER BY id DESC LIMIT 1",
                        (int(daire_id),)
                    ).fetchone()
                    if cur:
                        con.execute("UPDATE sakinler SET ad_soyad=?, telefon=? WHERE id=?",
                                    (ad_final, tel_final, int(cur[0])))
                    updated_flats += 1

            # tahakkuk upsert
            created_at = iso_today()
            try:
                row_id = con.execute(
                    "SELECT id FROM tahakkuk WHERE donem=? AND daire_id=?",
                    (donem, int(dmap[daire_no]))
                ).fetchone()

                if row_id:
                    con.execute("UPDATE tahakkuk SET tutar=?, created_at=? WHERE id=?",
                                (float(tutar_f), created_at, int(row_id[0])))
                else:
                    con.execute(
                        "INSERT INTO tahakkuk(donem, daire_id, tutar, created_at) VALUES(?,?,?,?)",
                        (donem, int(dmap[daire_no]), float(tutar_f), created_at)
                    )
                ok += 1
            except Exception as e:
                errors += 1
                msgs.append(f"Satır {r}: DB hatası -> {e}")

        con.commit()
        con.close()

        self.refresh_all()

        text = (
            "İçe aktarma tamamlandı.\n"
            f"Borç kaydı (tahakkuk): {ok}\n"
            f"Oluşturulan daire: {created_flats}\n"
            f"Güncellenen daire: {updated_flats}\n"
            f"Atlandı: {skipped}\n"
            f"Hata: {errors}"
        )
        if msgs:
            text += "\n\nDetay (ilk 15):\n" + "\n".join(msgs[:15])
        QMessageBox.information(self, "Sonuç", text)
    # --- payments (refactored) ---
    def _validate_payment_inputs(self) -> tuple:
        """
        Ödeme girdilerini valide et.
        Return: (valid, daire_id, donem0, ay_sayisi, yontem, tarih_iso, acik, tutar_raw)
        """
        donem0 = self.o_in_donem.text().strip()
        if not validate_period(donem0):
            QMessageBox.warning(self, "Uyarı", "Dönem formatı YYYY-MM olmalı.")
            return (False, None, None, None, None, None, None, None)
    
        daire_id = self.o_cmb_daire.currentData()
        if not daire_id:
            QMessageBox.warning(self, "Uyarı", "Daire seçin.")
            return (False, None, None, None, None, None, None, None)
    
        tek_donem = (self.o_cmb_mode.currentText() == "Tek Dönem Borç Kapatma")
        ay_sayisi = 1 if tek_donem else int(self.o_sp_ay_sayisi.value())
    
        yontem = self.o_cmb_yontem.currentText()
        tarih_iso = self.o_dt_tarih.date().toPython().isoformat()
        acik = self.o_in_acik.text().strip()
        tutar_raw = self.o_in_tutar.text().strip()
    
        return (True, daire_id, donem0, ay_sayisi, yontem, tarih_iso, acik, tutar_raw)

    def _calculate_payment_amount(self, daire_id: int, donem0: str, tek_donem: bool, 
                             ay_sayisi: int, tutar_raw: str) -> tuple:
        """
        Ödeme tutarını hesapla ve ödeme kaydı listesini oluştur.
        Return: (valid, total, payment_rows) 
        payment_rows = [(donem, tutar), ...]
        """
        con = connect()
    
        if tek_donem:
            # TEK DÖNEM MOD: Seçili dönemin bakiyesi
            borc, odeme, bakiye = self._balance_for_daire_period(int(daire_id), donem0)
        
            if borc <= 0:
                con.close()
                QMessageBox.warning(self, "Uyarı", f"{donem0} dönemi için tahakkuk bulunamadı.")
                return (False, 0, [])
        
            # Tutar belirle
            if not tutar_raw:
                total = round(max(bakiye, 0.0), 2)
            else:
                try:
                    total = round(safe_float(tutar_raw), 2)
                except Exception:
                    con.close()
                    QMessageBox.warning(self, "Uyarı", "Toplam tutar sayısal olmalı.")
                    return (False, 0, [])
        
            if total <= 0:
                con.close()
                QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
                return (False, 0, [])
        
            if total > bakiye + 1e-9:
                con.close()
                QMessageBox.warning(self, "Uyarı", f"Fazla ödeme kabul edilmiyor. Kalan: {bakiye:.2f} TL")
                return (False, 0, [])
        
            con.close()
            return (True, total, [(donem0, total)])
    
        else:
            # PEŞİN ÇOK DÖNEM MOD: Ay sayısı kadar dönemi kaplasın
            d = con.execute(
                "SELECT aidat FROM daireler WHERE id=?",
                (int(daire_id),)
            ).fetchone()
            if not d:
                con.close()
                QMessageBox.warning(self, "Uyarı", "Daire bulunamadı.")
                return (False, 0, [])
        
            aidat = float(d[0])
        
            # Tutar belirle
            if not tutar_raw:
                total = round(aidat * ay_sayisi, 2)
            else:
                try:
                    total = round(safe_float(tutar_raw), 2)
                except Exception:
                    con.close()
                    QMessageBox.warning(self, "Uyarı", "Toplam tutar sayısal olmalı.")
                    return (False, 0, [])
        
            if total <= 0:
                con.close()
                QMessageBox.warning(self, "Uyarı", "Toplam tutar 0'dan büyük olmalı.")
                return (False, 0, [])
        
            # Ay bazında bölüştür (son aya kalan fark ekle)
            per = round(total / ay_sayisi, 2)
            amounts = [per] * ay_sayisi
            fix = round(total - sum(amounts), 2)
            amounts[-1] = round(amounts[-1] + fix, 2)
        
            payment_rows = []
            for i in range(ay_sayisi):
                donem = add_months(donem0, i)
                payment_rows.append((donem, amounts[i]))
        
            con.close()
            return (True, total, payment_rows)
    def load_payment_history(self):
        """Seçili dairenin geçmiş ödemelerini yükle"""
        daire_id = self.o_cmb_daire.currentData()
    
        self.tbl_odeme_gecmis.setRowCount(0)
    
        if not daire_id:
            return
    
        con = connect()
        rows = con.execute("""
            SELECT id, donem, tarih, tutar, yontem, makbuz_no
            FROM odemeler
            WHERE daire_id = ?
            ORDER BY tarih DESC, donem DESC
        """, (int(daire_id),)).fetchall()
        con.close()
    
        for oid, donem, tarih, tutar, yontem, makbuz in rows:
            row = self.tbl_odeme_gecmis.rowCount()
            self.tbl_odeme_gecmis.insertRow(row)
        
            vals = [str(oid), donem, tarih, f"{float(tutar):.2f}", yontem, makbuz or ""]
        
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                if col == 3:  # Tutar
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item.setBackground(QColor(240, 248, 255))  # Açık mavi
                self.tbl_odeme_gecmis.setItem(row, col, item)

    def open_payment_for_edit(self, index):
        """Çift tıklanan ödemeyi düzenleme penceresinde aç"""
        row = index.row()
        col = index.column()
    
        try:
            payment_id = int(self.tbl_odeme_gecmis.item(row, 0).text())
        except Exception:
            QMessageBox.warning(self, "Hata", "Ödeme ID'si alınamadı.")
            return
    
        # Modal dialog aç
        dialog = EditPaymentDialog(payment_id, self)
        dialog.exec()
    
        # Dialog kapandıktan sonra tabloyu yenile
        self.load_payment_history()
        self.refresh_all()    

    def _save_payment_records(self, daire_id: int, payment_rows: list, makbuz_no: str, 
                         tarih_iso: str, yontem: str, acik: str) -> bool:
        """
        Ödeme kayıtlarını DB'ye kaydet.
        payment_rows = [(donem, tutar), ...]
        """
        con = connect()
        try:
            # Peşin modda (i > 0) otomatik tahakkuk oluştur
            for i, (donem, amount) in enumerate(payment_rows):
                if i > 0:  # Peşin mode - gelecek dönemler için tahakkuk yok, oluştur
                    d = con.execute(
                        "SELECT aidat FROM daireler WHERE id=?",
                        (int(daire_id),)
                    ).fetchone()
                    aidat = float(d[0]) if d else 0.0
                    try:
                        con.execute(
                            "INSERT INTO tahakkuk(donem, daire_id, tutar, created_at) VALUES(?,?,?,?)",
                            (donem, int(daire_id), aidat, iso_today())
                        )
                    except sqlite3.IntegrityError:
                        pass  # Zaten var, atla
            
                # Ödeme kaydı ekle
                con.execute("""
                    INSERT INTO odemeler(donem, tarih, daire_id, tutar, yontem, makbuz_no, aciklama)
                    VALUES(?,?,?,?,?,?,?)
                """, (donem, tarih_iso, int(daire_id), float(amount), yontem, makbuz_no, acik))
        
            con.commit()
            con.close()
            return True
        except Exception as e:
            con.close()
            QMessageBox.warning(self, "Hata", f"DB'ye kaydetme hatası:\n{e}")
            return False

    def _get_all_debt_periods(self, daire_id: int) -> list:
        """
        Dairenin tüm borçlu dönemlerini al.
        Return: [(donem, kalan_tutar), ...] (sadece kalan > 0)
        """
        con = connect()
        rows = con.execute("""
            SELECT d.donem,
                   COALESCE(b.borc, 0) - COALESCE(o.odeme, 0) as kalan
            FROM (
                SELECT donem FROM tahakkuk
                UNION
                SELECT donem FROM odemeler
            ) d
            LEFT JOIN (
            SELECT donem, SUM(tutar) as borc
            FROM tahakkuk WHERE daire_id=?
            GROUP BY donem
        ) b ON d.donem = b.donem
        LEFT JOIN (
                SELECT donem, SUM(tutar) as odeme
                FROM odemeler WHERE daire_id=?
                GROUP BY donem
            ) o ON d.donem = o.donem
            WHERE (COALESCE(b.borc, 0) - COALESCE(o.odeme, 0)) > 0.00001
            ORDER BY d.donem
        """, (int(daire_id), int(daire_id))).fetchall()
        con.close()
        return rows    

        # --- payments ---
   
    def odeme_save_multi(self):
        """Ana ödeme kaydet fonksiyonu"""
        # Girdileri valide et
        valid, daire_id, donem0, ay_sayisi, yontem, tarih_iso, acik, tutar_raw = self._validate_payment_inputs()
        if not valid:
            return
    
        # Daire bilgilerini al
        con = connect()
        d = con.execute(
            "SELECT daire_no, ad_soyad, telefon FROM daireler WHERE id=?",
            (int(daire_id),)
        ).fetchone()
        con.close()
    
        if not d:
            QMessageBox.warning(self, "Uyarı", "Daire bulunamadı.")
            return
    
        daire_no, ad, tel = d
        tek_donem = (self.o_cmb_mode.currentText() == "Tek Dönem Borç Kapatma")
    
        # Tutarı hesapla
        valid, total, payment_rows = self._calculate_payment_amount(
            int(daire_id), donem0, tek_donem, ay_sayisi, tutar_raw
        )
        if not valid:
            return
    
        # Makbuz no oluştur
        makbuz_no = receipt_next_for_period(donem0)
    
        # DB'ye kaydet
        if not self._save_payment_records(int(daire_id), payment_rows, makbuz_no, tarih_iso, yontem, acik):
            return
    
        # PDF makbuz oluştur ve aç
        pdf_path = None
        try:
            pdf_path = self.create_receipt_pdf(
                makbuz_no=makbuz_no,
                tarih_iso=tarih_iso,
                daire_no=str(daire_no),
                ad=str(ad),
                yontem=str(yontem),
                kalemler=payment_rows,
                aciklama=acik
        )
        except Exception as e:
            QMessageBox.warning(self, "Uyarı", f"Ödeme kaydedildi ama PDF üretilemedi:\n{e}")
    
        # Kullanıcıya bildir ve PDF aç
        msg = f"Ödeme kaydedildi.\nMakbuz No: {makbuz_no}"
        if pdf_path:
            msg += "\n\nMakbuzu açmak ister misiniz?"
            reply = QMessageBox.question(
                self, "Ödeme Kaydedildi", msg,
                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes
            )
            if reply == QMessageBox.Yes:
                webbrowser.open(Path(pdf_path).resolve().as_uri())
        else:
            QMessageBox.information(self, "OK", msg)
    
        # Formu temizle ve yenile
        self.o_in_tutar.clear()
        self.o_in_acik.clear()
        self.o_sp_ay_sayisi.setValue(1)
        self.refresh_all() 


    def create_receipt_pdf(self, makbuz_no: str, tarih_iso: str, daire_no: str, ad: str,
                           yontem: str, kalemler: list, aciklama: str):
        out_dir = Path.cwd() / "makbuzlar"
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / f"makbuz_{makbuz_no}.pdf"

        c = canvas.Canvas(str(out_path), pagesize=A4)
        w, h = A4
        y = h - 60

        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, y, "AİDAT MAKBUZU")
        y -= 25

        c.setFont("Helvetica", 11)
        c.drawString(50, y, f"Makbuz No: {makbuz_no}")
        c.drawString(300, y, f"Tarih: {tarih_iso}")
        y -= 18
        c.drawString(50, y, f"Daire: {daire_no}")
        c.drawString(300, y, f"Ad Soyad: {ad}")
        y -= 18
        c.drawString(50, y, f"Ödeme Yöntemi: {yontem}")
        y -= 25

        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, y, "Dönem")
        c.drawString(150, y, "Tutar (TL)")
        y -= 12
        c.line(50, y, 250, y)
        y -= 18

        total = 0.0
        c.setFont("Helvetica", 11)
        for donem, tutar in kalemler:
            c.drawString(50, y, str(donem))
            c.drawRightString(230, y, f"{float(tutar):.2f}")
            total += float(tutar)
            y -= 16
            if y < 120:
                c.showPage()
                y = h - 60

        y -= 10
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "TOPLAM")
        c.drawRightString(230, y, f"{total:.2f} TL")
        y -= 25

        if aciklama:
            c.setFont("Helvetica", 10)
            c.drawString(50, y, f"Açıklama: {aciklama[:90]}")
            y -= 18

        y -= 30
        c.setFont("Helvetica", 10)
        c.drawString(50, y, "İmza:")
        c.line(80, y - 2, 250, y - 2)

        c.save()
        return str(out_path)

    def delete_selected_payment(self):
        rows = self.tbl_odeme.selectionModel().selectedRows()
        if not rows:
            QMessageBox.warning(self, "Uyarı", "Silmek için tablodan bir ödeme seçin.")
            return
        r = rows[0].row()
        pid = int(self.tbl_odeme.item(r, 0).text())
        if QMessageBox.question(self, "Onay", "Seçili ödemeyi silmek istiyor musunuz?") != QMessageBox.Yes:
            return
        con = connect()
        con.execute("DELETE FROM odemeler WHERE id=?", (pid,))
        con.commit()
        con.close()
        self.refresh_all()

    # --- gider ---
    def gider_save(self):
        donem = self.g_in_donem.text().strip()
        if not validate_period(donem):
            QMessageBox.warning(self, "Uyarı", "Dönem formatı YYYY-MM olmalı.")
            return

        tarih_iso = self.g_dt_tarih.date().toPython().isoformat()
        kat = self.g_cmb_kat.currentText()
        try:
            tutar = safe_float(self.g_in_tutar.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Tutar sayısal olmalı.")
            return
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        yontem = self.g_cmb_yontem.currentText()
        acik = self.g_in_acik.text().strip()

        con = connect()
        con.execute("""
            INSERT INTO giderler(donem, tarih, kategori, tutar, yontem, aciklama)
            VALUES(?,?,?,?,?,?)
        """, (donem, tarih_iso, kat, float(tutar), yontem, acik))
        con.commit()
        con.close()

        QMessageBox.information(self, "OK", "Gider kaydedildi.")
        self.g_in_tutar.clear()
        self.g_in_acik.clear()
        self.refresh_all()

    def delete_selected_expense(self):
        rows = self.tbl_gider.selectionModel().selectedRows()
        if not rows:
            QMessageBox.warning(self, "Uyarı", "Silmek için tablodan bir gider seçin.")
            return
        r = rows[0].row()
        gid = int(self.tbl_gider.item(r, 0).text())
        if QMessageBox.question(self, "Onay", "Seçili gideri silmek istiyor musunuz?") != QMessageBox.Yes:
            return
        con = connect()
        con.execute("DELETE FROM giderler WHERE id=?", (gid,))
        con.commit()
        con.close()
        self.refresh_all()

    # --- announcements ---
    def duyuru_save(self):
        tarih_iso = self.duy_dt.date().toPython().isoformat()
        baslik = self.duy_baslik.text().strip()
        mesaj = self.duy_mesaj.toPlainText().strip()
        if not baslik or not mesaj:
            QMessageBox.warning(self, "Uyarı", "Başlık ve mesaj zorunlu.")
            return
        con = connect()
        con.execute("INSERT INTO duyurular(tarih, baslik, mesaj) VALUES(?,?,?)",
                    (tarih_iso, baslik, mesaj))
        con.commit()
        con.close()
        QMessageBox.information(self, "OK", "Duyuru kaydedildi.")
        self.duy_baslik.clear()
        self.duy_mesaj.clear()
        self.refresh_all()

    def duyuru_copy(self):
        baslik = self.duy_baslik.text().strip()
        mesaj = self.duy_mesaj.toPlainText().strip()
        if not baslik and not mesaj:
            QMessageBox.information(self, "Bilgi", "Kopyalanacak metin yok.")
            return
        text = (baslik + "\n\n" + mesaj).strip()
        QApplication.clipboard().setText(text)
        QMessageBox.information(self, "OK", "Metin panoya kopyalandı.")

    def duyuru_delete_selected(self):
        rows = self.tbl_duyuru.selectionModel().selectedRows()
        if not rows:
            QMessageBox.warning(self, "Uyarı", "Silmek için tablodan bir duyuru seçin.")
            return
        r = rows[0].row()
        did = int(self.tbl_duyuru.item(r, 0).text())
        if QMessageBox.question(self, "Onay", "Seçili duyuru silinsin mi?") != QMessageBox.Yes:
            return
        con = connect()
        con.execute("DELETE FROM duyurular WHERE id=?", (did,))
        con.commit()
        con.close()
        self.refresh_all()

    # --- WhatsApp ---
    def open_whatsapp_for_payment_tab(self):
        donem = self.o_in_donem.text().strip()
        daire_id = self.o_cmb_daire.currentData()
        if not daire_id:
            QMessageBox.warning(self, "Uyarı", "Daire seçin.")
            return
        self._open_whatsapp_for_daire(int(daire_id), donem)

    def report_whatsapp_selected(self):
        donem = self.r_in_donem.text().strip()
        rows = self.tbl_report.selectionModel().selectedRows()
        if not rows:
            QMessageBox.warning(self, "Uyarı", "Rapor tablosundan bir daire seçin.")
            return
        r = rows[0].row()
        daire_no = self.tbl_report.item(r, 0).text()
        con = connect()
        row = con.execute("SELECT id FROM daireler WHERE daire_no=?", (daire_no,)).fetchone()
        con.close()
        if not row:
            QMessageBox.warning(self, "Uyarı", "Daire bulunamadı.")
            return
        self._open_whatsapp_for_daire(int(row[0]), donem)
        
    def _open_whatsapp_for_daire(self, daire_id: int, donem: str = None):
        """WhatsApp mesaj sayfasını aç"""
        con = connect()
        d = con.execute(
            "SELECT daire_no, ad_soyad, telefon FROM daireler WHERE id=?",
            (int(daire_id),)
        ).fetchone()
        con.close()
    
        if not d:
            QMessageBox.warning(self, "Uyarı", "Daire bulunamadı.")
            return
    
        daire_no, ad, tel = d
    
        if not tel:
            QMessageBox.information(self, "Bilgi", "Bu daire için telefon kayıtlı değil.")
            return
    
        # Borçlu dönemleri al
        debt_periods = self._get_all_debt_periods(int(daire_id))
    
        if not debt_periods:
            QMessageBox.information(self, "Bilgi", "Bu dairenin borcu yok.")
            return
    
        # Mesaj oluştur
        satirlar = [f"{d[0]} dönemi {d[1]:,.2f} TL" for d in debt_periods]
        toplam = sum(d[1] for d in debt_periods)
    
        # Türkçe format (virgül)
        satir_text = "\n".join(satirlar).replace(",", "X").replace(".", ",").replace("X", ".")
        toplam_text = f"{toplam:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    
        mesaj = f"""Sayın {ad} (Daire {daire_no})

    {satir_text}

    Toplam {toplam_text} TL borcunuz bulunmaktadır.

    Bilginize."""
    
        # Telefon numarasını düzelt (90 formatı)
        digits = "".join([c for c in tel if c.isdigit()])
        if digits.startswith("0"):
            digits = digits[1:]
        if not digits.startswith("90"):
            digits = "90" + digits
    
        # WhatsApp aç
        url = "https://wa.me/" + digits + "?text=" + quote(mesaj)
        webbrowser.open(url)
        


    def copy_bulk_message_text(self):
        donem = self.r_in_donem.text().strip()
        if not validate_period(donem):
            QMessageBox.warning(self, "Uyarı", "Dönem formatı YYYY-MM olmalı.")
            return

        rows = self._report_rows(donem)
        debtors = [r for r in rows if r["toplam_bakiye"] > 0.00001]
        if not debtors:
            QMessageBox.information(self, "Bilgi", "Toplam borcu olan daire yok.")
            return

        vade = compute_vade_date(donem).isoformat()
        msg = f"{donem} dönemi aidat hatırlatması (Vade: {vade}).\n\n"
        for r in debtors:
            msg += (
                f"- Daire {r['daire_no']} {r['ad']}: "
                f"Dönem Kalan {r['period_bakiye']:.2f} TL | "
                f"Toplam Kalan {r['toplam_bakiye']:.2f} TL ({r['durum']})\n"
            )
        msg += "\nTeşekkürler."
        QApplication.clipboard().setText(msg)
        QMessageBox.information(self, "OK", "Toplu mesaj metni panoya kopyalandı.")

    # ---------------- refresh/report ----------------
    def refresh_all(self):
        d = self.in_donem.text().strip()

        if validate_period(d):
            if hasattr(self, "t_in_from") and not self.t_in_from.text().strip():
                self.t_in_from.setText(d)
            if hasattr(self, "t_in_to") and not self.t_in_to.text().strip():
                self.t_in_to.setText(d)
            if hasattr(self, "g_in_donem"):
                self.g_in_donem.setText(d)
            if hasattr(self, "r_in_donem"):
                self.r_in_donem.setText(d)
            if hasattr(self, "o_in_donem") and not validate_period(self.o_in_donem.text().strip()):
                self.o_in_donem.setText(d)

        self.refresh_daire_table()
        self.refresh_daire_combo_all()
        self.refresh_tahakkuk_table()
        self.refresh_payments_table()
        self.refresh_expenses_table()
        self.load_expense_history()  # ⭐ YENİ: Geçmiş giderleri yükle
        self.refresh_duyuru_table()
        self.refresh_report()

        if hasattr(self, "o_cmb_mode"):
            self._set_payment_mode_ui()
            self.update_payment_period_balance()

        did = self._selected_daire_id()
        if did:
            self._ensure_current_sakin(did)
            self.refresh_sakin_history_table(did)
            self.refresh_selected_summary(did)    


    def refresh_daire_table(self):
        con = connect()
        rows = con.execute("""
            SELECT id, daire_no, ad_soyad, telefon, aidat, aktif
            FROM daireler
            ORDER BY CAST(daire_no AS INT), daire_no
        """).fetchall()
        con.close()

        self.tbl_daire.setRowCount(0)
        for r in rows:
            row = self.tbl_daire.rowCount()
            self.tbl_daire.insertRow(row)
            for c, val in enumerate(r):
                it = QTableWidgetItem(str(val))
                if c in (0, 5):
                    it.setTextAlignment(Qt.AlignCenter)
                if c == 4:
                    it.setText(f"{float(val):.2f}")
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tbl_daire.setItem(row, c, it)
    def refresh_daire_combo_all(self):
        con = connect()
        rows = con.execute("""
            SELECT id, daire_no, ad_soyad, aktif
            FROM daireler
            ORDER BY CAST(daire_no AS INT), daire_no
        """).fetchall()
        con.close()

        selected_t = self.t_cmb_daire.currentData() if hasattr(self, "t_cmb_daire") else None
        selected_o = self.o_cmb_daire.currentData() if hasattr(self, "o_cmb_daire") else None
        selected_e = self.e_cmb_daire.currentData() if hasattr(self, "e_cmb_daire") else None

        if hasattr(self, "t_cmb_daire"):
            self.t_cmb_daire.blockSignals(True)
            self.t_cmb_daire.clear()
            self.t_cmb_daire.addItem("-- seç --", None)
            t_idx = 0
            for did, dno, ad, aktif in rows:
                tag = "" if int(aktif) == 1 else " (pasif)"
                self.t_cmb_daire.addItem(f"Daire {dno} - {ad}{tag}", did)
                if selected_t is not None and int(did) == int(selected_t):
                    t_idx = self.t_cmb_daire.count() - 1
            self.t_cmb_daire.setCurrentIndex(t_idx)
            self.t_cmb_daire.blockSignals(False)

        if hasattr(self, "o_cmb_daire"):
            self.o_cmb_daire.blockSignals(True)
            self.o_cmb_daire.clear()
            self.o_cmb_daire.addItem("-- seç --", None)
            o_idx = 0
            for did, dno, ad, aktif in rows:
                tag = "" if int(aktif) == 1 else " (pasif)"
                self.o_cmb_daire.addItem(f"Daire {dno} - {ad}{tag}", did)
                if selected_o is not None and int(did) == int(selected_o):
                    o_idx = self.o_cmb_daire.count() - 1
            self.o_cmb_daire.setCurrentIndex(o_idx)
            self.o_cmb_daire.blockSignals(False)

        if hasattr(self, "e_cmb_daire"):
            self.e_cmb_daire.blockSignals(True)
            self.e_cmb_daire.clear()
            self.e_cmb_daire.addItem("-- seç --", None)
            e_idx = 0
            for did, dno, ad, aktif in rows:
                tag = "" if int(aktif) == 1 else " (pasif)"
                self.e_cmb_daire.addItem(f"Daire {dno} - {ad}{tag}", did)
                if selected_e is not None and int(did) == int(selected_e):
                    e_idx = self.e_cmb_daire.count() - 1
            self.e_cmb_daire.setCurrentIndex(e_idx)
            self.e_cmb_daire.blockSignals(False)

        if hasattr(self, "o_lbl_bakiye"):
            self.update_payment_period_balance()
        # Kasa raporunu da yenile
        if hasattr(self, "btn_k_refresh"):
            self.refresh_kasa_report()    
    def refresh_tahakkuk_table(self):
        con = connect()
        rows = con.execute("""
            SELECT t.id,
                   d.daire_no || ' - ' || d.ad_soyad AS daire,
                   t.donem,
                   t.tutar,
                   t.created_at
            FROM tahakkuk t
            JOIN daireler d ON d.id = t.daire_id
            ORDER BY t.donem DESC, CAST(d.daire_no AS INT), d.daire_no
        """).fetchall()
        con.close()

        self.tbl_tah.setRowCount(0)
        for rec in rows:
            row = self.tbl_tah.rowCount()
            self.tbl_tah.insertRow(row)
            for c, val in enumerate(rec):
                it = QTableWidgetItem(str(val))
                if c == 0:
                    it.setTextAlignment(Qt.AlignCenter)
                if c == 3:
                    it.setText(f"{float(val):.2f}")
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tbl_tah.setItem(row, c, it)


    def refresh_payments_table(self):
        donem = self.o_in_donem.text().strip()
        if not validate_period(donem):
            self.tbl_odeme.setRowCount(0)
            return

        con = connect()
        rows = con.execute("""
            SELECT o.id, o.donem, o.tarih, d.daire_no, o.tutar, o.yontem, o.makbuz_no
            FROM odemeler o
            JOIN daireler d ON d.id=o.daire_id
            WHERE o.donem=?
            ORDER BY o.id DESC
        """, (donem,)).fetchall()
        con.close()

        self.tbl_odeme.setRowCount(0)
        for r in rows:
            row = self.tbl_odeme.rowCount()
            self.tbl_odeme.insertRow(row)
            for c, val in enumerate(r):
                it = QTableWidgetItem(str(val))
                if c in (0, 3):
                    it.setTextAlignment(Qt.AlignCenter)
                if c == 4:
                    it.setText(f"{float(val):.2f}")
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tbl_odeme.setItem(row, c, it)

    def refresh_expenses_table(self):
        donem = self.g_in_donem.text().strip()
        if not validate_period(donem):
            self.tbl_gider.setRowCount(0)
            return

        con = connect()
        rows = con.execute("""
            SELECT id, donem, tarih, kategori, tutar, yontem, aciklama
            FROM giderler
            WHERE donem=?
            ORDER BY id DESC
        """, (donem,)).fetchall()
        con.close()

        self.tbl_gider.setRowCount(0)
        for r in rows:
            row = self.tbl_gider.rowCount()
            self.tbl_gider.insertRow(row)
            for c, val in enumerate(r):
                it = QTableWidgetItem(str(val))
                if c == 0:
                    it.setTextAlignment(Qt.AlignCenter)
                if c == 4:
                    it.setText(f"{float(val):.2f}")
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tbl_gider.setItem(row, c, it)

    def refresh_duyuru_table(self):
        con = connect()
        rows = con.execute("""
            SELECT id, tarih, baslik, mesaj
            FROM duyurular
            ORDER BY id DESC
        """).fetchall()
        con.close()

        self.tbl_duyuru.setRowCount(0)
        for did, t, b, m in rows:
            row = self.tbl_duyuru.rowCount()
            self.tbl_duyuru.insertRow(row)
            self.tbl_duyuru.setItem(row, 0, QTableWidgetItem(str(did)))
            self.tbl_duyuru.setItem(row, 1, QTableWidgetItem(str(t)))
            self.tbl_duyuru.setItem(row, 2, QTableWidgetItem(str(b)))
            short = (m[:80] + "…") if len(m) > 80 else m
            self.tbl_duyuru.setItem(row, 3, QTableWidgetItem(short))



    def refresh_report(self):
        self.refresh_report_ui()

    def refresh_report_ui(self):
        donem = self.r_in_donem.text().strip()

        if not validate_period(donem):
            self.lbl_summary.setText("Özet: Dönem formatı yanlış.")
            self.tbl_report.setRowCount(0)
            return

        only_unpaid = self.r_chk_only_unpaid.isChecked()
        only_overdue = self.r_chk_only_overdue.isChecked()

        rows_all = self._report_rows(donem)
    
        if only_unpaid:
            rows = [r for r in rows_all if r["toplam_bakiye"] > 0.00001]
        elif only_overdue:
            rows = [r for r in rows_all if r["durum"] == "Gecikmiş"]
        else:
            rows = rows_all

        gelir = self._sum_payments(donem)
        gider = self._sum_expenses(donem)
        net = gelir - gider

        toplam_borc_genel = sum(r["toplam_borc"] for r in rows_all)
        toplam_odeme_genel = sum(r["toplam_odeme"] for r in rows_all)
        toplam_kalan_genel = sum(r["toplam_bakiye"] for r in rows_all)

        self.lbl_summary.setText(
            f"{donem} | Gelir: {gelir:.2f} TL | Gider: {gider:.2f} TL | Net: {net:.2f} TL "
            f"| Toplam Borç: {toplam_borc_genel:.2f} TL "
            f"| Toplam Kalan: {toplam_kalan_genel:.2f} TL"
        )

        self.tbl_report.setRowCount(0)

        for r in rows:
            row = self.tbl_report.rowCount()
            self.tbl_report.insertRow(row)

            vals = [
                r["daire_no"],
                r["ad"],
                r["tel"],
                f"{r['aidat']:.2f}",
                f"{r['period_borc']:.2f}",
                f"{r['period_odeme']:.2f}",
                f"{r['toplam_bakiye']:.2f}",
                r["durum"]
            ]

            for col, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
            
                if col in (3, 4, 5, 6):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

                # 🎨 RENK
                if r["toplam_bakiye"] > 0.00001:
                    item.setBackground(QColor(255, 200, 200))
                else:
                    item.setBackground(QColor(200, 255, 200))

                self.tbl_report.setItem(row, col, item)   
   

    def _report_rows(self, donem: str):
        con = connect()
        daireler = con.execute("""
            SELECT id, daire_no, ad_soyad, telefon, aidat, aktif
            FROM daireler
            ORDER BY CAST(daire_no AS INT), daire_no
        """).fetchall()
        con.close()

        out = []
        for did, dno, ad, tel, aidat, aktif in daireler:
            period_borc, period_odeme, period_bakiye = self._balance_for_daire_period(int(did), donem)
            toplam_borc, toplam_odeme, toplam_bakiye = self._balance_for_daire_total(int(did))
            dur = status_for_balance(donem, period_bakiye)

            out.append({
                "daire_id": int(did),
                "daire_no": dno,
                "ad": ad,
                "tel": tel or "",
                "aidat": float(aidat),
                "period_borc": float(period_borc),
                "period_odeme": float(period_odeme),
                "period_bakiye": float(period_bakiye),
                "toplam_borc": float(toplam_borc),
                "toplam_odeme": float(toplam_odeme),
                "toplam_bakiye": float(toplam_bakiye),
                "borc": float(toplam_borc),
                "odeme": float(toplam_odeme),
                "bakiye": float(toplam_bakiye),
                "durum": dur,
                "aktif": int(aktif)
            })
        return out
    def _balance_for_daire_total(self, daire_id: int):
        con = connect()
        borc = con.execute("""
            SELECT COALESCE(SUM(tutar),0)
            FROM tahakkuk
            WHERE daire_id=?
        """, (int(daire_id),)).fetchone()[0] or 0.0

        odeme = con.execute("""
            SELECT COALESCE(SUM(tutar),0)
            FROM odemeler
            WHERE daire_id=?
        """, (int(daire_id),)).fetchone()[0] or 0.0

        con.close()
        borc = float(borc)
        odeme = float(odeme)
        return borc, odeme, borc - odeme
    
    def _balance_for_daire_period(self, daire_id: int, donem: str):
        con = connect()
        borc = con.execute("""
            SELECT COALESCE(SUM(tutar),0) FROM tahakkuk
            WHERE donem=? AND daire_id=?
        """, (donem, daire_id)).fetchone()[0] or 0.0

        odeme = con.execute("""
            SELECT COALESCE(SUM(tutar),0) FROM odemeler
            WHERE donem=? AND daire_id=?
        """, (donem, daire_id)).fetchone()[0] or 0.0
        con.close()
        borc = float(borc)
        odeme = float(odeme)
        return borc, odeme, borc - odeme

    def _sum_payments(self, donem: str) -> float:
        con = connect()
        v = con.execute("SELECT COALESCE(SUM(tutar),0) FROM odemeler WHERE donem=?", (donem,)).fetchone()[0] or 0.0
        con.close()
        return float(v)

    def _sum_expenses(self, donem: str) -> float:
        con = connect()
        v = con.execute("SELECT COALESCE(SUM(tutar),0) FROM giderler WHERE donem=?", (donem,)).fetchone()[0] or 0.0
        con.close()
        return float(v)

    # ---------------- Excel export ----------------
    def refresh_kasa_report(self):
        """Kasa raporu verilerini yükle ve göster"""
        date_from = self.k_dt_from.date().toPython().isoformat()
        date_to = self.k_dt_to.date().toPython().isoformat()

        con = connect()

        # Gelirler
        gelir_rows = con.execute("""
            SELECT o.donem, o.tarih, d.daire_no, o.yontem, o.tutar
            FROM odemeler o
            JOIN daireler d ON d.id = o.daire_id
            WHERE o.tarih >= ? AND o.tarih <= ?
            ORDER BY o.tarih DESC
        """, (date_from, date_to)).fetchall()

        toplam_gelir = sum(r[4] for r in gelir_rows)
        gelir_say = len(gelir_rows)

        # Giderler
        gider_rows = con.execute("""
            SELECT donem, tarih, kategori, yontem, tutar
            FROM giderler
            WHERE tarih >= ? AND tarih <= ?
            ORDER BY tarih DESC
        """, (date_from, date_to)).fetchall()

        toplam_gider = sum(r[4] for r in gider_rows)
        gider_say = len(gider_rows)

        con.close()

        net = toplam_gelir - toplam_gider
        durum = "POZİTİF ✅" if net >= 0 else "NEGATİF ❌"
        durum_renk = "#ccffcc" if net >= 0 else "#ffcccc"

        # Özet güncelle
        gelir_text = f"{toplam_gelir:,.2f} TL".replace(",", "X").replace(".", ",").replace("X", ".")
        gider_text = f"{toplam_gider:,.2f} TL".replace(",", "X").replace(".", ",").replace("X", ".")
        net_text = f"{net:,.2f} TL".replace(",", "X").replace(".", ",").replace("X", ".")

        self.k_lbl_toplam_gelir.setText(gelir_text)
        self.k_lbl_toplam_gider.setText(gider_text)
        self.k_lbl_net.setText(net_text)
        self.k_lbl_net.setStyleSheet(f"background-color: {durum_renk}; padding: 8px; border-radius: 4px; font-weight: bold;")
        self.k_lbl_durum.setText(durum)

        gelir_enko_text = f"{gelir_say} kayıt | Toplam: " + gelir_text
        gider_enko_text = f"{gider_say} kayıt | Toplam: " + gider_text

        self.k_lbl_gelir_enko.setText(gelir_enko_text)
        self.k_lbl_gider_enko.setText(gider_enko_text)

        # Gelir tablosu
        self.tbl_k_gelir.setRowCount(0)
        for donem, tarih, daire_no, yontem, tutar in gelir_rows:
            row = self.tbl_k_gelir.rowCount()
            self.tbl_k_gelir.insertRow(row)
            vals = [donem, tarih, daire_no, yontem, f"{float(tutar):.2f}"]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                if c == 4:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item.setBackground(QColor(220, 245, 230))
                self.tbl_k_gelir.setItem(row, c, item)

        # Gider tablosu
        self.tbl_k_gider.setRowCount(0)
        for donem, tarih, kategori, yontem, tutar in gider_rows:
            row = self.tbl_k_gider.rowCount()
            self.tbl_k_gider.insertRow(row)
            vals = [donem, tarih, kategori, yontem, f"{float(tutar):.2f}"]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                if c == 4:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item.setBackground(QColor(255, 235, 235))
                self.tbl_k_gider.setItem(row, c, item)

    def export_kasa_excel(self):
        """Kasa raporunu Excel'e aktar"""
        date_from = self.k_dt_from.date().toPython().isoformat()
        date_to = self.k_dt_to.date().toPython().isoformat()

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Kasa Raporu Excel Kaydet",
            f"kasa_raporu_{date_from}_{date_to}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        con = connect()

        gelir_rows = con.execute("""
            SELECT o.donem, o.tarih, d.daire_no, o.yontem, o.tutar, o.makbuz_no
            FROM odemeler o
            JOIN daireler d ON d.id = o.daire_id
            WHERE o.tarih >= ? AND o.tarih <= ?
            ORDER BY o.tarih DESC
        """, (date_from, date_to)).fetchall()

        gider_rows = con.execute("""
            SELECT donem, tarih, kategori, yontem, tutar
            FROM giderler
            WHERE tarih >= ? AND tarih <= ?
            ORDER BY tarih DESC
        """, (date_from, date_to)).fetchall()

        con.close()

        toplam_gelir = sum(r[4] for r in gelir_rows)
        toplam_gider = sum(r[4] for r in gider_rows)
        net = toplam_gelir - toplam_gider

        # Excel oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Kasa Raporu"

        # Başlık
        ws.append(["KASA RAPORU"])
        ws.append(["Tarih Aralığı", f"{date_from} - {date_to}"])
        ws.append([])
        ws.append(["Toplam Gelir (Tahsilat)", toplam_gelir])
        ws.append(["Toplam Gider", toplam_gider])
        ws.append(["NET KALAN", net])
        ws.append([])

        # Gelirler
        ws.append(["GELIRLER (ÖDEMELER)"])
        ws.append(["Dönem", "Tarih", "Daire", "Yöntem", "Tutar", "Makbuz No"])
        for donem, tarih, dno, yontem, tutar, makbuz in gelir_rows:
            ws.append([donem, tarih, dno, yontem, tutar, makbuz])
        ws.append([])

        # Giderler
        ws.append(["GİDERLER"])
        ws.append(["Dönem", "Tarih", "Kategori", "Yöntem", "Tutar"])
        for donem, tarih, kat, yontem, tutar in gider_rows:
            ws.append([donem, tarih, kat, yontem, tutar])

        autosize_worksheet(ws)
        wb.save(path)
        QMessageBox.information(self, "OK", f"Kasa raporu Excel oluşturuldu:\n{path}")
    
    def export_excel(self):
        donem = self.r_in_donem.text().strip()
        if not validate_period(donem):
            QMessageBox.warning(self, "Uyarı", "Dönem formatı YYYY-MM olmalı.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Excel Kaydet",
            f"aidat_ozet_{donem}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Özet"

        gelir = self._sum_payments(donem)
        gider = self._sum_expenses(donem)
        net = gelir - gider
        vade = compute_vade_date(donem).isoformat()

        ws.append(["Dönem", donem])
        ws.append(["Vade", vade])
        ws.append(["Gelir", gelir])
        ws.append(["Gider", gider])
        ws.append(["Net", net])
        ws.append([])
        ws.append(["Daire", "İsim", "Telefon", "Borç (Dönem)", "Ödeme (Dönem)", "Kalan (Dönem)", "Toplam Borç", "Toplam Ödeme", "Toplam Kalan", "Durum", "Aktif"])

        report_rows = self._report_rows(donem)
        for r in report_rows:
            ws.append([
                r["daire_no"], r["ad"], r["tel"],
                r["period_borc"], r["period_odeme"], r["period_bakiye"],
                r["toplam_borc"], r["toplam_odeme"], r["toplam_bakiye"],
                r["durum"], r["aktif"]
            ])
        autosize_worksheet(ws)

        ws2 = wb.create_sheet("Ödemeler")
        ws2.append(["ID", "Dönem", "Tarih", "Daire", "Tutar", "Yöntem", "Makbuz", "Açıklama"])
        con = connect()
        pays = con.execute("""
            SELECT o.id, o.donem, o.tarih, d.daire_no, o.tutar, o.yontem, o.makbuz_no, o.aciklama
            FROM odemeler o
            JOIN daireler d ON d.id=o.daire_id
            WHERE o.donem=?
            ORDER BY o.id
        """, (donem,)).fetchall()
        for p in pays:
            ws2.append(list(p))
        autosize_worksheet(ws2)

        ws3 = wb.create_sheet("Giderler")
        ws3.append(["ID", "Dönem", "Tarih", "Kategori", "Tutar", "Yöntem", "Açıklama"])
        exps = con.execute("""
            SELECT id, donem, tarih, kategori, tutar, yontem, aciklama
            FROM giderler
            WHERE donem=?
            ORDER BY id
        """, (donem,)).fetchall()
        for e in exps:
            ws3.append(list(e))
        autosize_worksheet(ws3)

        ws4 = wb.create_sheet("Duyurular")
        ws4.append(["ID", "Tarih", "Başlık", "Mesaj"])
        drows = con.execute("SELECT id, tarih, baslik, mesaj FROM duyurular ORDER BY id DESC").fetchall()
        con.close()
        for r in drows:
            ws4.append(list(r))
        autosize_worksheet(ws4)

        wb.save(path)
        QMessageBox.information(self, "OK", f"Excel oluşturuldu:\n{path}")


def main():
    app = QApplication(sys.argv)
    w = ApartmanAidatApp()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()



