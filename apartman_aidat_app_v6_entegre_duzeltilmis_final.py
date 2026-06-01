# -*- coding: utf-8 -*-
"""
Apartman Aidat Sistemi v6 - Entegre Versiyon - TAMAMLANMIŞ
Temiz Ödeme Sistemi ile

Özellikler:
✅ Daire yönetimi (CRUD, aktif/pasif, devir)
✅ Sakin geçmişi (dönem bazlı)
✅ Tahakkuk (dönem yazma, Excel import)
✅ TEMIZ ÖDEME SİSTEMİ (checkbox tabanlı, çoklu dönem seçimi)
✅ Gider yönetimi
✅ Duyuru sistemi
✅ Rapor (borçlu/gecikmiş filtre)
✅ Kasa raporu (YENI)
✅ Ekstre (daire bazlı - YENI)
✅ Geçmiş ödeme tablosu (YENI)
✅ WhatsApp ödeme bildirimi (YENI)
"""

import sys
import sqlite3
from pathlib import Path
from datetime import date, timedelta
import webbrowser
from urllib.parse import quote

from PySide6.QtCore import Qt, QDate
from PySide6.QtWidgets import (
    QApplication, QWidget, QTabWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QMessageBox, QComboBox, QGroupBox, QFileDialog, QSpinBox, QCheckBox,
    QTextEdit, QDateEdit, QSplitter, QHeaderView, QScrollArea, QDialog
)
from PySide6.QtGui import QColor

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


DB_PATH = Path("apartman_aidat.db")


# ============ HELPER FUNCTIONS ============

def connect():
    con = sqlite3.connect(str(DB_PATH))
    con.execute("PRAGMA foreign_keys = ON;")
    return con


def iso_today() -> str:
    return date.today().isoformat()


def ym_of_today() -> str:
    d = date.today()
    return f"{d.year:04d}-{d.month:02d}"


def safe_float(s: str) -> float:
    t = (s or "").strip()
    if not t:
        return 0.0
    t = t.replace(" ", "")
    t = t.replace(".", "")
    t = t.replace(",", ".")
    return float(t)


def validate_period(donem: str) -> bool:
    if not donem or len(donem) != 7 or donem[4] != "-":
        return False
    try:
        y = int(donem[:4])
        m = int(donem[5:])
        return 2000 <= y <= 2100 and 1 <= m <= 12
    except Exception:
        return False


def add_months(ym: str, add: int) -> str:
    y = int(ym[:4])
    m = int(ym[5:])
    m2 = m + add
    y += (m2 - 1) // 12
    m2 = ((m2 - 1) % 12) + 1
    return f"{y:04d}-{m2:02d}"


def months_range(ym_start: str, ym_end: str):
    ys = int(ym_start[:4]); ms = int(ym_start[5:])
    ye = int(ym_end[:4]); me = int(ym_end[5:])
    cur_y, cur_m = ys, ms
    out = []
    while (cur_y, cur_m) <= (ye, me):
        out.append(f"{cur_y:04d}-{cur_m:02d}")
        cur_m += 1
        if cur_m == 13:
            cur_m = 1
            cur_y += 1
    return out


def ensure_db():
    con = connect()
    con.executescript("""
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS daireler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        daire_no TEXT NOT NULL UNIQUE,
        ad_soyad TEXT NOT NULL,
        telefon TEXT DEFAULT '',
        aidat REAL NOT NULL DEFAULT 0,
        aktif INTEGER NOT NULL DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS sakinler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        daire_id INTEGER NOT NULL,
        ad_soyad TEXT NOT NULL,
        telefon TEXT DEFAULT '',
        baslangic_tarihi TEXT NOT NULL,
        bitis_tarihi TEXT,
        FOREIGN KEY(daire_id) REFERENCES daireler(id)
    );
    CREATE INDEX IF NOT EXISTS idx_sakinler_daire ON sakinler(daire_id);
    CREATE INDEX IF NOT EXISTS idx_sakinler_aktif ON sakinler(daire_id, bitis_tarihi);

    CREATE TABLE IF NOT EXISTS tahakkuk (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        donem TEXT NOT NULL,
        daire_id INTEGER NOT NULL,
        tutar REAL NOT NULL,
        created_at TEXT NOT NULL,
        UNIQUE(donem, daire_id),
        FOREIGN KEY(daire_id) REFERENCES daireler(id)
    );

    CREATE TABLE IF NOT EXISTS odemeler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        donem TEXT NOT NULL,
        tarih TEXT NOT NULL,
        daire_id INTEGER NOT NULL,
        tutar REAL NOT NULL,
        yontem TEXT NOT NULL,
        makbuz_no TEXT NOT NULL,
        aciklama TEXT DEFAULT '',
        FOREIGN KEY(daire_id) REFERENCES daireler(id)
    );

    CREATE INDEX IF NOT EXISTS idx_odemeler_donem ON odemeler(donem);
    CREATE INDEX IF NOT EXISTS idx_odemeler_daire ON odemeler(daire_id);

    CREATE TABLE IF NOT EXISTS giderler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        donem TEXT NOT NULL,
        tarih TEXT NOT NULL,
        kategori TEXT NOT NULL,
        tutar REAL NOT NULL,
        yontem TEXT NOT NULL,
        aciklama TEXT DEFAULT ''
    );
    CREATE INDEX IF NOT EXISTS idx_giderler_donem ON giderler(donem);

    CREATE TABLE IF NOT EXISTS duyurular (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tarih TEXT NOT NULL,
        baslik TEXT NOT NULL,
        mesaj TEXT NOT NULL
    );
    """)
    con.commit()

    def set_default(k, v):
        if con.execute("SELECT value FROM settings WHERE key=?", (k,)).fetchone() is None:
            con.execute("INSERT INTO settings(key,value) VALUES(?,?)", (k, str(v)))

    set_default("vade_gun", 10)
    set_default("gecikme_gun", 5)
    con.commit()
    con.close()


def get_setting_int(key: str, default: int) -> int:
    con = connect()
    row = con.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    con.close()
    try:
        return int(row[0]) if row else default
    except Exception:
        return default


def set_setting_int(key: str, value: int):
    con = connect()
    con.execute(
        "INSERT INTO settings(key,value) VALUES(?,?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, str(int(value)))
    )
    con.commit()
    con.close()


def receipt_next_for_period(donem: str) -> str:
    """Makbuz numarasını oluştur (YYYYMM-00001 formatında)"""
    yyyymm = donem.replace("-", "")
    con = connect()
    row = con.execute("""
        SELECT makbuz_no FROM odemeler
        WHERE donem=?
        ORDER BY id DESC
        LIMIT 1
    """, (donem,)).fetchone()
    con.close()
    
    if not row:
        return f"{yyyymm}-00001"
    
    last = row[0]
    try:
        seq = int(last.split("-")[1])
    except Exception:
        seq = 0
    return f"{yyyymm}-{seq+1:05d}"


def compute_vade_date(donem: str) -> date:
    vade_gun = get_setting_int("vade_gun", 10)
    y, m = donem.split("-")
    y = int(y); m = int(m)
    if m == 12:
        next_month = date(y + 1, 1, 1)
    else:
        next_month = date(y, m + 1, 1)
    last_day = next_month - timedelta(days=1)
    day = min(vade_gun, last_day.day)
    return date(y, m, day)


def status_for_balance(donem: str, bakiye: float) -> str:
    if bakiye <= 0.00001:
        return "Ödedi"
    vade = compute_vade_date(donem)
    gecikme = get_setting_int("gecikme_gun", 5)
    limit = vade + timedelta(days=gecikme)
    if date.today() > limit:
        return "Gecikmiş"
    return "Borçlu"


def autosize_worksheet(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            v = str(v)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)


def create_whatsapp_message(daire_no: str, ad: str, doneler: list, toplam: float, makbuz_no: str) -> str:
    """WhatsApp mesajı oluştur"""
    msg = f"*AİDAT ÖDEME ONAYLANDI* ✅\n\n"
    msg += f"Daire: {daire_no}\n"
    msg += f"Sakin: {ad}\n\n"
    msg += f"Makbuz No: {makbuz_no}\n"
    msg += f"Ödenen Dönemler:\n"
    for d in doneler:
        msg += f"  • {d}\n"
    msg += f"\n*TOPLAM: {toplam:.2f} TL*\n"
    msg += f"Tarih: {date.today().isoformat()}\n"
    return msg


def open_whatsapp_chat(telefon: str, mesaj: str):
    """WhatsApp Web üzerinden sohbeti aç"""
    if not telefon or len(telefon) < 10:
        QMessageBox.warning(None, "Uyarı", "Geçersiz telefon numarası.")
        return
    
    # Telefon numarasını temizle (sadece rakamlar)
    tel_clean = ''.join(filter(str.isdigit, telefon))
    if not tel_clean.startswith('90'):
        if tel_clean.startswith('0'):
            tel_clean = '90' + tel_clean[1:]
        else:
            tel_clean = '90' + tel_clean
    
    url = f"https://wa.me/{tel_clean}?text={quote(mesaj)}"
    webbrowser.open(url)


# ============ MODAL DIALOGS ============

class EditPaymentDialog(QDialog):
    """Ödeme düzenleme modalı"""
    def __init__(self, payment_id: int, parent=None):
        super().__init__(parent)
        self.payment_id = payment_id
        self.setWindowTitle("Ödemeyi Düzenle")
        self.setGeometry(100, 100, 500, 300)
        self.setModal(True)
        self.init_ui()
        self.load_payment()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        form = QFormLayout()
        
        self.in_donem = QLineEdit()
        self.in_donem.setReadOnly(True)
        self.dt_tarih = QDateEdit()
        self.dt_tarih.setCalendarPopup(True)
        self.dt_tarih.setDisplayFormat("dd-MM-yyyy")
        self.in_daire = QLineEdit()
        self.in_daire.setReadOnly(True)
        self.in_tutar = QLineEdit()
        self.cmb_yontem = QComboBox()
        self.cmb_yontem.addItems(["Banka", "Elden"])
        self.in_makbuz = QLineEdit()
        self.in_acik = QLineEdit()
        
        form.addRow("Dönem", self.in_donem)
        form.addRow("Tarih", self.dt_tarih)
        form.addRow("Daire", self.in_daire)
        form.addRow("Tutar (TL)", self.in_tutar)
        form.addRow("Yöntem", self.cmb_yontem)
        form.addRow("Makbuz No", self.in_makbuz)
        form.addRow("Açıklama", self.in_acik)
        
        layout.addLayout(form)
        
        btns = QHBoxLayout()
        self.btn_save = QPushButton("Güncelle")
        self.btn_save.clicked.connect(self.save_payment)
        self.btn_delete = QPushButton("Sil")
        self.btn_delete.clicked.connect(self.delete_payment)
        self.btn_close = QPushButton("Kapat")
        self.btn_close.clicked.connect(self.close)
        
        btns.addWidget(self.btn_save)
        btns.addWidget(self.btn_delete)
        btns.addStretch(1)
        btns.addWidget(self.btn_close)
        
        layout.addLayout(btns)
    
    def load_payment(self):
        con = connect()
        row = con.execute("""
            SELECT o.donem, o.tarih, d.daire_no, o.tutar, o.yontem, o.makbuz_no, o.aciklama
            FROM odemeler o
            JOIN daireler d ON d.id = o.daire_id
            WHERE o.id = ?
        """, (self.payment_id,)).fetchone()
        con.close()
        
        if not row:
            QMessageBox.warning(self, "Hata", "Ödeme bulunamadı.")
            self.close()
            return
        
        donem, tarih, daire_no, tutar, yontem, makbuz_no, acik = row
        
        self.in_donem.setText(donem)
        self.dt_tarih.setDate(QDate.fromString(tarih, "yyyy-MM-dd"))
        self.in_daire.setText(daire_no)
        self.in_tutar.setText(f"{float(tutar):.2f}")
        self.cmb_yontem.setCurrentText(yontem)
        self.in_makbuz.setText(makbuz_no or "")
        self.in_acik.setText(acik or "")
    
    def save_payment(self):
        try:
            tutar = safe_float(self.in_tutar.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Tutar sayısal olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        tarih = self.dt_tarih.date().toPython().isoformat()
        yontem = self.cmb_yontem.currentText()
        acik = self.in_acik.text().strip()
        
        con = connect()
        con.execute("""
            UPDATE odemeler
            SET tarih=?, tutar=?, yontem=?, aciklama=?
            WHERE id=?
        """, (tarih, float(tutar), yontem, acik, self.payment_id))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Ödeme güncellendi.")
        self.accept()
    
    def delete_payment(self):
        reply = QMessageBox.question(
            self, "Onay", "Seçili ödemeyi silmek istiyor musunuz?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        
        con = connect()
        con.execute("DELETE FROM odemeler WHERE id=?", (self.payment_id,))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Ödeme silindi.")
        self.accept()


class EditExpenseDialog(QDialog):
    """Gider düzenleme modalı"""
    def __init__(self, expense_id: int, parent=None):
        super().__init__(parent)
        self.expense_id = expense_id
        self.setWindowTitle("Gideri Düzenle")
        self.setGeometry(100, 100, 500, 300)
        self.setModal(True)
        self.init_ui()
        self.load_expense()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        form = QFormLayout()
        
        self.in_donem = QLineEdit()
        self.in_donem.setReadOnly(True)
        
        self.dt_tarih = QDateEdit()
        self.dt_tarih.setCalendarPopup(True)
        self.dt_tarih.setDisplayFormat("dd-MM-yyyy")
        
        self.cmb_kat = QComboBox()
        self.cmb_kat.addItems(["Elektrik", "Temizlik", "Bakım", "Diğer"])
        
        self.in_tutar = QLineEdit()
        
        self.cmb_yontem = QComboBox()
        self.cmb_yontem.addItems(["Banka", "Elden"])
        
        self.in_acik = QLineEdit()
        
        form.addRow("Dönem", self.in_donem)
        form.addRow("Tarih", self.dt_tarih)
        form.addRow("Kategori", self.cmb_kat)
        form.addRow("Tutar (TL)", self.in_tutar)
        form.addRow("Yöntem", self.cmb_yontem)
        form.addRow("Açıklama", self.in_acik)
        
        layout.addLayout(form)
        
        btns = QHBoxLayout()
        self.btn_save = QPushButton("Güncelle")
        self.btn_save.clicked.connect(self.save_expense)
        self.btn_delete = QPushButton("Sil")
        self.btn_delete.clicked.connect(self.delete_expense)
        self.btn_close = QPushButton("Kapat")
        self.btn_close.clicked.connect(self.close)
        
        btns.addWidget(self.btn_save)
        btns.addWidget(self.btn_delete)
        btns.addStretch(1)
        btns.addWidget(self.btn_close)
        
        layout.addLayout(btns)
    
    def load_expense(self):
        con = connect()
        row = con.execute("""
            SELECT donem, tarih, kategori, tutar, yontem, aciklama
            FROM giderler
            WHERE id = ?
        """, (self.expense_id,)).fetchone()
        con.close()
        
        if not row:
            QMessageBox.warning(self, "Hata", "Gider bulunamadı.")
            self.close()
            return
        
        donem, tarih, kat, tutar, yontem, acik = row
        
        self.in_donem.setText(donem)
        self.dt_tarih.setDate(QDate.fromString(tarih, "yyyy-MM-dd"))
        self.cmb_kat.setCurrentText(kat)
        self.in_tutar.setText(f"{float(tutar):.2f}")
        self.cmb_yontem.setCurrentText(yontem)
        self.in_acik.setText(acik or "")
    
    def save_expense(self):
        try:
            tutar = safe_float(self.in_tutar.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Tutar sayısal olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        tarih = self.dt_tarih.date().toPython().isoformat()
        kat = self.cmb_kat.currentText()
        yontem = self.cmb_yontem.currentText()
        acik = self.in_acik.text().strip()
        
        con = connect()
        con.execute("""
            UPDATE giderler
            SET tarih=?, kategori=?, tutar=?, yontem=?, aciklama=?
            WHERE id=?
        """, (tarih, kat, float(tutar), yontem, acik, self.expense_id))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Gider güncellendi.")
        self.accept()
    
    def delete_expense(self):
        reply = QMessageBox.question(
            self, "Onay", "Seçili gideri silmek istiyor musunuz?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        
        con = connect()
        con.execute("DELETE FROM giderler WHERE id=?", (self.expense_id,))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Gider silindi.")
        self.accept()


# ============ MAIN APPLICATION ============

class ApartmanAidatApp(QWidget):
    """Ana Apartman Aidat Uygulaması - Temiz Ödeme Sistemi ile"""
    
    def __init__(self):
        super().__init__()
        self.selected_daire_id = None
        ensure_db()
        self.setWindowTitle("Apartman Aidat Sistemi v6 - Temiz Ödeme")
        screen = QApplication.primaryScreen().availableGeometry()
        w = int(screen.width() * 0.85)
        h = int(screen.height() * 0.85)
        self.resize(w, h)
        self.setMinimumSize(min(1100, w), min(700, h))
        self._build_ui()
        self.refresh_all()
    
    def _build_ui(self):
        root = QVBoxLayout(self)
        
        topbar = QHBoxLayout()
        root.addLayout(topbar)
        
        topbar.addWidget(QLabel("Dönem:"))
        self.in_donem = QLineEdit(ym_of_today())
        self.in_donem.setMaximumWidth(110)
        topbar.addWidget(self.in_donem)
        
        self.btn_refresh = QPushButton("Yenile")
        self.btn_refresh.clicked.connect(self.refresh_all)
        topbar.addWidget(self.btn_refresh)
        
        topbar.addStretch(1)
        
        topbar.addWidget(QLabel("Vade günü:"))
        self.sp_vade = QSpinBox()
        self.sp_vade.setRange(1, 28)
        self.sp_vade.setValue(get_setting_int("vade_gun", 10))
        topbar.addWidget(self.sp_vade)
        
        topbar.addWidget(QLabel("Gecikme + gün:"))
        self.sp_gec = QSpinBox()
        self.sp_gec.setRange(0, 60)
        self.sp_gec.setValue(get_setting_int("gecikme_gun", 5))
        topbar.addWidget(self.sp_gec)
        
        self.btn_save_settings = QPushButton("Ayarları Kaydet")
        self.btn_save_settings.clicked.connect(self.save_settings)
        topbar.addWidget(self.btn_save_settings)
        
        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)
        
        self._tab_daireler()
        self._tab_tahakkuk()
        self._tab_odeme()  # ✨ TEMIZ ÖDEME SİSTEMİ
        self._tab_odeme_gecmisi()  # YENI: Geçmiş Ödemeler
        self._tab_gider()
        self._tab_duyuru()
        self._tab_rapor()
        self._tab_kasa()  # YENI: Kasa Raporu
        self._tab_ekstre()  # YENI: Ekstre
    
    def _tab_daireler(self):
        """Daire yönetimi sekmesi"""
        w = QWidget()
        lay = QVBoxLayout(w)
        
        gb = QGroupBox("Daire Ekle / Güncelle")
        form = QFormLayout(gb)
        self.d_in_no = QLineEdit()
        self.d_in_ad = QLineEdit()
        self.d_in_tel = QLineEdit()
        self.d_in_aidat = QLineEdit()
        self.d_chk_aktif = QCheckBox("Aktif")
        self.d_chk_aktif.setChecked(True)
        
        form.addRow("Daire No", self.d_in_no)
        form.addRow("Ad Soyad", self.d_in_ad)
        form.addRow("Telefon", self.d_in_tel)
        form.addRow("Aidat (TL)", self.d_in_aidat)
        form.addRow("", self.d_chk_aktif)
        
        btns = QHBoxLayout()
        self.btn_d_save = QPushButton("Kaydet")
        self.btn_d_save.clicked.connect(self.daire_save)
        self.btn_d_clear = QPushButton("Temizle")
        self.btn_d_clear.clicked.connect(self.daire_clear)
        
        btns.addWidget(self.btn_d_save)
        btns.addWidget(self.btn_d_clear)
        btns.addStretch(1)
        
        lay.addWidget(gb)
        lay.addLayout(btns)
        
        self.tbl_daire = QTableWidget(0, 6)
        self.tbl_daire.setHorizontalHeaderLabels(["ID", "Daire", "Ad", "Tel", "Aidat", "Aktif"])
        self.tbl_daire.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_daire.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_daire.cellClicked.connect(self.daire_row_clicked)
        self.tbl_daire.horizontalHeader().setStretchLastSection(True)
        self.tbl_daire.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_daire, 1)
        
        self.tabs.addTab(w, "Daireler")
    
    def refresh_daire_table(self):
        con = connect()
        rows = con.execute("""
            SELECT id, daire_no, ad_soyad, telefon, aidat, aktif
            FROM daireler
            ORDER BY CAST(daire_no AS INTEGER), daire_no
        """).fetchall()
        con.close()
        
        self.tbl_daire.setRowCount(0)
        for r in rows:
            row = self.tbl_daire.rowCount()
            self.tbl_daire.insertRow(row)
            for c, val in enumerate(r):
                it = QTableWidgetItem(str(val))
                if c in (0, 5):
                    it.setTextAlignment(Qt.AlignCenter)
                if c == 4:
                    it.setText(f"{float(val):.2f}")
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tbl_daire.setItem(row, c, it)
    
    def daire_row_clicked(self, row, col):
        try:
            did = int(self.tbl_daire.item(row, 0).text())
            dno = self.tbl_daire.item(row, 1).text()
            ad = self.tbl_daire.item(row, 2).text()
            tel = self.tbl_daire.item(row, 3).text()
            aidat = self.tbl_daire.item(row, 4).text()
            aktif = self.tbl_daire.item(row, 5).text() == "1"
        except Exception:
            return
        
        self.selected_daire_id = did
        self.d_in_no.setText(dno)
        self.d_in_ad.setText(ad)
        self.d_in_tel.setText(tel)
        self.d_in_aidat.setText(aidat)
        self.d_chk_aktif.setChecked(aktif)
    
    def daire_clear(self):
        self.selected_daire_id = None
        self.d_in_no.clear()
        self.d_in_ad.clear()
        self.d_in_tel.clear()
        self.d_in_aidat.clear()
        self.d_chk_aktif.setChecked(True)
        self.tbl_daire.clearSelection()
    
    def daire_save(self):
        dno = self.d_in_no.text().strip()
        ad = self.d_in_ad.text().strip()
        tel = self.d_in_tel.text().strip()
        
        try:
            aidat = safe_float(self.d_in_aidat.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Aidat sayısal olmalı.")
            return
        
        aktif = 1 if self.d_chk_aktif.isChecked() else 0
        
        if not dno or not ad:
            QMessageBox.warning(self, "Uyarı", "Daire No ve Ad zorunlu.")
            return
        
        con = connect()
        try:
            if self.selected_daire_id:
                conflict = con.execute(
                    "SELECT id FROM daireler WHERE daire_no=? AND id<>?",
                    (dno, int(self.selected_daire_id))
                ).fetchone()
                
                if conflict:
                    QMessageBox.warning(self, "Uyarı", "Bu daire no başka bir kayıtta kullanılıyor.")
                    con.close()
                    return
                
                con.execute("""
                    UPDATE daireler
                    SET daire_no=?, ad_soyad=?, telefon=?, aidat=?, aktif=?
                    WHERE id=?
                """, (dno, ad, tel, float(aidat), int(aktif), int(self.selected_daire_id)))
                
                did = int(self.selected_daire_id)
            else:
                con.execute("""
                    INSERT INTO daireler(daire_no, ad_soyad, telefon, aidat, aktif)
                    VALUES(?,?,?,?,?)
                """, (dno, ad, tel, float(aidat), int(aktif)))
                
                did = int(con.execute(
                    "SELECT id FROM daireler WHERE daire_no=?",
                    (dno,)
                ).fetchone()[0])
            
            con.commit()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Uyarı", "Bu daire no zaten var.")
        finally:
            con.close()
        
        self.refresh_all()
        self.daire_clear()
    
    def _tab_tahakkuk(self):
        """Tahakkuk sekmesi"""
        w = QWidget()
        lay = QVBoxLayout(w)
        
        top = QGroupBox("Tahakkuk / Borç")
        form = QFormLayout(top)
        
        self.t_cmb_daire = QComboBox()
        self.t_in_from = QLineEdit(ym_of_today())
        self.t_in_to = QLineEdit(ym_of_today())
        self.t_in_tutar = QLineEdit()
        self.t_in_tutar.setPlaceholderText("Örn: 400")
        
        btns = QHBoxLayout()
        self.btn_t_add = QPushButton("Tahakkuk Oluştur")
        self.btn_t_add.clicked.connect(self.add_tahakkuk_range)
        btns.addWidget(self.btn_t_add)
        btns.addStretch()
        
        form.addRow("Daire", self.t_cmb_daire)
        form.addRow("Başlangıç", self.t_in_from)
        form.addRow("Bitiş", self.t_in_to)
        form.addRow("Tutar", self.t_in_tutar)
        form.addRow("", btns)
        lay.addWidget(top)
        
        self.tbl_tah = QTableWidget(0, 5)
        self.tbl_tah.setHorizontalHeaderLabels(["ID", "Daire", "Dönem", "Tutar", "Tarih"])
        self.tbl_tah.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_tah.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_tah.horizontalHeader().setStretchLastSection(True)
        self.tbl_tah.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_tah, 1)
        
        self.tabs.addTab(w, "Tahakkuk")
    
    def refresh_tahakkuk_table(self):
        con = connect()
        rows = con.execute("""
            SELECT t.id,
                   d.daire_no || ' - ' || d.ad_soyad AS daire,
                   t.donem,
                   t.tutar,
                   t.created_at
            FROM tahakkuk t
            JOIN daireler d ON d.id = t.daire_id
            ORDER BY t.donem DESC, CAST(d.daire_no AS INT), d.daire_no
        """).fetchall()
        con.close()
        
        self.tbl_tah.setRowCount(0)
        for rec in rows:
            row = self.tbl_tah.rowCount()
            self.tbl_tah.insertRow(row)
            for c, val in enumerate(rec):
                it = QTableWidgetItem(str(val))
                if c == 0:
                    it.setTextAlignment(Qt.AlignCenter)
                if c == 3:
                    it.setText(f"{float(val):.2f}")
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tbl_tah.setItem(row, c, it)
    
    def add_tahakkuk_range(self):
        daire_id = self.t_cmb_daire.currentData()
        if not daire_id:
            QMessageBox.warning(self, "Uyarı", "Daire seçin.")
            return
        
        start = self.t_in_from.text().strip()
        end = self.t_in_to.text().strip()
        
        if not validate_period(start) or not validate_period(end):
            QMessageBox.warning(self, "Uyarı", "Dönem formatı YYYY-MM olmalı.")
            return
        
        try:
            tutar = safe_float(self.t_in_tutar.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Tutar sayısal olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        con = connect()
        adet = 0
        for donem in months_range(start, end):
            con.execute("""
                INSERT INTO tahakkuk(donem, daire_id, tutar, created_at)
                VALUES(?,?,?,?)
                ON CONFLICT(donem, daire_id)
                DO UPDATE SET tutar=excluded.tutar, created_at=excluded.created_at
            """, (donem, int(daire_id), float(tutar), iso_today()))
            adet += 1
        
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", f"{adet} dönem için tahakkuk oluşturuldu.")
        self.refresh_all()
    
    def _tab_odeme(self):
        """✨ TEMIZ ÖDEME SİSTEMİ - Checkbox tabanlı çoklu dönem seçimi"""
        w = QWidget()
        lay = QVBoxLayout(w)
        
        # Daire seçimi
        top_form = QFormLayout()
        self.o_cmb_daire = QComboBox()
        self.o_cmb_daire.currentIndexChanged.connect(self.on_payment_daire_changed)
        top_form.addRow("Daire Seç:", self.o_cmb_daire)
        lay.addLayout(top_form)
        
        # Borçlu dönemler listesi
        gb_borc = QGroupBox("Borçlu Dönemler (Checkbox ile Ödenecekleri Seç)")
        layout_borc = QVBoxLayout(gb_borc)
        
        self.tbl_borc = QTableWidget()
        self.tbl_borc.setColumnCount(4)
        self.tbl_borc.setHorizontalHeaderLabels(["Seç", "Dönem", "Borç (TL)", "Durum"])
        self.tbl_borc.setSelectionBehavior(QTableWidget.SelectRows)
        layout_borc.addWidget(self.tbl_borc)
        lay.addWidget(gb_borc, 1)
        
        # Ödeme formu
        gb_form = QGroupBox("Ödeme Bilgileri")
        form_payment = QFormLayout(gb_form)
        
        self.o_dt_tarih = QDateEdit()
        self.o_dt_tarih.setCalendarPopup(True)
        self.o_dt_tarih.setDisplayFormat("dd-MM-yyyy")
        self.o_dt_tarih.setDate(QDate.currentDate())
        
        self.o_cmb_yontem = QComboBox()
        self.o_cmb_yontem.addItems(["Banka", "Elden"])
        
        self.o_in_acik = QLineEdit()
        self.o_in_acik.setPlaceholderText("İsteğe bağlı açıklama")
        
        self.o_lbl_ozet = QLabel("Seçili Dönemler: Yok | Toplam: 0.00 TL")
        self.o_lbl_ozet.setStyleSheet("font-weight: bold; color: #0066cc; padding: 8px;")
        
        form_payment.addRow("Ödeme Tarihi:", self.o_dt_tarih)
        form_payment.addRow("Ödeme Yöntemi:", self.o_cmb_yontem)
        form_payment.addRow("Açıklama:", self.o_in_acik)
        form_payment.addRow("", self.o_lbl_ozet)
        
        lay.addWidget(gb_form)
        
        # Butonlar
        btns = QHBoxLayout()
        
        btn_kaydet = QPushButton("✅ Ödemeyi Kaydet")
        btn_kaydet.clicked.connect(self.payment_save)
        btn_kaydet.setMinimumHeight(40)
        btn_kaydet.setStyleSheet("background-color: #28a745; color: white; font-weight: bold;")
        
        btn_whatsapp = QPushButton("💬 WhatsApp Gönder")
        btn_whatsapp.clicked.connect(self.payment_send_whatsapp)
        btn_whatsapp.setMinimumHeight(40)
        btn_whatsapp.setStyleSheet("background-color: #25D366; color: white; font-weight: bold;")
        
        btn_temizle = QPushButton("🔄 Temizle")
        btn_temizle.clicked.connect(self.payment_clear)
        
        btns.addWidget(btn_kaydet, 2)
        btns.addWidget(btn_whatsapp, 2)
        btns.addWidget(btn_temizle, 1)
        btns.addStretch(1)
        lay.addLayout(btns)
        
        self.tabs.addTab(w, "Ödeme")
    
    def payment_send_whatsapp(self):
        """Ödeme OnayıWhatsApp ile gönder"""
        daire_id = self.o_cmb_daire.currentData()
        
        if not daire_id:
            QMessageBox.warning(self, "Uyarı", "Daire seçin.")
            return
        
        # Seçili dönemleri topla
        odenecek_doneler = []
        toplam_tutar = 0.0
        
        for row in range(self.tbl_borc.rowCount()):
            chk = self.tbl_borc.cellWidget(row, 0)
            if chk and chk.isChecked():
                donem = self.tbl_borc.item(row, 1).text()
                tutar = float(self.tbl_borc.item(row, 2).text())
                odenecek_doneler.append(donem)
                toplam_tutar += tutar
        
        if not odenecek_doneler:
            QMessageBox.warning(self, "Uyarı", "Ödeme yapılacak dönem seçin.")
            return
        
        # Daire bilgilerini al
        daire_info = self._get_daire_info(int(daire_id))
        makbuz_no = receipt_next_for_period(odenecek_doneler[0])
        
        # WhatsApp mesajı oluştur
        mesaj = create_whatsapp_message(
            daire_info['daire_no'],
            daire_info['ad'],
            odenecek_doneler,
            toplam_tutar,
            makbuz_no
        )
        
        # WhatsApp'ı aç
        open_whatsapp_chat(daire_info['tel'], mesaj)
    
    def refresh_odeme_daire_combo(self):
        con = connect()
        rows = con.execute("""
            SELECT id, daire_no, ad_soyad, aktif
            FROM daireler
            ORDER BY CAST(daire_no AS INT), daire_no
        """).fetchall()
        con.close()
        
        self.o_cmb_daire.blockSignals(True)
        self.o_cmb_daire.clear()
        self.o_cmb_daire.addItem("-- seç --", None)
        
        for did, dno, ad, aktif in rows:
            tag = "" if int(aktif) == 1 else " (pasif)"
            self.o_cmb_daire.addItem(f"Daire {dno} - {ad}{tag}", did)
        
        self.o_cmb_daire.blockSignals(False)
    
    def on_payment_daire_changed(self):
        """Ödeme dairesi seçildiğinde borçlu dönemleri yükle"""
        daire_id = self.o_cmb_daire.currentData()
        
        if not daire_id:
            self.tbl_borc.setRowCount(0)
            self.update_payment_summary()
            return
        
        self.load_payment_borc_doneleri(int(daire_id))
        self.update_payment_summary()
    
    def load_payment_borc_doneleri(self, daire_id: int):
        """Ödeme dönemleri - borçlu dönemleri yükle"""
        con = connect()
        
        rows = con.execute("""
            SELECT 
                COALESCE(t.donem, o.donem) as donem,
                COALESCE(SUM(t.tutar), 0) as tahakkuk,
                COALESCE(SUM(o.tutar), 0) as odeme
            FROM (
                SELECT DISTINCT donem FROM tahakkuk WHERE daire_id = ?
                UNION
                SELECT DISTINCT donem FROM odemeler WHERE daire_id = ?
            ) d
            LEFT JOIN tahakkuk t ON t.donem = d.donem AND t.daire_id = ?
            LEFT JOIN odemeler o ON o.donem = d.donem AND o.daire_id = ?
            GROUP BY COALESCE(t.donem, o.donem)
            HAVING (COALESCE(SUM(t.tutar), 0) - COALESCE(SUM(o.tutar), 0)) > 0.00001
            ORDER BY COALESCE(t.donem, o.donem) DESC
        """, (daire_id, daire_id, daire_id, daire_id)).fetchall()
        
        con.close()
        
        self.tbl_borc.setRowCount(0)
        
        for donem, tahakkuk, odeme in rows:
            kalan = float(tahakkuk) - float(odeme)
            
            row = self.tbl_borc.rowCount()
            self.tbl_borc.insertRow(row)
            
            chk = QCheckBox()
            chk.stateChanged.connect(self.update_payment_summary)
            self.tbl_borc.setCellWidget(row, 0, chk)
            
            item_donem = QTableWidgetItem(str(donem))
            item_donem.setTextAlignment(Qt.AlignCenter)
            self.tbl_borc.setItem(row, 1, item_donem)
            
            item_tutar = QTableWidgetItem(f"{kalan:.2f}")
            item_tutar.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_borc.setItem(row, 2, item_tutar)
            
            if kalan > float(tahakkuk) * 0.5:
                durum = "Yüksek"
                color = QColor(255, 200, 200)
            else:
                durum = "Kısmi"
                color = QColor(255, 255, 150)
            
            item_durum = QTableWidgetItem(durum)
            item_durum.setBackground(color)
            self.tbl_borc.setItem(row, 3, item_durum)
            
            for col in range(4):
                if col != 0:
                    self.tbl_borc.item(row, col).setBackground(QColor(245, 250, 255))
    
    def update_payment_summary(self):
        """Seçili dönemlerin özetini güncelle"""
        toplam = 0.0
        secili_doneler = []
        
        for row in range(self.tbl_borc.rowCount()):
            chk = self.tbl_borc.cellWidget(row, 0)
            if chk and chk.isChecked():
                donem = self.tbl_borc.item(row, 1).text()
                tutar = float(self.tbl_borc.item(row, 2).text())
                toplam += tutar
                secili_doneler.append(donem)
        
        if secili_doneler:
            ozet_text = f"Seçili Dönemler: {', '.join(secili_doneler)} | Toplam: {toplam:.2f} TL"
            self.o_lbl_ozet.setStyleSheet("font-weight: bold; color: #28a745; padding: 8px; background-color: #f0f8f0;")
        else:
            ozet_text = "Seçili Dönemler: Yok | Toplam: 0.00 TL"
            self.o_lbl_ozet.setStyleSheet("font-weight: bold; color: #0066cc; padding: 8px;")
        
        self.o_lbl_ozet.setText(ozet_text)
    
    def payment_clear(self):
        """Ödeme formunu temizle"""
        for row in range(self.tbl_borc.rowCount()):
            chk = self.tbl_borc.cellWidget(row, 0)
            if chk:
                chk.setChecked(False)
        
        self.update_payment_summary()
    
    def payment_save(self):
        """Ödemeyi kaydet - TEMIZ SISTEM"""
        daire_id = self.o_cmb_daire.currentData()
        
        if not daire_id:
            QMessageBox.warning(self, "Uyarı", "Daire seçin.")
            return
        
        # Seçili dönemleri topla
        odenecek_doneler = []
        toplam_tutar = 0.0
        
        for row in range(self.tbl_borc.rowCount()):
            chk = self.tbl_borc.cellWidget(row, 0)
            if chk and chk.isChecked():
                donem = self.tbl_borc.item(row, 1).text()
                tutar = float(self.tbl_borc.item(row, 2).text())
                odenecek_doneler.append((donem, tutar))
                toplam_tutar += tutar
        
        if not odenecek_doneler:
            QMessageBox.warning(self, "Uyarı", "Ödeme yapılacak dönem seçin.")
            return
        
        tarih = self.o_dt_tarih.date().toPython().isoformat()
        yontem = self.o_cmb_yontem.currentText()
        acik = self.o_in_acik.text().strip()
        
        ilk_donem = odenecek_doneler[0][0]
        makbuz_no = receipt_next_for_period(ilk_donem)
        
        # DB'ye kaydet
        con = connect()
        try:
            for donem, tutar in odenecek_doneler:
                con.execute("""
                    INSERT INTO odemeler(donem, tarih, daire_id, tutar, yontem, makbuz_no, aciklama)
                    VALUES(?, ?, ?, ?, ?, ?, ?)
                """, (donem, tarih, int(daire_id), float(tutar), yontem, makbuz_no, acik))
            
            con.commit()
            con.close()
            
            # Makbuz oluştur
            daire_info = self._get_daire_info(int(daire_id))
            pdf_path = self.create_receipt_pdf(
                makbuz_no=makbuz_no,
                tarih=tarih,
                daire_no=daire_info['daire_no'],
                ad=daire_info['ad'],
                yontem=yontem,
                odenecek_doneler=odenecek_doneler,
                toplam=toplam_tutar,
                aciklama=acik
            )
            
            # Onay mesajı
            msg = f"✅ Ödeme Kaydedildi!\n\nMakbuz No: {makbuz_no}\nToplam: {toplam_tutar:.2f} TL\n\nMakbuzı açmak ister misiniz?"
            reply = QMessageBox.question(
                self, "Ödeme Başarılı", msg,
                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes and pdf_path:
                webbrowser.open(Path(pdf_path).resolve().as_uri())
            
            # Formu temizle ve yenile
            self.payment_clear()
            self.o_in_acik.clear()
            self.o_dt_tarih.setDate(QDate.currentDate())
            self.on_payment_daire_changed()
            self.refresh_all()
            
        except Exception as e:
            con.close()
            QMessageBox.critical(self, "Hata", f"Ödeme kaydedilemedi:\n{str(e)}")
    
    def _tab_odeme_gecmisi(self):
        """YENI: Geçmiş ödemeler tablosu"""
        w = QWidget()
        lay = QVBoxLayout(w)
        
        top = QHBoxLayout()
        top.addWidget(QLabel("Dönem:"))
        self.og_in_donem = QLineEdit(ym_of_today())
        self.og_in_donem.setMaximumWidth(110)
        self.og_in_donem.textChanged.connect(self.refresh_odeme_gecmisi_table)
        top.addWidget(self.og_in_donem)
        
        top.addWidget(QLabel("Daire:"))
        self.og_cmb_daire = QComboBox()
        self.og_cmb_daire.addItem("Tümü", None)
        self.og_cmb_daire.currentIndexChanged.connect(self.refresh_odeme_gecmisi_table)
        top.addWidget(self.og_cmb_daire)
        
        top.addStretch(1)
        lay.addLayout(top)
        
        self.tbl_odeme_gecmis = QTableWidget(0, 8)
        self.tbl_odeme_gecmis.setHorizontalHeaderLabels(["ID", "Dönem", "Daire", "Ad", "Tarih", "Tutar", "Yöntem", "Makbuz"])
        self.tbl_odeme_gecmis.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_odeme_gecmis.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_odeme_gecmis.doubleClicked.connect(self.open_payment_for_edit)
        self.tbl_odeme_gecmis.horizontalHeader().setStretchLastSection(True)
        self.tbl_odeme_gecmis.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_odeme_gecmis, 1)
        
        self.tabs.addTab(w, "Geçmiş Ödemeler")
    
    def refresh_odeme_gecmisi_daire_combo(self):
        """Geçmiş ödeme ekranı için daire combo'yu doldur"""
        con = connect()
        rows = con.execute("""
            SELECT DISTINCT id, daire_no, ad_soyad
            FROM daireler
            ORDER BY CAST(daire_no AS INT), daire_no
        """).fetchall()
        con.close()
        
        self.og_cmb_daire.blockSignals(True)
        self.og_cmb_daire.clear()
        self.og_cmb_daire.addItem("Tümü", None)
        
        for did, dno, ad in rows:
            self.og_cmb_daire.addItem(f"{dno} - {ad}", did)
        
        self.og_cmb_daire.blockSignals(False)
    
    def refresh_odeme_gecmisi_table(self):
        """Geçmiş ödemeleri yükle"""
        donem = self.og_in_donem.text().strip()
        daire_id = self.og_cmb_daire.currentData()
        
        con = connect()
        
        if not validate_period(donem):
            query = """
                SELECT o.id, o.donem, d.daire_no, d.ad_soyad, o.tarih, o.tutar, o.yontem, o.makbuz_no
                FROM odemeler o
                JOIN daireler d ON d.id = o.daire_id
                WHERE 1=1
            """
            params = []
        else:
            query = """
                SELECT o.id, o.donem, d.daire_no, d.ad_soyad, o.tarih, o.tutar, o.yontem, o.makbuz_no
                FROM odemeler o
                JOIN daireler d ON d.id = o.daire_id
                WHERE o.donem = ?
            """
            params = [donem]
        
        if daire_id:
            query += " AND o.daire_id = ?"
            params.append(daire_id)
        
        query += " ORDER BY o.id DESC"
        
        rows = con.execute(query, params).fetchall()
        con.close()
        
        self.tbl_odeme_gecmis.setRowCount(0)
        for rec in rows:
            row = self.tbl_odeme_gecmis.rowCount()
            self.tbl_odeme_gecmis.insertRow(row)
            
            for c, val in enumerate(rec):
                it = QTableWidgetItem(str(val))
                
                if c == 0:
                    it.setTextAlignment(Qt.AlignCenter)
                if c == 5:  # Tutar
                    it.setText(f"{float(val):.2f}")
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                
                it.setBackground(QColor(245, 250, 255))
                self.tbl_odeme_gecmis.setItem(row, c, it)
    
    def open_payment_for_edit(self, index):
        """Ödemeyi düzenle modalı aç"""
        row = index.row()
        
        try:
            payment_id = int(self.tbl_odeme_gecmis.item(row, 0).text())
        except Exception:
            QMessageBox.warning(self, "Hata", "Ödeme ID'si alınamadı.")
            return
        
        dialog = EditPaymentDialog(payment_id, self)
        dialog.exec()
        
        self.refresh_odeme_gecmisi_table()
        self.refresh_all()
    
    def _get_daire_info(self, daire_id: int) -> dict:
        """Daire bilgilerini al"""
        con = connect()
        row = con.execute(
            "SELECT daire_no, ad_soyad, telefon FROM daireler WHERE id=?",
            (daire_id,)
        ).fetchone()
        con.close()
        
        if not row:
            return {'daire_no': '?', 'ad': '?', 'tel': ''}
        
        return {
            'daire_no': row[0],
            'ad': row[1],
            'tel': row[2]
        }
    
    def create_receipt_pdf(self, makbuz_no: str, tarih: str, daire_no: str, ad: str,
                          yontem: str, odenecek_doneler: list, toplam: float, aciklama: str) -> str:
        """Makbuz PDF oluştur"""
        
        out_dir = Path.cwd() / "makbuzlar"
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / f"makbuz_{makbuz_no}.pdf"
        
        c = canvas.Canvas(str(out_path), pagesize=A4)
        w, h = A4
        
        y = h - 50
        
        # Başlık
        c.setFont("Helvetica-Bold", 18)
        c.drawString(50, y, "AİDAT MAKBUZU")
        y -= 30
        
        # Üst bilgiler
        c.setFont("Helvetica", 11)
        c.drawString(50, y, f"Makbuz No: {makbuz_no}")
        c.drawRightString(w - 50, y, f"Tarih: {tarih}")
        y -= 20
        
        c.drawString(50, y, f"Daire: {daire_no}")
        c.drawRightString(w - 50, y, f"Ad Soyad: {ad}")
        y -= 20
        
        c.drawString(50, y, f"Ödeme Yöntemi: {yontem}")
        y -= 30
        
        # Ödenen dönemler başlığı
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "ÖDENEN DÖNEMLER")
        y -= 20
        
        # Tablo başlığı
        c.setFont("Helvetica-Bold", 10)
        c.drawString(50, y, "Dönem")
        c.drawRightString(200, y, "Tutar (TL)")
        y -= 15
        c.line(50, y, 250, y)
        y -= 15
        
        # Dönemler
        c.setFont("Helvetica", 10)
        for donem, tutar in odenecek_doneler:
            c.drawString(50, y, donem)
            c.drawRightString(200, y, f"{tutar:.2f}")
            y -= 15
        
        # Ayırıcı çizgi
        y -= 5
        c.line(50, y, 250, y)
        y -= 15
        
        # Toplam
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "TOPLAM")
        c.drawRightString(200, y, f"{toplam:.2f} TL")
        y -= 30
        
        # Açıklama
        if aciklama:
            c.setFont("Helvetica", 9)
            c.drawString(50, y, f"Not: {aciklama[:80]}")
            y -= 20
        
        # İmza alanı
        y -= 30
        c.setFont("Helvetica", 9)
        c.drawString(50, y, "Kasiyerin İmzası:")
        c.line(180, y - 2, 350, y - 2)
        
        c.save()
        return str(out_path)
    
    def _tab_gider(self):
        """Gider sekmesi"""
        w = QWidget()
        lay = QVBoxLayout(w)
        
        gb = QGroupBox("Gider Ekle")
        form = QFormLayout(gb)
        
        self.g_in_donem = QLineEdit(ym_of_today())
        self.g_dt_tarih = QDateEdit()
        self.g_dt_tarih.setCalendarPopup(True)
        self.g_dt_tarih.setDisplayFormat("dd-MM-yyyy")
        self.g_dt_tarih.setDate(QDate.currentDate())
        
        self.g_cmb_kat = QComboBox()
        self.g_cmb_kat.addItems(["Elektrik", "Temizlik", "Bakım", "Diğer"])
        
        self.g_in_tutar = QLineEdit()
        
        self.g_cmb_yontem = QComboBox()
        self.g_cmb_yontem.addItems(["Banka", "Elden"])
        
        self.g_in_acik = QLineEdit()
        
        self.btn_g_save = QPushButton("Gider Kaydet")
        self.btn_g_save.clicked.connect(self.gider_save)
        
        form.addRow("Dönem", self.g_in_donem)
        form.addRow("Tarih", self.g_dt_tarih)
        form.addRow("Kategori", self.g_cmb_kat)
        form.addRow("Tutar (TL)", self.g_in_tutar)
        form.addRow("Yöntem", self.g_cmb_yontem)
        form.addRow("Açıklama", self.g_in_acik)
        form.addRow("", self.btn_g_save)
        lay.addWidget(gb)
        
        self.tbl_gider = QTableWidget(0, 7)
        self.tbl_gider.setHorizontalHeaderLabels(["ID", "Dönem", "Tarih", "Kategori", "Tutar", "Yöntem", "Açıklama"])
        self.tbl_gider.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_gider.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_gider.doubleClicked.connect(self.open_expense_for_edit)
        self.tbl_gider.horizontalHeader().setStretchLastSection(True)
        self.tbl_gider.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_gider, 1)
        
        self.tabs.addTab(w, "Gider")
    
    def refresh_gider_table(self):
        donem = self.g_in_donem.text().strip()
        if not validate_period(donem):
            self.tbl_gider.setRowCount(0)
            return
        
        con = connect()
        rows = con.execute("""
            SELECT id, donem, tarih, kategori, tutar, yontem, aciklama
            FROM giderler
            WHERE donem=?
            ORDER BY id DESC
        """, (donem,)).fetchall()
        con.close()
        
        self.tbl_gider.setRowCount(0)
        for r in rows:
            row = self.tbl_gider.rowCount()
            self.tbl_gider.insertRow(row)
            for c, val in enumerate(r):
                it = QTableWidgetItem(str(val))
                if c == 0:
                    it.setTextAlignment(Qt.AlignCenter)
                if c == 4:
                    it.setText(f"{float(val):.2f}")
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tbl_gider.setItem(row, c, it)
    
    def gider_save(self):
        donem = self.g_in_donem.text().strip()
        if not validate_period(donem):
            QMessageBox.warning(self, "Uyarı", "Dönem formatı YYYY-MM olmalı.")
            return
        
        tarih = self.g_dt_tarih.date().toPython().isoformat()
        kat = self.g_cmb_kat.currentText()
        try:
            tutar = safe_float(self.g_in_tutar.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Tutar sayısal olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        yontem = self.g_cmb_yontem.currentText()
        acik = self.g_in_acik.text().strip()
        
        con = connect()
        con.execute("""
            INSERT INTO giderler(donem, tarih, kategori, tutar, yontem, aciklama)
            VALUES(?,?,?,?,?,?)
        """, (donem, tarih, kat, float(tutar), yontem, acik))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Gider kaydedildi.")
        self.g_in_tutar.clear()
        self.g_in_acik.clear()
        self.refresh_gider_table()
    
    def open_expense_for_edit(self, index):
        row = index.row()
        
        try:
            expense_id = int(self.tbl_gider.item(row, 0).text())
        except Exception:
            QMessageBox.warning(self, "Hata", "Gider ID'si alınamadı.")
            return
        
        dialog = EditExpenseDialog(expense_id, self)
        dialog.exec()
        
        self.refresh_gider_table()
        self.refresh_all()
    
    def _tab_duyuru(self):
        """Duyuru sekmesi"""
        w = QWidget()
        lay = QVBoxLayout(w)
        
        gb = QGroupBox("Apartman Duyurusu")
        form = QFormLayout(gb)
        
        self.duy_dt = QDateEdit()
        self.duy_dt.setCalendarPopup(True)
        self.duy_dt.setDisplayFormat("dd-MM-yyyy")
        self.duy_dt.setDate(QDate.currentDate())
        
        self.duy_baslik = QLineEdit()
        self.duy_mesaj = QTextEdit()
        
        self.btn_duy_save = QPushButton("Duyuru Kaydet")
        self.btn_duy_save.clicked.connect(self.duyuru_save)
        
        form.addRow("Tarih", self.duy_dt)
        form.addRow("Başlık", self.duy_baslik)
        form.addRow("Mesaj", self.duy_mesaj)
        form.addRow("", self.btn_duy_save)
        lay.addWidget(gb)
        
        self.tbl_duyuru = QTableWidget(0, 4)
        self.tbl_duyuru.setHorizontalHeaderLabels(["ID", "Tarih", "Başlık", "Mesaj"])
        self.tbl_duyuru.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_duyuru.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_duyuru.horizontalHeader().setStretchLastSection(True)
        self.tbl_duyuru.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_duyuru, 1)
        
        btns = QHBoxLayout()
        self.btn_duy_del = QPushButton("Seçileni Sil")
        self.btn_duy_del.clicked.connect(self.duyuru_delete_selected)
        btns.addWidget(self.btn_duy_del)
        btns.addStretch(1)
        lay.addLayout(btns)
        
        self.tabs.addTab(w, "Duyuru")
    
    def refresh_duyuru_table(self):
        con = connect()
        rows = con.execute("""
            SELECT id, tarih, baslik, mesaj
            FROM duyurular
            ORDER BY id DESC
        """).fetchall()
        con.close()
        
        self.tbl_duyuru.setRowCount(0)
        for did, t, b, m in rows:
            row = self.tbl_duyuru.rowCount()
            self.tbl_duyuru.insertRow(row)
            self.tbl_duyuru.setItem(row, 0, QTableWidgetItem(str(did)))
            self.tbl_duyuru.setItem(row, 1, QTableWidgetItem(str(t)))
            self.tbl_duyuru.setItem(row, 2, QTableWidgetItem(str(b)))
            short = (m[:80] + "…") if len(m) > 80 else m
            self.tbl_duyuru.setItem(row, 3, QTableWidgetItem(short))
    
    def duyuru_save(self):
        tarih = self.duy_dt.date().toPython().isoformat()
        baslik = self.duy_baslik.text().strip()
        mesaj = self.duy_mesaj.toPlainText().strip()
        
        if not baslik or not mesaj:
            QMessageBox.warning(self, "Uyarı", "Başlık ve mesaj zorunlu.")
            return
        
        con = connect()
        con.execute("INSERT INTO duyurular(tarih, baslik, mesaj) VALUES(?,?,?)",
                    (tarih, baslik, mesaj))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Duyuru kaydedildi.")
        self.duy_baslik.clear()
        self.duy_mesaj.clear()
        self.refresh_duyuru_table()
    
    def duyuru_delete_selected(self):
        rows = self.tbl_duyuru.selectionModel().selectedRows()
        if not rows:
            QMessageBox.warning(self, "Uyarı", "Silmek için tablodan bir duyuru seçin.")
            return
        
        r = rows[0].row()
        did = int(self.tbl_duyuru.item(r, 0).text())
        
        if QMessageBox.question(self, "Onay", "Seçili duyuru silinsin mi?") != QMessageBox.Yes:
            return
        
        con = connect()
        con.execute("DELETE FROM duyurular WHERE id=?", (did,))
        con.commit()
        con.close()
        
        self.refresh_duyuru_table()
    
    def _tab_rapor(self):
        """Rapor sekmesi"""
        w = QWidget()
        lay = QVBoxLayout(w)
        
        head = QHBoxLayout()
        lay.addLayout(head)
        
        head.addWidget(QLabel("Rapor Dönem:"))
        self.r_in_donem = QLineEdit(ym_of_today())
        self.r_in_donem.setMaximumWidth(110)
        head.addWidget(self.r_in_donem)
        
        self.r_chk_only_unpaid = QCheckBox("Sadece borçlu")
        self.r_chk_only_overdue = QCheckBox("Sadece geciken")
        head.addWidget(self.r_chk_only_unpaid)
        head.addWidget(self.r_chk_only_overdue)
        
        self.btn_r_calc = QPushButton("Hesapla/Yenile")
        self.btn_r_calc.clicked.connect(self.refresh_report)
        head.addWidget(self.btn_r_calc)
        
        head.addStretch(1)
        
        self.lbl_summary = QLabel("Özet: -")
        self.lbl_summary.setTextInteractionFlags(Qt.TextSelectableByMouse)
        lay.addWidget(self.lbl_summary)
        
        self.tbl_report = QTableWidget(0, 8)
        self.tbl_report.setHorizontalHeaderLabels([
            "Daire", "Ad", "Tel", "Aidat", "Borç", "Ödeme", "Bakiye", "Durum"
        ])
        self.tbl_report.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_report.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl_report.horizontalHeader().setStretchLastSection(True)
        self.tbl_report.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_report, 1)
        
        self.tabs.addTab(w, "Rapor")
    
    def refresh_report(self):
        donem = self.r_in_donem.text().strip()
        
        if not validate_period(donem):
            self.lbl_summary.setText("Özet: Dönem formatı yanlış.")
            self.tbl_report.setRowCount(0)
            return
        
        only_unpaid = self.r_chk_only_unpaid.isChecked()
        only_overdue = self.r_chk_only_overdue.isChecked()
        
        con = connect()
        daireler = con.execute("""
            SELECT id, daire_no, ad_soyad, telefon, aidat
            FROM daireler
            WHERE aktif = 1
            ORDER BY CAST(daire_no AS INTEGER), daire_no
        """).fetchall()
        con.close()
        
        self.tbl_report.setRowCount(0)
        
        toplam_tahakkuk = 0
        toplam_odeme = 0
        
        for did, dno, ad, tel, aidat in daireler:
            con = connect()
            
            row_tahakkuk = con.execute("""
                SELECT COALESCE(SUM(tutar), 0)
                FROM tahakkuk
                WHERE daire_id=? AND donem=?
            """, (did, donem)).fetchone()
            tahakkuk = float(row_tahakkuk[0]) if row_tahakkuk[0] else 0
            
            row_odeme = con.execute("""
                SELECT COALESCE(SUM(tutar), 0)
                FROM odemeler
                WHERE daire_id=? AND donem=?
            """, (did, donem)).fetchone()
            odeme = float(row_odeme[0]) if row_odeme[0] else 0
            
            con.close()
            
            bakiye = tahakkuk - odeme
            durum = status_for_balance(donem, bakiye)
            
            if only_unpaid and bakiye <= 0.00001:
                continue
            if only_overdue and durum != "Gecikmiş":
                continue
            
            row = self.tbl_report.rowCount()
            self.tbl_report.insertRow(row)
            
            vals = [
                dno,
                ad,
                tel or "",
                f"{float(aidat):.2f}",
                f"{tahakkuk:.2f}",
                f"{odeme:.2f}",
                f"{bakiye:.2f}",
                durum
            ]
            
            for col, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                
                if col in (3, 4, 5, 6):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                
                if bakiye > 0.00001:
                    item.setBackground(QColor(255, 200, 200))
                else:
                    item.setBackground(QColor(200, 255, 200))
                
                self.tbl_report.setItem(row, col, item)
            
            toplam_tahakkuk += tahakkuk
            toplam_odeme += odeme
        
        net = toplam_tahakkuk - toplam_odeme
        self.lbl_summary.setText(
            f"Dönem: {donem} | Tahakkuk: {toplam_tahakkuk:.2f} TL | "
            f"Ödeme: {toplam_odeme:.2f} TL | Bakiye: {net:.2f} TL"
        )
    
    def _tab_kasa(self):
        """YENI: Kasa Raporu Sekmesi"""
        w = QWidget()
        lay = QVBoxLayout(w)
        
        top = QHBoxLayout()
        top.addWidget(QLabel("Dönem:"))
        self.k_in_donem = QLineEdit(ym_of_today())
        self.k_in_donem.setMaximumWidth(110)
        top.addWidget(self.k_in_donem)
        
        self.btn_k_calc = QPushButton("Hesapla")
        self.btn_k_calc.clicked.connect(self.refresh_kasa)
        top.addWidget(self.btn_k_calc)
        
        top.addStretch(1)
        lay.addLayout(top)
        
        # Özet bilgileri
        self.lbl_k_summary = QLabel()
        self.lbl_k_summary.setStyleSheet("""
            background-color: #f0f0f0; 
            padding: 15px; 
            font-size: 13px; 
            border-radius: 5px;
            font-weight: bold;
        """)
        lay.addWidget(self.lbl_k_summary)
        
        # Gelir/Gider tablosu
        self.tbl_kasa = QTableWidget(0, 3)
        self.tbl_kasa.setHorizontalHeaderLabels(["Kategori", "Dönem", "Tutar (TL)"])
        self.tbl_kasa.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_kasa.horizontalHeader().setStretchLastSection(True)
        self.tbl_kasa.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_kasa, 1)
        
        self.tabs.addTab(w, "Kasa")
    
    def refresh_kasa(self):
        """Kasa raporunu hesapla"""
        donem = self.k_in_donem.text().strip()
        
        if not validate_period(donem):
            QMessageBox.warning(self, "Uyarı", "Dönem formatı YYYY-MM olmalı.")
            return
        
        con = connect()
        
        # Gelir = Tahakkuk
        tahakkuk_row = con.execute("""
            SELECT COALESCE(SUM(tutar), 0)
            FROM tahakkuk
            WHERE donem=?
        """, (donem,)).fetchone()
        tahakkuk_toplam = float(tahakkuk_row[0]) if tahakkuk_row[0] else 0
        
        # Ödeme geliri
        odeme_row = con.execute("""
            SELECT COALESCE(SUM(tutar), 0)
            FROM odemeler
            WHERE donem=?
        """, (donem,)).fetchone()
        odeme_toplam = float(odeme_row[0]) if odeme_row[0] else 0
        
        # Kategorili giderler
        gider_rows = con.execute("""
            SELECT kategori, COALESCE(SUM(tutar), 0)
            FROM giderler
            WHERE donem=?
            GROUP BY kategori
            ORDER BY kategori
        """, (donem,)).fetchall()
        
        toplam_gider = sum(float(r[1]) for r in gider_rows)
        
        con.close()
        
        # Tabloyu doldur
        self.tbl_kasa.setRowCount(0)
        
        # Gelir başlığı
        row = self.tbl_kasa.rowCount()
        self.tbl_kasa.insertRow(row)
        item = QTableWidgetItem("GELİRLER")
        item.setStyleSheet("background-color: #90EE90; font-weight: bold;")
        self.tbl_kasa.setItem(row, 0, item)
        
        # Tahakkuk
        row = self.tbl_kasa.rowCount()
        self.tbl_kasa.insertRow(row)
        self.tbl_kasa.setItem(row, 0, QTableWidgetItem("  Tahakkuk Geliri"))
        self.tbl_kasa.setItem(row, 1, QTableWidgetItem(donem))
        item = QTableWidgetItem(f"{tahakkuk_toplam:.2f}")
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.tbl_kasa.setItem(row, 2, item)
        
        # Ödeme geliri
        row = self.tbl_kasa.rowCount()
        self.tbl_kasa.insertRow(row)
        self.tbl_kasa.setItem(row, 0, QTableWidgetItem("  Ödeme Gelirleri"))
        self.tbl_kasa.setItem(row, 1, QTableWidgetItem(donem))
        item = QTableWidgetItem(f"{odeme_toplam:.2f}")
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.tbl_kasa.setItem(row, 2, item)
        
        # Gider başlığı
        row = self.tbl_kasa.rowCount()
        self.tbl_kasa.insertRow(row)
        item = QTableWidgetItem("GİDERLER")
        item.setStyleSheet("background-color: #FFB6C6; font-weight: bold;")
        self.tbl_kasa.setItem(row, 0, item)
        
        # Kategorili giderler
        for kat, tutar in gider_rows:
            row = self.tbl_kasa.rowCount()
            self.tbl_kasa.insertRow(row)
            self.tbl_kasa.setItem(row, 0, QTableWidgetItem(f"  {kat}"))
            self.tbl_kasa.setItem(row, 1, QTableWidgetItem(donem))
            item = QTableWidgetItem(f"{float(tutar):.2f}")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_kasa.setItem(row, 2, item)
        
        # Bakiye
        bakiye = tahakkuk_toplam + odeme_toplam - toplam_gider
        
        summary_text = (
            f"Dönem: {donem}\n"
            f"Tahakkuk Geliri: {tahakkuk_toplam:.2f} TL\n"
            f"Ödeme Gelirleri: {odeme_toplam:.2f} TL\n"
            f"TOPLAM GELİR: {tahakkuk_toplam + odeme_toplam:.2f} TL\n\n"
            f"TOPLAM GİDER: {toplam_gider:.2f} TL\n\n"
            f"KASA BAKIYE: {bakiye:.2f} TL"
        )
        
        if bakiye >= 0:
            self.lbl_k_summary.setStyleSheet("""
                background-color: #E8F5E9; 
                padding: 15px; 
                font-size: 13px; 
                border-radius: 5px;
                font-weight: bold;
                color: #2E7D32;
            """)
        else:
            self.lbl_k_summary.setStyleSheet("""
                background-color: #FFEBEE; 
                padding: 15px; 
                font-size: 13px; 
                border-radius: 5px;
                font-weight: bold;
                color: #C62828;
            """)
        
        self.lbl_k_summary.setText(summary_text)
    
    def _tab_ekstre(self):
        """YENI: Ekstre Sekmesi (Daire bazlı hareketler)"""
        w = QWidget()
        lay = QVBoxLayout(w)
        
        top = QHBoxLayout()
        
        top.addWidget(QLabel("Daire:"))
        self.e_cmb_daire = QComboBox()
        self.e_cmb_daire.currentIndexChanged.connect(self.refresh_ekstre)
        top.addWidget(self.e_cmb_daire)
        
        top.addWidget(QLabel("Başlangıç Dönem:"))
        self.e_in_from = QLineEdit(add_months(ym_of_today(), -12))
        self.e_in_from.setMaximumWidth(110)
        top.addWidget(self.e_in_from)
        
        top.addWidget(QLabel("Bitiş Dönem:"))
        self.e_in_to = QLineEdit(ym_of_today())
        self.e_in_to.setMaximumWidth(110)
        top.addWidget(self.e_in_to)
        
        btn_e_refresh = QPushButton("Yenile")
        btn_e_refresh.clicked.connect(self.refresh_ekstre)
        top.addWidget(btn_e_refresh)
        
        top.addStretch(1)
        lay.addLayout(top)
        
        # Özet
        self.lbl_e_ozet = QLabel("Özet: -")
        self.lbl_e_ozet.setStyleSheet("font-weight: bold; padding: 8px; background-color: #f0f0f0;")
        lay.addWidget(self.lbl_e_ozet)
        
        # Ekstre tablosu
        self.tbl_ekstre = QTableWidget(0, 6)
        self.tbl_ekstre.setHorizontalHeaderLabels(["Tarih", "Dönem", "Hareket Türü", "Açıklama", "Tutar", "Bakiye"])
        self.tbl_ekstre.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl_ekstre.horizontalHeader().setStretchLastSection(True)
        self.tbl_ekstre.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        lay.addWidget(self.tbl_ekstre, 1)
        
        self.tabs.addTab(w, "Ekstre")
    
    def refresh_ekstre_daire_combo(self):
        """Ekstre daire combo'sunu doldur"""
        con = connect()
        rows = con.execute("""
            SELECT id, daire_no, ad_soyad
            FROM daireler
            WHERE aktif = 1
            ORDER BY CAST(daire_no AS INT), daire_no
        """).fetchall()
        con.close()
        
        self.e_cmb_daire.blockSignals(True)
        self.e_cmb_daire.clear()
        
        for did, dno, ad in rows:
            self.e_cmb_daire.addItem(f"{dno} - {ad}", did)
        
        self.e_cmb_daire.blockSignals(False)
    
    def refresh_ekstre(self):
        """Ekstre tablosunu yükle"""
        daire_id = self.e_cmb_daire.currentData()
        
        if not daire_id:
            self.tbl_ekstre.setRowCount(0)
            self.lbl_e_ozet.setText("Özet: Daire seçin")
            return
        
        start = self.e_in_from.text().strip()
        end = self.e_in_to.text().strip()
        
        if not validate_period(start) or not validate_period(end):
            self.tbl_ekstre.setRowCount(0)
            self.lbl_e_ozet.setText("Özet: Dönem formatı YYYY-MM olmalı")
            return
        
        con = connect()
        
        # Dönemleri al
        donemler = months_range(start, end)
        
        self.tbl_ekstre.setRowCount(0)
        
        bakiye = 0.0
        toplam_tahakkuk = 0.0
        toplam_odeme = 0.0
        
        # Her dönem için tahakkuk ve ödemeleri göster
        for donem in donemler:
            row_tahakkuk = con.execute("""
                SELECT COALESCE(SUM(tutar), 0)
                FROM tahakkuk
                WHERE daire_id=? AND donem=?
            """, (daire_id, donem)).fetchone()
            tahakkuk = float(row_tahakkuk[0]) if row_tahakkuk[0] else 0
            
            row_odemeler = con.execute("""
                SELECT tarih, tutar
                FROM odemeler
                WHERE daire_id=? AND donem=?
                ORDER BY tarih
            """, (daire_id, donem)).fetchall()
            
            odeme_toplam_donem = sum(float(r[1]) for r in row_odemeler)
            
            # Tahakkuk satırı
            if tahakkuk > 0:
                bakiye += tahakkuk
                row = self.tbl_ekstre.rowCount()
                self.tbl_ekstre.insertRow(row)
                
                self.tbl_ekstre.setItem(row, 0, QTableWidgetItem(""))
                self.tbl_ekstre.setItem(row, 1, QTableWidgetItem(donem))
                
                item_type = QTableWidgetItem("Tahakkuk")
                item_type.setBackground(QColor(255, 240, 245))
                self.tbl_ekstre.setItem(row, 2, item_type)
                
                self.tbl_ekstre.setItem(row, 3, QTableWidgetItem("Aylık Aidat"))
                
                item_tutar = QTableWidgetItem(f"{tahakkuk:.2f}")
                item_tutar.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_tutar.setBackground(QColor(255, 240, 245))
                self.tbl_ekstre.setItem(row, 4, item_tutar)
                
                item_bakiye = QTableWidgetItem(f"{bakiye:.2f}")
                item_bakiye.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_bakiye.setBackground(QColor(255, 240, 245))
                self.tbl_ekstre.setItem(row, 5, item_bakiye)
                
                toplam_tahakkuk += tahakkuk
            
            # Ödeme satırları
            for tarih, tutar in row_odemeler:
                bakiye -= float(tutar)
                row = self.tbl_ekstre.rowCount()
                self.tbl_ekstre.insertRow(row)
                
                self.tbl_ekstre.setItem(row, 0, QTableWidgetItem(tarih))
                self.tbl_ekstre.setItem(row, 1, QTableWidgetItem(donem))
                
                item_type = QTableWidgetItem("Ödeme")
                item_type.setBackground(QColor(240, 255, 240))
                self.tbl_ekstre.setItem(row, 2, item_type)
                
                self.tbl_ekstre.setItem(row, 3, QTableWidgetItem("Ödeme"))
                
                item_tutar = QTableWidgetItem(f"-{float(tutar):.2f}")
                item_tutar.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_tutar.setBackground(QColor(240, 255, 240))
                self.tbl_ekstre.setItem(row, 4, item_tutar)
                
                item_bakiye = QTableWidgetItem(f"{bakiye:.2f}")
                item_bakiye.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_bakiye.setBackground(QColor(240, 255, 240))
                self.tbl_ekstre.setItem(row, 5, item_bakiye)
                
                toplam_odeme += float(tutar)
        
        con.close()
        
        # Özet
        ozet = (
            f"Dönem: {start} → {end} | "
            f"Tahakkuk: {toplam_tahakkuk:.2f} TL | "
            f"Ödeme: {toplam_odeme:.2f} TL | "
            f"Bakiye: {bakiye:.2f} TL"
        )
        self.lbl_e_ozet.setText(ozet)
    
    # ============ REFRESH & SETTINGS ============
    
    def save_settings(self):
        set_setting_int("vade_gun", int(self.sp_vade.value()))
        set_setting_int("gecikme_gun", int(self.sp_gec.value()))
        QMessageBox.information(self, "OK", "Ayarlar kaydedildi.")
        self.refresh_all()
    
    def refresh_all(self):
        d = self.in_donem.text().strip()
        
        if validate_period(d):
            if hasattr(self, "t_in_from") and not self.t_in_from.text().strip():
                self.t_in_from.setText(d)
            if hasattr(self, "t_in_to") and not self.t_in_to.text().strip():
                self.t_in_to.setText(d)
            if hasattr(self, "g_in_donem"):
                self.g_in_donem.setText(d)
            if hasattr(self, "r_in_donem"):
                self.r_in_donem.setText(d)
        
        self.refresh_daire_table()
        self.refresh_odeme_daire_combo()
        self.refresh_tahakkuk_table()
        self.refresh_gider_table()
        self.refresh_duyuru_table()
        self.refresh_report()
        self.refresh_odeme_gecmisi_daire_combo()
        self.refresh_odeme_gecmisi_table()
        self.refresh_ekstre_daire_combo()
        self.refresh_ekstre()


def main():
    app = QApplication(sys.argv)
    window = ApartmanAidatApp()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
