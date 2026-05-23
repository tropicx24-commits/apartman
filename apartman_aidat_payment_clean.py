# -*- coding: utf-8 -*-
"""
Temiz Ödeme Sistemi - v6 uygulaması için
Daire seçildiğinde borçlu dönemler otomatik listelenir
Checkbox ile seçilenleri öde, makbuz oluştur
"""

import sqlite3
from pathlib import Path
from datetime import date
import webbrowser
from urllib.parse import quote

from PySide6.QtCore import Qt, QDate
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout, QGroupBox,
    QComboBox, QTableWidget, QTableWidgetItem, QPushButton,
    QMessageBox, QDateEdit, QLineEdit, QCheckBox, QFileDialog, QLabel, QDialog
)
from PySide6.QtGui import QColor

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


DB_PATH = Path("apartman_aidat.db")


def connect():
    con = sqlite3.connect(str(DB_PATH))
    con.execute("PRAGMA foreign_keys = ON;")
    return con


def iso_today() -> str:
    return date.today().isoformat()


def safe_float(s: str) -> float:
    t = (s or "").strip()
    if not t:
        return 0.0
    t = t.replace(" ", "")
    t = t.replace(".", "")
    t = t.replace(",", ".")
    return float(t)


def validate_period(donem: str) -> bool:
    if not donem or len(donem) != 7 or donem[4] != "-":
        return False
    try:
        y = int(donem[:4])
        m = int(donem[5:])
        return 2000 <= y <= 2100 and 1 <= m <= 12
    except Exception:
        return False


def receipt_next_for_period(donem: str) -> str:
    """Makbuz numarasını oluştur (YYYYMM-00001 formatında)"""
    yyyymm = donem.replace("-", "")
    con = connect()
    row = con.execute("""
        SELECT makbuz_no FROM odemeler
        WHERE donem=?
        ORDER BY id DESC
        LIMIT 1
    """, (donem,)).fetchone()
    con.close()
    
    if not row:
        return f"{yyyymm}-00001"
    
    last = row[0]
    try:
        seq = int(last.split("-")[1])
    except Exception:
        seq = 0
    return f"{yyyymm}-{seq+1:05d}"


def autosize_worksheet(ws):
    """Excel sütunlarını otomatik genişlet"""
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            v = str(v)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)


class PaymentSystemClean(QWidget):
    """Temiz Ödeme Sistemi"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Temiz Ödeme Sistemi")
        self.setGeometry(100, 100, 1200, 800)
        self.init_ui()
    
    def init_ui(self):
        root = QVBoxLayout(self)
        
        # ============ DAIRE SEÇİMİ ============
        top_form = QFormLayout()
        
        self.cmb_daire = QComboBox()
        self.cmb_daire.currentIndexChanged.connect(self.on_daire_changed)
        
        top_form.addRow("Daire Seç:", self.cmb_daire)
        root.addLayout(top_form)
        
        # ============ BORÇLU DÖNEMLER LİSTESİ ============
        gb_borc = QGroupBox("Borçlu Dönemler (Checkbox ile Ödenecekleri Seç)")
        layout_borc = QVBoxLayout(gb_borc)
        
        self.tbl_borç = QTableWidget()
        self.tbl_borç.setColumnCount(4)
        self.tbl_borç.setHorizontalHeaderLabels(["Seç", "Dönem", "Borç (TL)", "Durum"])
        self.tbl_borç.setSelectionBehavior(QTableWidget.SelectRows)
        layout_borc.addWidget(self.tbl_borç)
        root.addWidget(gb_borc, 1)
        
        # ============ ÖDEME FORMU ============
        gb_form = QGroupBox("Ödeme Bilgileri")
        form_payment = QFormLayout(gb_form)
        
        self.dt_tarih = QDateEdit()
        self.dt_tarih.setCalendarPopup(True)
        self.dt_tarih.setDisplayFormat("dd-MM-yyyy")
        self.dt_tarih.setDate(QDate.currentDate())
        
        self.cmb_yontem = QComboBox()
        self.cmb_yontem.addItems(["Banka", "Elden"])
        
        self.in_acik = QLineEdit()
        self.in_acik.setPlaceholderText("İsteğe bağlı açıklama")
        
        # Özet
        self.lbl_ozet = QLabel("Seçili Dönemler: Yok | Toplam: 0.00 TL")
        self.lbl_ozet.setStyleSheet("font-weight: bold; color: #0066cc; padding: 8px;")
        
        form_payment.addRow("Ödeme Tarihi:", self.dt_tarih)
        form_payment.addRow("Ödeme Yöntemi:", self.cmb_yontem)
        form_payment.addRow("Açıklama:", self.in_acik)
        form_payment.addRow("", self.lbl_ozet)
        
        root.addWidget(gb_form)
        
        # ============ BUTONLAR ============
        btns = QHBoxLayout()
        
        btn_kaydet = QPushButton("✅ Ödemeyi Kaydet")
        btn_kaydet.clicked.connect(self.save_payment)
        btn_kaydet.setMinimumHeight(40)
        btn_kaydet.setStyleSheet("background-color: #28a745; color: white; font-weight: bold;")
        
        btn_temizle = QPushButton("🔄 Temizle")
        btn_temizle.clicked.connect(self.clear_selection)
        
        btns.addWidget(btn_kaydet, 2)
        btns.addWidget(btn_temizle, 1)
        btns.addStretch(1)
        root.addLayout(btns)
        
        self.setLayout(root)
        self.refresh_daireler()
    
    def refresh_daireler(self):
        """Daireleri yükle"""
        con = connect()
        rows = con.execute("""
            SELECT id, daire_no, ad_soyad
            FROM daireler
            WHERE aktif = 1
            ORDER BY CAST(daire_no AS INTEGER), daire_no
        """).fetchall()
        con.close()
        
        self.cmb_daire.blockSignals(True)
        self.cmb_daire.clear()
        self.cmb_daire.addItem("-- Daire Seç --", None)
        
        for did, dno, ad in rows:
            self.cmb_daire.addItem(f"Daire {dno} - {ad}", did)
        
        self.cmb_daire.blockSignals(False)
    
    def on_daire_changed(self):
        """Daire seçildiğinde borçlu dönemleri yükle"""
        daire_id = self.cmb_daire.currentData()
        
        if not daire_id:
            self.tbl_borç.setRowCount(0)
            self.update_summary()
            return
        
        self.load_borc_doneleri(int(daire_id))
        self.update_summary()
    
    def load_borc_doneleri(self, daire_id: int):
        """Dairenin borçlu dönemlerini yükle"""
        con = connect()
        
        # Tüm dönemleri al (tahakkuk + ödeme)
        rows = con.execute("""
            SELECT 
                COALESCE(t.donem, o.donem) as donem,
                COALESCE(SUM(t.tutar), 0) as tahakkuk,
                COALESCE(SUM(o.tutar), 0) as odeme
            FROM (
                SELECT DISTINCT donem FROM tahakkuk WHERE daire_id = ?
                UNION
                SELECT DISTINCT donem FROM odemeler WHERE daire_id = ?
            ) d
            LEFT JOIN tahakkuk t ON t.donem = d.donem AND t.daire_id = ?
            LEFT JOIN odemeler o ON o.donem = d.donem AND o.daire_id = ?
            GROUP BY COALESCE(t.donem, o.donem)
            HAVING (COALESCE(SUM(t.tutar), 0) - COALESCE(SUM(o.tutar), 0)) > 0.00001
            ORDER BY COALESCE(t.donem, o.donem) DESC
        """, (daire_id, daire_id, daire_id, daire_id)).fetchall()
        
        con.close()
        
        self.tbl_borç.setRowCount(0)
        
        for donem, tahakkuk, odeme in rows:
            kalan = float(tahakkuk) - float(odeme)
            
            row = self.tbl_borç.rowCount()
            self.tbl_borç.insertRow(row)
            
            # Checkbox
            chk = QCheckBox()
            self.tbl_borç.setCellWidget(row, 0, chk)
            
            # Dönem
            item_donem = QTableWidgetItem(str(donem))
            item_donem.setTextAlignment(Qt.AlignCenter)
            self.tbl_borç.setItem(row, 1, item_donem)
            
            # Borç
            item_tutar = QTableWidgetItem(f"{kalan:.2f}")
            item_tutar.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_borç.setItem(row, 2, item_tutar)
            
            # Durum
            if kalan > float(tahakkuk) * 0.5:
                durum = "Yüksek"
                color = QColor(255, 200, 200)
            else:
                durum = "Kısmi"
                color = QColor(255, 255, 150)
            
            item_durum = QTableWidgetItem(durum)
            item_durum.setBackground(color)
            self.tbl_borç.setItem(row, 3, item_durum)
            
            # Tüm hücreleri hafif arka plana ayarla
            for col in range(4):
                if col != 0:
                    self.tbl_borç.item(row, col).setBackground(QColor(245, 250, 255))
    
    def update_summary(self):
        """Seçili dönemlerin özetini göster"""
        toplam = 0.0
        secili_doneler = []
        
        for row in range(self.tbl_borç.rowCount()):
            chk = self.tbl_borç.cellWidget(row, 0)
            if chk and chk.isChecked():
                donem = self.tbl_borç.item(row, 1).text()
                tutar = float(self.tbl_borç.item(row, 2).text())
                toplam += tutar
                secili_doneler.append(donem)
        
        if secili_doneler:
            ozet_text = f"Seçili Dönemler: {', '.join(secili_doneler)} | Toplam: {toplam:.2f} TL"
            self.lbl_ozet.setStyleSheet("font-weight: bold; color: #28a745; padding: 8px; background-color: #f0f8f0;")
        else:
            ozet_text = "Seçili Dönemler: Yok | Toplam: 0.00 TL"
            self.lbl_ozet.setStyleSheet("font-weight: bold; color: #0066cc; padding: 8px;")
        
        self.lbl_ozet.setText(ozet_text)
    
    def clear_selection(self):
        """Tüm checkbox'ları temizle"""
        for row in range(self.tbl_borç.rowCount()):
            chk = self.tbl_borç.cellWidget(row, 0)
            if chk:
                chk.setChecked(False)
        
        self.update_summary()
    
    def save_payment(self):
        """Ödemeyi kaydet"""
        daire_id = self.cmb_daire.currentData()
        
        if not daire_id:
            QMessageBox.warning(self, "Uyarı", "Daire seçin.")
            return
        
        # Seçili dönemleri topla
        odenecek_doneler = []
        toplam_tutar = 0.0
        
        for row in range(self.tbl_borç.rowCount()):
            chk = self.tbl_borç.cellWidget(row, 0)
            if chk and chk.isChecked():
                donem = self.tbl_borç.item(row, 1).text()
                tutar = float(self.tbl_borç.item(row, 2).text())
                odenecek_doneler.append((donem, tutar))
                toplam_tutar += tutar
        
        if not odenecek_doneler:
            QMessageBox.warning(self, "Uyarı", "Ödeme yapılacak dönem seçin.")
            return
        
        tarih = self.dt_tarih.date().toPython().isoformat()
        yontem = self.cmb_yontem.currentText()
        acik = self.in_acik.text().strip()
        
        # Makbuz numarası (ilk seçili dönemden)
        ilk_donem = odenecek_doneler[0][0]
        makbuz_no = receipt_next_for_period(ilk_donem)
        
        # DB'ye kaydet
        con = connect()
        try:
            for donem, tutar in odenecek_doneler:
                con.execute("""
                    INSERT INTO odemeler(donem, tarih, daire_id, tutar, yontem, makbuz_no, aciklama)
                    VALUES(?, ?, ?, ?, ?, ?, ?)
                """, (donem, tarih, int(daire_id), float(tutar), yontem, makbuz_no, acik))
            
            con.commit()
            con.close()
            
            # Makbuz oluştur
            daire_info = self._get_daire_info(int(daire_id))
            pdf_path = self.create_receipt_pdf(
                makbuz_no=makbuz_no,
                tarih=tarih,
                daire_no=daire_info['daire_no'],
                ad=daire_info['ad'],
                yontem=yontem,
                odenecek_doneler=odenecek_doneler,
                toplam=toplam_tutar,
                aciklama=acik
            )
            
            # Onay mesajı
            msg = f"✅ Ödeme Kaydedildi!\n\nMakbuz No: {makbuz_no}\nToplam: {toplam_tutar:.2f} TL\n\nMakbuzı açmak ister misiniz?"
            reply = QMessageBox.question(
                self, "Ödeme Başarılı", msg,
                QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes and pdf_path:
                webbrowser.open(Path(pdf_path).resolve().as_uri())
            
            # Formu temizle ve yenile
            self.clear_selection()
            self.in_acik.clear()
            self.dt_tarih.setDate(QDate.currentDate())
            self.on_daire_changed()
            
        except Exception as e:
            con.close()
            QMessageBox.critical(self, "Hata", f"Ödeme kaydedilemedi:\n{str(e)}")
    
    def _get_daire_info(self, daire_id: int) -> dict:
        """Daire bilgilerini al"""
        con = connect()
        row = con.execute(
            "SELECT daire_no, ad_soyad, telefon FROM daireler WHERE id=?",
            (daire_id,)
        ).fetchone()
        con.close()
        
        if not row:
            return {'daire_no': '?', 'ad': '?', 'tel': ''}
        
        return {
            'daire_no': row[0],
            'ad': row[1],
            'tel': row[2]
        }
    
    def create_receipt_pdf(self, makbuz_no: str, tarih: str, daire_no: str, ad: str,
                          yontem: str, odenecek_doneler: list, toplam: float, aciklama: str) -> str:
        """Makbuz PDF oluştur"""
        
        out_dir = Path.cwd() / "makbuzlar"
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / f"makbuz_{makbuz_no}.pdf"
        
        c = canvas.Canvas(str(out_path), pagesize=A4)
        w, h = A4
        
        y = h - 50
        
        # Başlık
        c.setFont("Helvetica-Bold", 18)
        c.drawString(50, y, "AİDAT MAKBUZU")
        y -= 30
        
        # Üst bilgiler
        c.setFont("Helvetica", 11)
        c.drawString(50, y, f"Makbuz No: {makbuz_no}")
        c.drawRightString(w - 50, y, f"Tarih: {tarih}")
        y -= 20
        
        c.drawString(50, y, f"Daire: {daire_no}")
        c.drawRightString(w - 50, y, f"Ad Soyad: {ad}")
        y -= 20
        
        c.drawString(50, y, f"Ödeme Yöntemi: {yontem}")
        y -= 30
        
        # Ödenen dönemler başlığı
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "ÖDENEN DÖNEMLER")
        y -= 20
        
        # Tablo başlığı
        c.setFont("Helvetica-Bold", 10)
        c.drawString(50, y, "Dönem")
        c.drawRightString(200, y, "Tutar (TL)")
        y -= 15
        c.line(50, y, 250, y)
        y -= 15
        
        # Dönemler
        c.setFont("Helvetica", 10)
        for donem, tutar in odenecek_doneler:
            c.drawString(50, y, donem)
            c.drawRightString(200, y, f"{tutar:.2f}")
            y -= 15
        
        # Ayırıcı çizgi
        y -= 5
        c.line(50, y, 250, y)
        y -= 15
        
        # Toplam
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "TOPLAM")
        c.drawRightString(200, y, f"{toplam:.2f} TL")
        y -= 30
        
        # Açıklama
        if aciklama:
            c.setFont("Helvetica", 9)
            c.drawString(50, y, f"Not: {aciklama[:80]}")
            y -= 20
        
        # İmza alanı
        y -= 30
        c.setFont("Helvetica", 9)
        c.drawString(50, y, "Kasiyerin İmzası:")
        c.line(180, y - 2, 350, y - 2)
        
        c.save()
        return str(out_path)


# ============ TEST ============
if __name__ == "__main__":
    from PySide6.QtWidgets import QApplication
    import sys
    
    app = QApplication(sys.argv)
    window = PaymentSystemClean()
    window.show()
    sys.exit(app.exec())
