# -*- coding: utf-8 -*-
"""
Apartman Aidat Sistemi (Mini v4) + Devir (Sakin Geçmişi) - UI Revamp

Masaüstü PySide6 uygulaması.

Özellikler:
- Daire ekle/güncelle/aktif-pasif
- Sakin geçmişi: Devir (taşınma) + geçmiş liste
- Tahakkuk (dönem borç yaz) + manuel geçmiş borç
- Excel'den borç içe aktar (daire yoksa otomatik oluşturur)
- Ödeme al (Banka/Elden) + peşin çok ay paylaştırma + Makbuz PDF
- Gider ekle/sil
- Duyuru ekle/sil + panoya kopyala
- Rapor: gelir/gider/net + borçlu/gecikmiş filtre + Excel export
- WhatsApp: güncel sakinin telefonuna wa.me linki açar

Not: WhatsApp otomatik göndermez, sadece tarayıcıda mesaj taslağı açar.
"""

import sys
import sqlite3
from pathlib import Path
from datetime import date, timedelta
import webbrowser
from urllib.parse import quote


from PySide6.QtCore import Qt, QDate
from PySide6.QtWidgets import (
    QApplication, QWidget, QTabWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QMessageBox, QComboBox, QGroupBox, QFileDialog, QSpinBox, QCheckBox,
    QTextEdit, QDateEdit, QSplitter, QHeaderView, QScrollArea, QDialog
)
from PySide6.QtGui import QColor

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


DB_PATH = Path("apartman_aidat.db")


# ----------------- helpers -----------------
def connect():
    con = sqlite3.connect(str(DB_PATH))
    con.execute("PRAGMA foreign_keys = ON;")
    return con


def iso_today() -> str:
    return date.today().isoformat()


def ym_of_today() -> str:
    d = date.today()
    return f"{d.year:04d}-{d.month:02d}"


def safe_float(s: str) -> float:
    t = (s or "").strip()
    if not t:
        return 0.0
    t = t.replace(" ", "")
    # TR: 1.234,56 -> 1234.56
    t = t.replace(".", "")
    t = t.replace(",", ".")
    return float(t)


def validate_period(donem: str) -> bool:
    if not donem or len(donem) != 7 or donem[4] != "-":
        return False
    try:
        y = int(donem[:4])
        m = int(donem[5:])
        return 2000 <= y <= 2100 and 1 <= m <= 12
    except Exception:
        return False


def add_months(ym: str, add: int) -> str:
    y = int(ym[:4])
    m = int(ym[5:])
    m2 = m + add
    y += (m2 - 1) // 12
    m2 = ((m2 - 1) % 12) + 1
    return f"{y:04d}-{m2:02d}"


def months_range(ym_start: str, ym_end: str):
    ys = int(ym_start[:4]); ms = int(ym_start[5:])
    ye = int(ym_end[:4]); me = int(ym_end[5:])
    cur_y, cur_m = ys, ms
    out = []
    while (cur_y, cur_m) <= (ye, me):
        out.append(f"{cur_y:04d}-{cur_m:02d}")
        cur_m += 1
        if cur_m == 13:
            cur_m = 1
            cur_y += 1
    return out


def ensure_db():
    con = connect()
    con.executescript("""
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS daireler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        daire_no TEXT NOT NULL UNIQUE,
        ad_soyad TEXT NOT NULL,
        telefon TEXT DEFAULT '',
        aidat REAL NOT NULL DEFAULT 0,
        aktif INTEGER NOT NULL DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS sakinler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        daire_id INTEGER NOT NULL,
        ad_soyad TEXT NOT NULL,
        telefon TEXT DEFAULT '',
        baslangic_tarihi TEXT NOT NULL,
        bitis_tarihi TEXT,
        FOREIGN KEY(daire_id) REFERENCES daireler(id)
    );
    CREATE INDEX IF NOT EXISTS idx_sakinler_daire ON sakinler(daire_id);
    CREATE INDEX IF NOT EXISTS idx_sakinler_aktif ON sakinler(daire_id, bitis_tarihi);

    CREATE TABLE IF NOT EXISTS tahakkuk (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        donem TEXT NOT NULL,
        daire_id INTEGER NOT NULL,
        tutar REAL NOT NULL,
        created_at TEXT NOT NULL,
        UNIQUE(donem, daire_id),
        FOREIGN KEY(daire_id) REFERENCES daireler(id)
    );

    CREATE TABLE IF NOT EXISTS odemeler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        donem TEXT NOT NULL,
        tarih TEXT NOT NULL,
        daire_id INTEGER NOT NULL,
        tutar REAL NOT NULL,
        yontem TEXT NOT NULL,
        makbuz_no TEXT NOT NULL,
        aciklama TEXT DEFAULT '',
        FOREIGN KEY(daire_id) REFERENCES daireler(id)
    );

    CREATE INDEX IF NOT EXISTS idx_odemeler_donem ON odemeler(donem);
    CREATE INDEX IF NOT EXISTS idx_odemeler_daire ON odemeler(daire_id);

    CREATE TABLE IF NOT EXISTS giderler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        donem TEXT NOT NULL,
        tarih TEXT NOT NULL,
        kategori TEXT NOT NULL,
        tutar REAL NOT NULL,
        yontem TEXT NOT NULL,
        aciklama TEXT DEFAULT ''
    );
    CREATE INDEX IF NOT EXISTS idx_giderler_donem ON giderler(donem);

    CREATE TABLE IF NOT EXISTS duyurular (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tarih TEXT NOT NULL,
        baslik TEXT NOT NULL,
        mesaj TEXT NOT NULL
    );
    """)
    con.commit()

    def set_default(k, v):
        if con.execute("SELECT value FROM settings WHERE key=?", (k,)).fetchone() is None:
            con.execute("INSERT INTO settings(key,value) VALUES(?,?)", (k, str(v)))

    set_default("vade_gun", 10)
    set_default("gecikme_gun", 5)
    con.commit()
    con.close()


def get_setting_int(key: str, default: int) -> int:
    con = connect()
    row = con.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    con.close()
    try:
        return int(row[0]) if row else default
    except Exception:
        return default


def set_setting_int(key: str, value: int):
    con = connect()
    con.execute(
        "INSERT INTO settings(key,value) VALUES(?,?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, str(int(value)))
    )
    con.commit()
    con.close()


def receipt_next_for_period(donem: str) -> str:
    yyyymm = donem.replace("-", "")
    con = connect()
    row = con.execute("""
        SELECT makbuz_no FROM odemeler
        WHERE donem=?
        ORDER BY id DESC
        LIMIT 1
    """, (donem,)).fetchone()
    con.close()
    if not row:
        return f"{yyyymm}-00001"
    last = row[0]
    try:
        seq = int(last.split("-")[1])
    except Exception:
        seq = 0
    return f"{yyyymm}-{seq+1:05d}"


def compute_vade_date(donem: str) -> date:
    vade_gun = get_setting_int("vade_gun", 10)
    y, m = donem.split("-")
    y = int(y); m = int(m)
    # last day
    if m == 12:
        next_month = date(y + 1, 1, 1)
    else:
        next_month = date(y, m + 1, 1)
    last_day = next_month - timedelta(days=1)
    day = min(vade_gun, last_day.day)
    return date(y, m, day)


def status_for_balance(donem: str, bakiye: float) -> str:
    if bakiye <= 0.00001:
        return "Ödedi"
    vade = compute_vade_date(donem)
    gecikme = get_setting_int("gecikme_gun", 5)
    limit = vade + timedelta(days=gecikme)
    if date.today() > limit:
        return "Gecikmiş"
    return "Borçlu"


def autosize_worksheet(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            v = str(v)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)


# ============ MODAL DIALOGS ============

class EditPaymentDialog(QDialog):
    """Ödeme düzenleme modalı"""
    def __init__(self, payment_id: int, parent=None):
        super().__init__(parent)
        self.payment_id = payment_id
        self.setWindowTitle("Ödemeyi Düzenle")
        self.setGeometry(100, 100, 500, 300)
        self.setModal(True)
        self.init_ui()
        self.load_payment()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        form = QFormLayout()
        
        self.in_donem = QLineEdit()
        self.in_donem.setReadOnly(True)
        self.dt_tarih = QDateEdit()
        self.dt_tarih.setCalendarPopup(True)
        self.dt_tarih.setDisplayFormat("dd-MM-yyyy")
        self.in_daire = QLineEdit()
        self.in_daire.setReadOnly(True)
        self.in_tutar = QLineEdit()
        self.cmb_yontem = QComboBox()
        self.cmb_yontem.addItems(["Banka", "Elden"])
        self.in_makbuz = QLineEdit()
        self.in_acik = QLineEdit()
        
        form.addRow("Dönem", self.in_donem)
        form.addRow("Tarih", self.dt_tarih)
        form.addRow("Daire", self.in_daire)
        form.addRow("Tutar (TL)", self.in_tutar)
        form.addRow("Yöntem", self.cmb_yontem)
        form.addRow("Makbuz No", self.in_makbuz)
        form.addRow("Açıklama", self.in_acik)
        
        layout.addLayout(form)
        
        btns = QHBoxLayout()
        self.btn_save = QPushButton("Güncelle")
        self.btn_save.clicked.connect(self.save_payment)
        self.btn_delete = QPushButton("Sil")
        self.btn_delete.clicked.connect(self.delete_payment)
        self.btn_close = QPushButton("Kapat")
        self.btn_close.clicked.connect(self.close)
        
        btns.addWidget(self.btn_save)
        btns.addWidget(self.btn_delete)
        btns.addStretch(1)
        btns.addWidget(self.btn_close)
        
        layout.addLayout(btns)
    
    def load_payment(self):
        """Ödeme verilerini yükle"""
        con = connect()
        row = con.execute("""
            SELECT o.donem, o.tarih, d.daire_no, o.tutar, o.yontem, o.makbuz_no, o.aciklama
            FROM odemeler o
            JOIN daireler d ON d.id = o.daire_id
            WHERE o.id = ?
        """, (self.payment_id,)).fetchone()
        con.close()
        
        if not row:
            QMessageBox.warning(self, "Hata", "Ödeme bulunamadı.")
            self.close()
            return
        
        donem, tarih, daire_no, tutar, yontem, makbuz_no, acik = row
        
        self.in_donem.setText(donem)
        self.dt_tarih.setDate(QDate.fromString(tarih, "yyyy-MM-dd"))
        self.in_daire.setText(daire_no)
        self.in_tutar.setText(f"{float(tutar):.2f}")
        self.cmb_yontem.setCurrentText(yontem)
        self.in_makbuz.setText(makbuz_no or "")
        self.in_acik.setText(acik or "")
    
    def save_payment(self):
        """Ödemeyi güncelle"""
        try:
            tutar = safe_float(self.in_tutar.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Tutar sayısal olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        tarih = self.dt_tarih.date().toPython().isoformat()
        yontem = self.cmb_yontem.currentText()
        acik = self.in_acik.text().strip()
        
        con = connect()
        con.execute("""
            UPDATE odemeler
            SET tarih=?, tutar=?, yontem=?, aciklama=?
            WHERE id=?
        """, (tarih, float(tutar), yontem, acik, self.payment_id))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Ödeme güncellendi.")
        self.accept()
    
    def delete_payment(self):
        """Ödemeyi sil"""
        reply = QMessageBox.question(
            self, "Onay", "Seçili ödemeyi silmek istiyor musunuz?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        
        con = connect()
        con.execute("DELETE FROM odemeler WHERE id=?", (self.payment_id,))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Ödeme silindi.")
        self.accept()


class EditExpenseDialog(QDialog):
    """Gider düzenleme modalı"""
    def __init__(self, expense_id: int, parent=None):
        super().__init__(parent)
        self.expense_id = expense_id
        self.setWindowTitle("Gideri Düzenle")
        self.setGeometry(100, 100, 500, 300)
        self.setModal(True)
        self.init_ui()
        self.load_expense()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        form = QFormLayout()
        
        self.in_donem = QLineEdit()
        self.in_donem.setReadOnly(True)
        
        self.dt_tarih = QDateEdit()
        self.dt_tarih.setCalendarPopup(True)
        self.dt_tarih.setDisplayFormat("dd-MM-yyyy")
        
        self.cmb_kat = QComboBox()
        self.cmb_kat.addItems(["Elektrik", "Temizlik", "Bakım", "Diğer"])
        
        self.in_tutar = QLineEdit()
        
        self.cmb_yontem = QComboBox()
        self.cmb_yontem.addItems(["Banka", "Elden"])
        
        self.in_acik = QLineEdit()
        
        form.addRow("Dönem", self.in_donem)
        form.addRow("Tarih", self.dt_tarih)
        form.addRow("Kategori", self.cmb_kat)
        form.addRow("Tutar (TL)", self.in_tutar)
        form.addRow("Yöntem", self.cmb_yontem)
        form.addRow("Açıklama", self.in_acik)
        
        layout.addLayout(form)
        
        btns = QHBoxLayout()
        self.btn_save = QPushButton("Güncelle")
        self.btn_save.clicked.connect(self.save_expense)
        self.btn_delete = QPushButton("Sil")
        self.btn_delete.clicked.connect(self.delete_expense)
        self.btn_close = QPushButton("Kapat")
        self.btn_close.clicked.connect(self.close)
        
        btns.addWidget(self.btn_save)
        btns.addWidget(self.btn_delete)
        btns.addStretch(1)
        btns.addWidget(self.btn_close)
        
        layout.addLayout(btns)
    
    def load_expense(self):
        """Gider verilerini yükle"""
        con = connect()
        row = con.execute("""
            SELECT donem, tarih, kategori, tutar, yontem, aciklama
            FROM giderler
            WHERE id = ?
        """, (self.expense_id,)).fetchone()
        con.close()
        
        if not row:
            QMessageBox.warning(self, "Hata", "Gider bulunamadı.")
            self.close()
            return
        
        donem, tarih, kat, tutar, yontem, acik = row
        
        self.in_donem.setText(donem)
        self.dt_tarih.setDate(QDate.fromString(tarih, "yyyy-MM-dd"))
        self.cmb_kat.setCurrentText(kat)
        self.in_tutar.setText(f"{float(tutar):.2f}")
        self.cmb_yontem.setCurrentText(yontem)
        self.in_acik.setText(acik or "")
    
    def save_expense(self):
        """Gideri güncelle"""
        try:
            tutar = safe_float(self.in_tutar.text())
        except Exception:
            QMessageBox.warning(self, "Uyarı", "Tutar sayısal olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        tarih = self.dt_tarih.date().toPython().isoformat()
        kat = self.cmb_kat.currentText()
        yontem = self.cmb_yontem.currentText()
        acik = self.in_acik.text().strip()
        
        con = connect()
        con.execute("""
            UPDATE giderler
            SET tarih=?, kategori=?, tutar=?, yontem=?, aciklama=?
            WHERE id=?
        """, (tarih, kat, float(tutar), yontem, acik, self.expense_id))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Gider güncellendi.")
        self.accept()
    
    def delete_expense(self):
        """Gideri sil"""
        reply = QMessageBox.question(
            self, "Onay", "Seçili gideri silmek istiyor musunuz?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        
        con = connect()
        con.execute("DELETE FROM giderler WHERE id=?", (self.expense_id,))
        con.commit()
        con.close()
        
        QMessageBox.information(self, "OK", "Gider silindi.")
        self.accept()


# ============ MAIN WINDOW ============

class ApartmanApp(QWidget):
    """Ana uygulama penceresi"""
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Apartman Aidat Sistemi v7")
        self.setGeometry(100, 100, 1400, 900)
        self.init_db()
        self.init_ui()
    
    def init_db(self):
        ensure_db()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        self.tabs = QTabWidget()
        
        # Sekmeler
        self.tab_daireler = self.create_tab_daireler()
        self.tab_sakinler = self.create_tab_sakinler()
        self.tab_tahakkuk = self.create_tab_tahakkuk()
        self.tab_odemeler = self.create_tab_odemeler()
        self.tab_giderler = self.create_tab_giderler()
        self.tab_duyurular = self.create_tab_duyurular()
        self.tab_rapor = self.create_tab_rapor()
        self.tab_ayarlar = self.create_tab_ayarlar()
        
        self.tabs.addTab(self.tab_daireler, "Daireler")
        self.tabs.addTab(self.tab_sakinler, "Sakinler")
        self.tabs.addTab(self.tab_tahakkuk, "Tahakkuk")
        self.tabs.addTab(self.tab_odemeler, "Ödemeler")
        self.tabs.addTab(self.tab_giderler, "Giderler")
        self.tabs.addTab(self.tab_duyurular, "Duyurular")
        self.tabs.addTab(self.tab_rapor, "Rapor")
        self.tabs.addTab(self.tab_ayarlar, "Ayarlar")
        
        layout.addWidget(self.tabs)
        self.setLayout(layout)
    
    def create_tab_daireler(self):
        """Daire yönetimi sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Form
        form = QFormLayout()
        self.in_daire_no = QLineEdit()
        self.in_ad_soyad = QLineEdit()
        self.in_telefon = QLineEdit()
        self.in_aidat = QLineEdit()
        self.chk_aktif = QCheckBox("Aktif")
        self.chk_aktif.setChecked(True)
        
        form.addRow("Daire No", self.in_daire_no)
        form.addRow("Ad Soyad", self.in_ad_soyad)
        form.addRow("Telefon", self.in_telefon)
        form.addRow("Aidat (TL)", self.in_aidat)
        form.addRow("", self.chk_aktif)
        
        layout.addLayout(form)
        
        # Butonlar
        btns = QHBoxLayout()
        btn_ekle = QPushButton("Ekle")
        btn_ekle.clicked.connect(self.add_daire)
        btn_guncelle = QPushButton("Güncelle")
        btn_guncelle.clicked.connect(self.update_daire)
        btns.addWidget(btn_ekle)
        btns.addWidget(btn_guncelle)
        btns.addStretch()
        layout.addLayout(btns)
        
        # Tablo
        self.tbl_daireler = QTableWidget()
        self.tbl_daireler.setColumnCount(6)
        self.tbl_daireler.setHorizontalHeaderLabels(["Daire No", "Ad Soyad", "Telefon", "Aidat", "Aktif", "ID"])
        self.tbl_daireler.itemSelectionChanged.connect(self.on_daire_selected)
        layout.addWidget(self.tbl_daireler)
        
        widget.setLayout(layout)
        self.refresh_daireler()
        return widget
    
    def refresh_daireler(self):
        """Daire tablosunu yenile"""
        con = connect()
        rows = con.execute("SELECT id, daire_no, ad_soyad, telefon, aidat, aktif FROM daireler ORDER BY daire_no").fetchall()
        con.close()
        
        self.tbl_daireler.setRowCount(0)
        for row in rows:
            r = self.tbl_daireler.rowCount()
            self.tbl_daireler.insertRow(r)
            self.tbl_daireler.setItem(r, 0, QTableWidgetItem(str(row[1])))
            self.tbl_daireler.setItem(r, 1, QTableWidgetItem(str(row[2])))
            self.tbl_daireler.setItem(r, 2, QTableWidgetItem(str(row[3])))
            self.tbl_daireler.setItem(r, 3, QTableWidgetItem(f"{float(row[4]):.2f}"))
            self.tbl_daireler.setItem(r, 4, QTableWidgetItem("Evet" if row[5] else "Hayır"))
            self.tbl_daireler.setItem(r, 5, QTableWidgetItem(str(row[0])))
        
        self.tbl_daireler.hideColumn(5)
    
    def on_daire_selected(self):
        """Daire seçildiğinde form doldur"""
        row = self.tbl_daireler.currentRow()
        if row < 0:
            return
        
        self.in_daire_no.setText(self.tbl_daireler.item(row, 0).text())
        self.in_ad_soyad.setText(self.tbl_daireler.item(row, 1).text())
        self.in_telefon.setText(self.tbl_daireler.item(row, 2).text())
        self.in_aidat.setText(self.tbl_daireler.item(row, 3).text())
        aktif_text = self.tbl_daireler.item(row, 4).text()
        self.chk_aktif.setChecked(aktif_text == "Evet")
    
    def add_daire(self):
        """Yeni daire ekle"""
        daire_no = self.in_daire_no.text().strip()
        ad_soyad = self.in_ad_soyad.text().strip()
        telefon = self.in_telefon.text().strip()
        aidat = safe_float(self.in_aidat.text())
        aktif = 1 if self.chk_aktif.isChecked() else 0
        
        if not daire_no or not ad_soyad:
            QMessageBox.warning(self, "Uyarı", "Daire No ve Ad Soyad gerekli.")
            return
        
        try:
            con = connect()
            con.execute("""
                INSERT INTO daireler(daire_no, ad_soyad, telefon, aidat, aktif)
                VALUES(?, ?, ?, ?, ?)
            """, (daire_no, ad_soyad, telefon, float(aidat), aktif))
            con.commit()
            con.close()
            
            QMessageBox.information(self, "OK", "Daire eklendi.")
            self.in_daire_no.clear()
            self.in_ad_soyad.clear()
            self.in_telefon.clear()
            self.in_aidat.clear()
            self.chk_aktif.setChecked(True)
            self.refresh_daireler()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Uyarı", "Bu daire numarası zaten var.")
        except Exception as e:
            QMessageBox.warning(self, "Hata", str(e))
    
    def update_daire(self):
        """Seçili dairenin bilgilerini güncelle"""
        row = self.tbl_daireler.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Güncellenecek daire seçin.")
            return
        
        daire_id = int(self.tbl_daireler.item(row, 5).text())
        ad_soyad = self.in_ad_soyad.text().strip()
        telefon = self.in_telefon.text().strip()
        aidat = safe_float(self.in_aidat.text())
        aktif = 1 if self.chk_aktif.isChecked() else 0
        
        if not ad_soyad:
            QMessageBox.warning(self, "Uyarı", "Ad Soyad gerekli.")
            return
        
        try:
            con = connect()
            con.execute("""
                UPDATE daireler
                SET ad_soyad=?, telefon=?, aidat=?, aktif=?
                WHERE id=?
            """, (ad_soyad, telefon, float(aidat), aktif, daire_id))
            con.commit()
            con.close()
            
            QMessageBox.information(self, "OK", "Daire güncellendi.")
            self.refresh_daireler()
        except Exception as e:
            QMessageBox.warning(self, "Hata", str(e))
    
    def create_tab_sakinler(self):
        """Sakin yönetimi sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Form
        form = QFormLayout()
        self.cmb_sakin_daire = QComboBox()
        self.in_sakin_ad = QLineEdit()
        self.in_sakin_tel = QLineEdit()
        self.dt_baslangic = QDateEdit()
        self.dt_baslangic.setCalendarPopup(True)
        self.dt_baslangic.setDisplayFormat("yyyy-MM-dd")
        self.dt_baslangic.setDate(QDate.fromString(iso_today(), "yyyy-MM-dd"))
        self.dt_bitis = QDateEdit()
        self.dt_bitis.setCalendarPopup(True)
        self.dt_bitis.setDisplayFormat("yyyy-MM-dd")
        self.chk_aktif_sakin = QCheckBox("Hala oturuyor")
        self.chk_aktif_sakin.setChecked(True)
        
        form.addRow("Daire", self.cmb_sakin_daire)
        form.addRow("Ad Soyad", self.in_sakin_ad)
        form.addRow("Telefon", self.in_sakin_tel)
        form.addRow("Başlangıç Tarihi", self.dt_baslangic)
        form.addRow("Bitiş Tarihi", self.dt_bitis)
        form.addRow("", self.chk_aktif_sakin)
        
        layout.addLayout(form)
        
        # Butonlar
        btns = QHBoxLayout()
        btn_sakin_ekle = QPushButton("Ekle")
        btn_sakin_ekle.clicked.connect(self.add_sakin)
        btns.addWidget(btn_sakin_ekle)
        btns.addStretch()
        layout.addLayout(btns)
        
        # Tablo
        self.tbl_sakinler = QTableWidget()
        self.tbl_sakinler.setColumnCount(6)
        self.tbl_sakinler.setHorizontalHeaderLabels(["Daire", "Ad Soyad", "Telefon", "Başlangıç", "Bitiş", "ID"])
        layout.addWidget(self.tbl_sakinler)
        
        widget.setLayout(layout)
        self.refresh_sakinler()
        return widget
    
    def refresh_sakinler(self):
        """Sakin tablosunu yenile"""
        self.cmb_sakin_daire.clear()
        con = connect()
        daireler = con.execute("SELECT id, daire_no FROM daireler WHERE aktif=1 ORDER BY daire_no").fetchall()
        for d_id, d_no in daireler:
            self.cmb_sakin_daire.addItem(d_no, d_id)
        
        sakinler = con.execute("""
            SELECT s.id, d.daire_no, s.ad_soyad, s.telefon, s.baslangic_tarihi, s.bitis_tarihi
            FROM sakinler s
            JOIN daireler d ON d.id = s.daire_id
            ORDER BY d.daire_no, s.baslangic_tarihi DESC
        """).fetchall()
        con.close()
        
        self.tbl_sakinler.setRowCount(0)
        for row in sakinler:
            r = self.tbl_sakinler.rowCount()
            self.tbl_sakinler.insertRow(r)
            self.tbl_sakinler.setItem(r, 0, QTableWidgetItem(str(row[1])))
            self.tbl_sakinler.setItem(r, 1, QTableWidgetItem(str(row[2])))
            self.tbl_sakinler.setItem(r, 2, QTableWidgetItem(str(row[3])))
            self.tbl_sakinler.setItem(r, 3, QTableWidgetItem(str(row[4])))
            self.tbl_sakinler.setItem(r, 4, QTableWidgetItem(str(row[5]) if row[5] else ""))
            self.tbl_sakinler.setItem(r, 5, QTableWidgetItem(str(row[0])))
        
        self.tbl_sakinler.hideColumn(5)
    
    def add_sakin(self):
        """Yeni sakin ekle"""
        daire_id = self.cmb_sakin_daire.currentData()
        ad_soyad = self.in_sakin_ad.text().strip()
        telefon = self.in_sakin_tel.text().strip()
        baslangic = self.dt_baslangic.date().toPython().isoformat()
        
        if not ad_soyad:
            QMessageBox.warning(self, "Uyarı", "Ad Soyad gerekli.")
            return
        
        bitis = None
        if not self.chk_aktif_sakin.isChecked():
            bitis = self.dt_bitis.date().toPython().isoformat()
        
        try:
            con = connect()
            con.execute("""
                INSERT INTO sakinler(daire_id, ad_soyad, telefon, baslangic_tarihi, bitis_tarihi)
                VALUES(?, ?, ?, ?, ?)
            """, (daire_id, ad_soyad, telefon, baslangic, bitis))
            con.commit()
            con.close()
            
            QMessageBox.information(self, "OK", "Sakin eklendi.")
            self.in_sakin_ad.clear()
            self.in_sakin_tel.clear()
            self.dt_baslangic.setDate(QDate.fromString(iso_today(), "yyyy-MM-dd"))
            self.chk_aktif_sakin.setChecked(True)
            self.refresh_sakinler()
        except Exception as e:
            QMessageBox.warning(self, "Hata", str(e))
    
    def create_tab_tahakkuk(self):
        """Tahakkuk sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Form
        form = QFormLayout()
        self.in_tahakkuk_donem = QLineEdit()
        self.in_tahakkuk_donem.setPlaceholderText("YYYY-MM")
        self.in_tahakkuk_tutar = QLineEdit()
        self.in_tahakkuk_daire = QLineEdit()
        self.in_tahakkuk_daire.setPlaceholderText("Boş = tüm daireler")
        
        form.addRow("Dönem", self.in_tahakkuk_donem)
        form.addRow("Tutar (TL)", self.in_tahakkuk_tutar)
        form.addRow("Daire (isteğe bağlı)", self.in_tahakkuk_daire)
        
        layout.addLayout(form)
        
        # Butonlar
        btns = QHBoxLayout()
        btn_tahakkuk = QPushButton("Tahakkuk Yaz")
        btn_tahakkuk.clicked.connect(self.create_tahakkuk)
        btns.addWidget(btn_tahakkuk)
        btns.addStretch()
        layout.addLayout(btns)
        
        # Tablo
        self.tbl_tahakkuk = QTableWidget()
        self.tbl_tahakkuk.setColumnCount(5)
        self.tbl_tahakkuk.setHorizontalHeaderLabels(["Dönem", "Daire", "Tutar", "Tarih", "ID"])
        layout.addWidget(self.tbl_tahakkuk)
        
        widget.setLayout(layout)
        self.refresh_tahakkuk()
        return widget
    
    def refresh_tahakkuk(self):
        """Tahakkuk tablosunu yenile"""
        con = connect()
        rows = con.execute("""
            SELECT t.id, t.donem, d.daire_no, t.tutar, t.created_at
            FROM tahakkuk t
            JOIN daireler d ON d.id = t.daire_id
            ORDER BY t.donem DESC, d.daire_no
        """).fetchall()
        con.close()
        
        self.tbl_tahakkuk.setRowCount(0)
        for row in rows:
            r = self.tbl_tahakkuk.rowCount()
            self.tbl_tahakkuk.insertRow(r)
            self.tbl_tahakkuk.setItem(r, 0, QTableWidgetItem(str(row[1])))
            self.tbl_tahakkuk.setItem(r, 1, QTableWidgetItem(str(row[2])))
            self.tbl_tahakkuk.setItem(r, 2, QTableWidgetItem(f"{float(row[3]):.2f}"))
            self.tbl_tahakkuk.setItem(r, 3, QTableWidgetItem(str(row[4])))
            self.tbl_tahakkuk.setItem(r, 4, QTableWidgetItem(str(row[0])))
        
        self.tbl_tahakkuk.hideColumn(4)
    
    def create_tahakkuk(self):
        """Tahakkuk yaz"""
        donem = self.in_tahakkuk_donem.text().strip()
        tutar = safe_float(self.in_tahakkuk_tutar.text())
        daire_no = self.in_tahakkuk_daire.text().strip()
        
        if not donem or not validate_period(donem):
            QMessageBox.warning(self, "Uyarı", "Dönem YYYY-MM formatında olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        try:
            con = connect()
            
            if daire_no:
                daire = con.execute("SELECT id FROM daireler WHERE daire_no=? AND aktif=1", (daire_no,)).fetchone()
                if not daire:
                    QMessageBox.warning(self, "Uyarı", f"Aktif daire '{daire_no}' bulunamadı.")
                    con.close()
                    return
                
                daire_ids = [daire[0]]
            else:
                daire_ids = [row[0] for row in con.execute("SELECT id FROM daireler WHERE aktif=1").fetchall()]
            
            for daire_id in daire_ids:
                con.execute("""
                    INSERT OR IGNORE INTO tahakkuk(donem, daire_id, tutar, created_at)
                    VALUES(?, ?, ?, ?)
                """, (donem, daire_id, float(tutar), iso_today()))
            
            con.commit()
            con.close()
            
            QMessageBox.information(self, "OK", "Tahakkuk yazıldı.")
            self.in_tahakkuk_donem.clear()
            self.in_tahakkuk_tutar.clear()
            self.in_tahakkuk_daire.clear()
            self.refresh_tahakkuk()
        except Exception as e:
            QMessageBox.warning(self, "Hata", str(e))
    
    def create_tab_odemeler(self):
        """Ödeme sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Form
        form = QFormLayout()
        self.cmb_odeme_daire = QComboBox()
        self.in_odeme_donem = QLineEdit()
        self.in_odeme_donem.setPlaceholderText("YYYY-MM")
        self.dt_odeme_tarih = QDateEdit()
        self.dt_odeme_tarih.setCalendarPopup(True)
        self.dt_odeme_tarih.setDisplayFormat("yyyy-MM-dd")
        self.dt_odeme_tarih.setDate(QDate.fromString(iso_today(), "yyyy-MM-dd"))
        self.in_odeme_tutar = QLineEdit()
        self.cmb_odeme_yontem = QComboBox()
        self.cmb_odeme_yontem.addItems(["Banka", "Elden"])
        self.in_odeme_acik = QLineEdit()
        
        form.addRow("Daire", self.cmb_odeme_daire)
        form.addRow("Dönem", self.in_odeme_donem)
        form.addRow("Tarih", self.dt_odeme_tarih)
        form.addRow("Tutar (TL)", self.in_odeme_tutar)
        form.addRow("Yöntem", self.cmb_odeme_yontem)
        form.addRow("Açıklama", self.in_odeme_acik)
        
        layout.addLayout(form)
        
        # Butonlar
        btns = QHBoxLayout()
        btn_odeme = QPushButton("Ödeme Kaydet")
        btn_odeme.clicked.connect(self.add_odeme)
        btns.addWidget(btn_odeme)
        btns.addStretch()
        layout.addLayout(btns)
        
        # Tablo (tıklanabilir)
        self.tbl_odemeler = QTableWidget()
        self.tbl_odemeler.setColumnCount(7)
        self.tbl_odemeler.setHorizontalHeaderLabels(["Dönem", "Daire", "Tarih", "Tutar", "Yöntem", "Makbuz", "ID"])
        self.tbl_odemeler.itemSelectionChanged.connect(self.on_odeme_row_clicked)
        layout.addWidget(self.tbl_odemeler)
        
        widget.setLayout(layout)
        self.refresh_odemeler()
        return widget
    
    def refresh_odemeler(self):
        """Ödeme tablosunu yenile"""
        self.cmb_odeme_daire.clear()
        con = connect()
        daireler = con.execute("SELECT id, daire_no FROM daireler WHERE aktif=1 ORDER BY daire_no").fetchall()
        for d_id, d_no in daireler:
            self.cmb_odeme_daire.addItem(d_no, d_id)
        
        odemeler = con.execute("""
            SELECT o.id, o.donem, d.daire_no, o.tarih, o.tutar, o.yontem, o.makbuz_no
            FROM odemeler o
            JOIN daireler d ON d.id = o.daire_id
            ORDER BY o.donem DESC, d.daire_no
        """).fetchall()
        con.close()
        
        self.tbl_odemeler.setRowCount(0)
        for row in odemeler:
            r = self.tbl_odemeler.rowCount()
            self.tbl_odemeler.insertRow(r)
            self.tbl_odemeler.setItem(r, 0, QTableWidgetItem(str(row[1])))
            self.tbl_odemeler.setItem(r, 1, QTableWidgetItem(str(row[2])))
            self.tbl_odemeler.setItem(r, 2, QTableWidgetItem(str(row[3])))
            self.tbl_odemeler.setItem(r, 3, QTableWidgetItem(f"{float(row[4]):.2f}"))
            self.tbl_odemeler.setItem(r, 4, QTableWidgetItem(str(row[5])))
            self.tbl_odemeler.setItem(r, 5, QTableWidgetItem(str(row[6])))
            self.tbl_odemeler.setItem(r, 6, QTableWidgetItem(str(row[0])))
        
        self.tbl_odemeler.hideColumn(6)
    
    def on_odeme_row_clicked(self):
        """Ödeme satırı tıklandı - düzenleme modalını aç"""
        row = self.tbl_odemeler.currentRow()
        if row >= 0:
            payment_id = int(self.tbl_odemeler.item(row, 6).text())
            dialog = EditPaymentDialog(payment_id, self)
            if dialog.exec() == QDialog.Accepted:
                self.refresh_odemeler()
    
    def add_odeme(self):
        """Yeni ödeme ekle"""
        daire_id = self.cmb_odeme_daire.currentData()
        donem = self.in_odeme_donem.text().strip()
        tutar = safe_float(self.in_odeme_tutar.text())
        yontem = self.cmb_odeme_yontem.currentText()
        tarih = self.dt_odeme_tarih.date().toPython().isoformat()
        acik = self.in_odeme_acik.text().strip()
        
        if not donem or not validate_period(donem):
            QMessageBox.warning(self, "Uyarı", "Dönem YYYY-MM formatında olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        try:
            makbuz_no = receipt_next_for_period(donem)
            
            con = connect()
            con.execute("""
                INSERT INTO odemeler(donem, tarih, daire_id, tutar, yontem, makbuz_no, aciklama)
                VALUES(?, ?, ?, ?, ?, ?, ?)
            """, (donem, tarih, daire_id, float(tutar), yontem, makbuz_no, acik))
            con.commit()
            con.close()
            
            QMessageBox.information(self, "OK", f"Ödeme kaydedildi. Makbuz: {makbuz_no}")
            self.in_odeme_donem.clear()
            self.in_odeme_tutar.clear()
            self.in_odeme_acik.clear()
            self.dt_odeme_tarih.setDate(QDate.fromString(iso_today(), "yyyy-MM-dd"))
            self.refresh_odemeler()
        except Exception as e:
            QMessageBox.warning(self, "Hata", str(e))
    
    def create_tab_giderler(self):
        """Gider sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Form
        form = QFormLayout()
        self.in_gider_donem = QLineEdit()
        self.in_gider_donem.setPlaceholderText("YYYY-MM")
        self.dt_gider_tarih = QDateEdit()
        self.dt_gider_tarih.setCalendarPopup(True)
        self.dt_gider_tarih.setDisplayFormat("yyyy-MM-dd")
        self.dt_gider_tarih.setDate(QDate.fromString(iso_today(), "yyyy-MM-dd"))
        self.cmb_gider_kat = QComboBox()
        self.cmb_gider_kat.addItems(["Elektrik", "Temizlik", "Bakım", "Diğer"])
        self.in_gider_tutar = QLineEdit()
        self.cmb_gider_yontem = QComboBox()
        self.cmb_gider_yontem.addItems(["Banka", "Elden"])
        self.in_gider_acik = QLineEdit()
        
        form.addRow("Dönem", self.in_gider_donem)
        form.addRow("Tarih", self.dt_gider_tarih)
        form.addRow("Kategori", self.cmb_gider_kat)
        form.addRow("Tutar (TL)", self.in_gider_tutar)
        form.addRow("Yöntem", self.cmb_gider_yontem)
        form.addRow("Açıklama", self.in_gider_acik)
        
        layout.addLayout(form)
        
        # Butonlar
        btns = QHBoxLayout()
        btn_gider = QPushButton("Gider Ekle")
        btn_gider.clicked.connect(self.add_gider)
        btns.addWidget(btn_gider)
        btns.addStretch()
        layout.addLayout(btns)
        
        # Tablo
        self.tbl_giderler = QTableWidget()
        self.tbl_giderler.setColumnCount(7)
        self.tbl_giderler.setHorizontalHeaderLabels(["Dönem", "Tarih", "Kategori", "Tutar", "Yöntem", "Açıklama", "ID"])
        self.tbl_giderler.itemSelectionChanged.connect(self.on_gider_row_clicked)
        layout.addWidget(self.tbl_giderler)
        
        widget.setLayout(layout)
        self.refresh_giderler()
        return widget
    
    def refresh_giderler(self):
        """Gider tablosunu yenile"""
        con = connect()
        giderler = con.execute("""
            SELECT id, donem, tarih, kategori, tutar, yontem, aciklama
            FROM giderler
            ORDER BY donem DESC, tarih DESC
        """).fetchall()
        con.close()
        
        self.tbl_giderler.setRowCount(0)
        for row in giderler:
            r = self.tbl_giderler.rowCount()
            self.tbl_giderler.insertRow(r)
            self.tbl_giderler.setItem(r, 0, QTableWidgetItem(str(row[1])))
            self.tbl_giderler.setItem(r, 1, QTableWidgetItem(str(row[2])))
            self.tbl_giderler.setItem(r, 2, QTableWidgetItem(str(row[3])))
            self.tbl_giderler.setItem(r, 3, QTableWidgetItem(f"{float(row[4]):.2f}"))
            self.tbl_giderler.setItem(r, 4, QTableWidgetItem(str(row[5])))
            self.tbl_giderler.setItem(r, 5, QTableWidgetItem(str(row[6]) if row[6] else ""))
            self.tbl_giderler.setItem(r, 6, QTableWidgetItem(str(row[0])))
        
        self.tbl_giderler.hideColumn(6)
    
    def on_gider_row_clicked(self):
        """Gider satırı tıklandı - düzenleme modalını aç"""
        row = self.tbl_giderler.currentRow()
        if row >= 0:
            expense_id = int(self.tbl_giderler.item(row, 6).text())
            dialog = EditExpenseDialog(expense_id, self)
            if dialog.exec() == QDialog.Accepted:
                self.refresh_giderler()
    
    def add_gider(self):
        """Yeni gider ekle"""
        donem = self.in_gider_donem.text().strip()
        tarih = self.dt_gider_tarih.date().toPython().isoformat()
        kat = self.cmb_gider_kat.currentText()
        tutar = safe_float(self.in_gider_tutar.text())
        yontem = self.cmb_gider_yontem.currentText()
        acik = self.in_gider_acik.text().strip()
        
        if not donem or not validate_period(donem):
            QMessageBox.warning(self, "Uyarı", "Dönem YYYY-MM formatında olmalı.")
            return
        
        if tutar <= 0:
            QMessageBox.warning(self, "Uyarı", "Tutar 0'dan büyük olmalı.")
            return
        
        try:
            con = connect()
            con.execute("""
                INSERT INTO giderler(donem, tarih, kategori, tutar, yontem, aciklama)
                VALUES(?, ?, ?, ?, ?, ?)
            """, (donem, tarih, kat, float(tutar), yontem, acik))
            con.commit()
            con.close()
            
            QMessageBox.information(self, "OK", "Gider eklendi.")
            self.in_gider_donem.clear()
            self.in_gider_tutar.clear()
            self.in_gider_acik.clear()
            self.dt_gider_tarih.setDate(QDate.fromString(iso_today(), "yyyy-MM-dd"))
            self.refresh_giderler()
        except Exception as e:
            QMessageBox.warning(self, "Hata", str(e))
    
    def create_tab_duyurular(self):
        """Duyuru sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Form
        form = QFormLayout()
        self.in_duyuru_baslik = QLineEdit()
        self.txt_duyuru_mesaj = QTextEdit()
        
        form.addRow("Başlık", self.in_duyuru_baslik)
        form.addRow("Mesaj", self.txt_duyuru_mesaj)
        
        layout.addLayout(form)
        
        # Butonlar
        btns = QHBoxLayout()
        btn_duyuru = QPushButton("Duyuru Ekle")
        btn_duyuru.clicked.connect(self.add_duyuru)
        btns.addWidget(btn_duyuru)
        btns.addStretch()
        layout.addLayout(btns)
        
        # Tablo
        self.tbl_duyurular = QTableWidget()
        self.tbl_duyurular.setColumnCount(4)
        self.tbl_duyurular.setHorizontalHeaderLabels(["Tarih", "Başlık", "Mesaj", "ID"])
        layout.addWidget(self.tbl_duyurular)
        
        widget.setLayout(layout)
        self.refresh_duyurular()
        return widget
    
    def refresh_duyurular(self):
        """Duyuru tablosunu yenile"""
        con = connect()
        duyurular = con.execute("""
            SELECT id, tarih, baslik, mesaj
            FROM duyurular
            ORDER BY tarih DESC
        """).fetchall()
        con.close()
        
        self.tbl_duyurular.setRowCount(0)
        for row in duyurular:
            r = self.tbl_duyurular.rowCount()
            self.tbl_duyurular.insertRow(r)
            self.tbl_duyurular.setItem(r, 0, QTableWidgetItem(str(row[1])))
            self.tbl_duyurular.setItem(r, 1, QTableWidgetItem(str(row[2])))
            self.tbl_duyurular.setItem(r, 2, QTableWidgetItem(str(row[3])[:50]))
            self.tbl_duyurular.setItem(r, 3, QTableWidgetItem(str(row[0])))
        
        self.tbl_duyurular.hideColumn(3)
    
    def add_duyuru(self):
        """Yeni duyuru ekle"""
        baslik = self.in_duyuru_baslik.text().strip()
        mesaj = self.txt_duyuru_mesaj.toPlainText().strip()
        
        if not baslik or not mesaj:
            QMessageBox.warning(self, "Uyarı", "Başlık ve Mesaj gerekli.")
            return
        
        try:
            con = connect()
            con.execute("""
                INSERT INTO duyurular(tarih, baslik, mesaj)
                VALUES(?, ?, ?)
            """, (iso_today(), baslik, mesaj))
            con.commit()
            con.close()
            
            QMessageBox.information(self, "OK", "Duyuru eklendi.")
            self.in_duyuru_baslik.clear()
            self.txt_duyuru_mesaj.clear()
            self.refresh_duyurular()
        except Exception as e:
            QMessageBox.warning(self, "Hata", str(e))
    
    def create_tab_rapor(self):
        """Rapor sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Filtreler
        form = QFormLayout()
        self.in_rapor_donem_start = QLineEdit()
        self.in_rapor_donem_start.setPlaceholderText("YYYY-MM")
        self.in_rapor_donem_end = QLineEdit()
        self.in_rapor_donem_end.setPlaceholderText("YYYY-MM")
        self.cmb_rapor_filtre = QComboBox()
        self.cmb_rapor_filtre.addItems(["Tümü", "Borçlu", "Gecikmiş", "Ödedi"])
        
        form.addRow("Dönem Başlangıcı", self.in_rapor_donem_start)
        form.addRow("Dönem Bitişi", self.in_rapor_donem_end)
        form.addRow("Filtre", self.cmb_rapor_filtre)
        
        layout.addLayout(form)
        
        # Butonlar
        btns = QHBoxLayout()
        btn_rapor = QPushButton("Rapor Oluştur")
        btn_rapor.clicked.connect(self.generate_report)
        btn_excel = QPushButton("Excel'e Aktar")
        btn_excel.clicked.connect(self.export_to_excel)
        btns.addWidget(btn_rapor)
        btns.addWidget(btn_excel)
        btns.addStretch()
        layout.addLayout(btns)
        
        # Tablo
        self.tbl_rapor = QTableWidget()
        self.tbl_rapor.setColumnCount(8)
        self.tbl_rapor.setHorizontalHeaderLabels(["Daire", "Dönem", "Tahakkuk", "Ödeme", "Bakiye", "Durum", "Sakin Telefon", "İşlem"])
        layout.addWidget(self.tbl_rapor)
        
        # Özet
        self.lbl_rapor_ozet = QLabel()
        layout.addWidget(self.lbl_rapor_ozet)
        
        widget.setLayout(layout)
        return widget
    
    def generate_report(self):
        """Rapor oluştur"""
        donem_start = self.in_rapor_donem_start.text().strip()
        donem_end = self.in_rapor_donem_end.text().strip()
        filtre = self.cmb_rapor_filtre.currentText()
        
        if not donem_start or not donem_end:
            QMessageBox.warning(self, "Uyarı", "Dönem aralığı belirtin.")
            return
        
        if not validate_period(donem_start) or not validate_period(donem_end):
            QMessageBox.warning(self, "Uyarı", "Dönem YYYY-MM formatında olmalı.")
            return
        
        try:
            con = connect()
            
            # Dönemi listele
            doneler = months_range(donem_start, donem_end)
            
            # Daireler
            daireler = con.execute("SELECT id, daire_no FROM daireler WHERE aktif=1 ORDER BY daire_no").fetchall()
            
            rapor_data = []
            
            for daire_id, daire_no in daireler:
                for donem in doneler:
                    tahakkuk = con.execute(
                        "SELECT tutar FROM tahakkuk WHERE daire_id=? AND donem=?",
                        (daire_id, donem)
                    ).fetchone()
                    tahakkuk_tutar = float(tahakkuk[0]) if tahakkuk else 0
                    
                    odemeler = con.execute(
                        "SELECT SUM(tutar) FROM odemeler WHERE daire_id=? AND donem=?",
                        (daire_id, donem)
                    ).fetchone()
                    odeme_tutar = float(odemeler[0]) if odemeler[0] else 0
                    
                    bakiye = tahakkuk_tutar - odeme_tutar
                    durum = status_for_balance(donem, bakiye)
                    
                    if filtre != "Tümü" and durum != filtre:
                        continue
                    
                    # Güncel sakin bilgisi
                    sakin = con.execute("""
                        SELECT ad_soyad, telefon
                        FROM sakinler
                        WHERE daire_id=? AND (bitis_tarihi IS NULL OR bitis_tarihi > ?)
                        ORDER BY baslangic_tarihi DESC
                        LIMIT 1
                    """, (daire_id, donem + "-01")).fetchone()
                    
                    sakin_tel = sakin[1] if sakin else ""
                    
                    rapor_data.append({
                        'daire': daire_no,
                        'donem': donem,
                        'tahakkuk': tahakkuk_tutar,
                        'odeme': odeme_tutar,
                        'bakiye': bakiye,
                        'durum': durum,
                        'tel': sakin_tel,
                        'daire_id': daire_id
                    })
            
            con.close()
            
            # Tablo doldur
            self.tbl_rapor.setRowCount(0)
            toplam_tahakkuk = 0
            toplam_odeme = 0
            toplam_bakiye = 0
            
            for item in rapor_data:
                r = self.tbl_rapor.rowCount()
                self.tbl_rapor.insertRow(r)
                
                self.tbl_rapor.setItem(r, 0, QTableWidgetItem(item['daire']))
                self.tbl_rapor.setItem(r, 1, QTableWidgetItem(item['donem']))
                self.tbl_rapor.setItem(r, 2, QTableWidgetItem(f"{item['tahakkuk']:.2f}"))
                self.tbl_rapor.setItem(r, 3, QTableWidgetItem(f"{item['odeme']:.2f}"))
                self.tbl_rapor.setItem(r, 4, QTableWidgetItem(f"{item['bakiye']:.2f}"))
                
                # Durum rengi
                durum_item = QTableWidgetItem(item['durum'])
                if item['durum'] == "Ödedi":
                    durum_item.setBackground(QColor(144, 238, 144))
                elif item['durum'] == "Borçlu":
                    durum_item.setBackground(QColor(255, 255, 153))
                elif item['durum'] == "Gecikmiş":
                    durum_item.setBackground(QColor(255, 153, 153))
                self.tbl_rapor.setItem(r, 5, durum_item)
                
                self.tbl_rapor.setItem(r, 6, QTableWidgetItem(item['tel']))
                
                # WhatsApp butonu
                btn_wa = QPushButton("📱 WhatsApp")
                btn_wa.clicked.connect(lambda checked, t=item['tel']: self.open_whatsapp(t))
                self.tbl_rapor.setCellWidget(r, 7, btn_wa)
                
                toplam_tahakkuk += item['tahakkuk']
                toplam_odeme += item['odeme']
                toplam_bakiye += item['bakiye']
            
            ozet = f"Toplam Tahakkuk: {toplam_tahakkuk:.2f} TL | Toplam Ödeme: {toplam_odeme:.2f} TL | Toplam Bakiye: {toplam_bakiye:.2f} TL"
            self.lbl_rapor_ozet.setText(ozet)
            
        except Exception as e:
            QMessageBox.warning(self, "Hata", str(e))
    
    def open_whatsapp(self, telefon: str):
        """WhatsApp'ı aç"""
        if not telefon:
            QMessageBox.warning(self, "Uyarı", "Telefon numarası yok.")
            return
        
        # Türkiye numarası için +90 ekle
        if telefon.startswith("0"):
            telefon = "+90" + telefon[1:]
        elif not telefon.startswith("+"):
            telefon = "+90" + telefon
        
        url = f"https://wa.me/{telefon}"
        webbrowser.open(url)
    
    def export_to_excel(self):
        """Raporu Excel'e aktar"""
        donem_start = self.in_rapor_donem_start.text().strip()
        donem_end = self.in_rapor_donem_end.text().strip()
        
        if not donem_start or not donem_end:
            QMessageBox.warning(self, "Uyarı", "Dönem aralığı belirtin.")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "Raporu Kaydet", f"rapor_{donem_start}_{donem_end}.xlsx", "Excel (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Rapor"
            
            # Başlık
            headers = ["Daire", "Dönem", "Tahakkuk", "Ödeme", "Bakiye", "Durum", "Sakin Telefon"]
            ws.append(headers)
            
            # Veriler
            for row in range(self.tbl_rapor.rowCount()):
                rowdata = []
                for col in range(7):
                    item = self.tbl_rapor.item(row, col)
                    rowdata.append(item.text() if item else "")
                ws.append(rowdata)
            
            autosize_worksheet(ws)
            wb.save(filename)
            QMessageBox.information(self, "OK", f"Rapor kaydedildi: {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Hata", str(e))
    
    def create_tab_ayarlar(self):
        """Ayarlar sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        form = QFormLayout()
        
        self.sp_vade_gun = QSpinBox()
        self.sp_vade_gun.setMinimum(1)
        self.sp_vade_gun.setMaximum(31)
        self.sp_vade_gun.setValue(get_setting_int("vade_gun", 10))
        
        self.sp_gecikme_gun = QSpinBox()
        self.sp_gecikme_gun.setMinimum(1)
        self.sp_gecikme_gun.setMaximum(31)
        self.sp_gecikme_gun.setValue(get_setting_int("gecikme_gun", 5))
        
        form.addRow("Vade (Gün)", self.sp_vade_gun)
        form.addRow("Gecikme (Gün)", self.sp_gecikme_gun)
        
        layout.addLayout(form)
        
        btn_kaydet = QPushButton("Ayarları Kaydet")
        btn_kaydet.clicked.connect(self.save_settings)
        layout.addWidget(btn_kaydet)
        
        layout.addStretch()
        
        widget.setLayout(layout)
        return widget
    
    def save_settings(self):
        """Ayarları kaydet"""
        vade_gun = self.sp_vade_gun.value()
        gecikme_gun = self.sp_gecikme_gun.value()
        
        set_setting_int("vade_gun", vade_gun)
        set_setting_int("gecikme_gun", gecikme_gun)
        
        QMessageBox.information(self, "OK", "Ayarlar kaydedildi.")


def main():
    app = QApplication(sys.argv)
    window = ApartmanApp()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
